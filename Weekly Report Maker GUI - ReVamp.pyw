#               Installing Pips Guide
#---------------------------------------------------------
#
#
#       Run the CIS PythonStartup Batch File
#
#   Last Update: 8/18/2023
#   Program Task: This program compiles excel sheets to
#                   generate the registration and outcomes report.
#                   Additionally, this only uses Openpyxl, win32com, and Ensurezip. 
#   
#
#                   ¯\_(ツ)_/¯
#
#                 Server Version
#---------------------------------------------------------

# Excel-Python Program for MSTs only (Openpyxl, Win32com, Ensurezip, and Pandas)
# Developed by CISHOT Data Entry Team
#
#   Day 2 Log - Chapter 2
#       The Second Start
#
#
#   Hello, Data person!
#   This is a program made for the CISNAV Thursday Weekly Registration and Close out report. All you need to do is put the file in Contact Logs folder then run the program using the file's name.
#   It is super simple~!
#
#   You can run the program just by double clicking or using IDLE python launcher. Next you will enter the file location that are prompted and it should open the file and begin its job!
#   'S:\\Programs\\Communities In Schools\\'+year+'\\6. Program Director Tools\\2023 Weekly Report Log\\' This is the path the program takes.
#
#See you soon,
#Data Dev.
#
#   Disclaimer: You will need to hand enter all information since the main excel is a living document.
#                       
#
# Required Information ------

from openpyxl import load_workbook
import openpyxl as opxl
import win32com.client as win32
import pywintypes
import os
from pywintypes import com_error
from win32com.client import constants
import pandas as pd
import numpy as np
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from mttkinter import *
import time
import threading

class backgroundInfo():
    path = ''
    previousTab = 0
    filename = ''
    def browse_button_One(eventOne = None):
        backgroundInfo.filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        pathEntry.delete(0, END)
        pathEntry.insert(END, backgroundInfo.filename)
        backgroundInfo.path = backgroundInfo.filename

    def recallPath():
        pathEntry.delete(0, END)
        pathEntry.insert(END, backgroundInfo.path)

    def closeNameWindow(eventClose = None):
            backgroundInfo.path = pathEntry.get()
            if(backgroundInfo.path != ""):
                backgroundInfo.previousTab = 1
                root.destroy()

    def close_window(eventLeave = None):
        exit();

    def previous_Button(eventBack = None):
        backgroundInfo.previousTab = 0
        loading.destroy()

def dispatch(app_name:str):
    try:
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    except AttributeError:
        # Corner case dependencies.
        import os
        import re
        import sys
        import shutil
        # Remove cache and try again.
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    return app

#filechange = ""
    #Phase 1 - Decisions
#while True:
#    file = input('====>Enter File Name<====\n')
#    filechange = file + '.xlsx'
#    year = input('What year is the file in? (Example: 2022-2023)\n')
#    proceed = input('Are you ready to continue? Y/N\n')
#    if(proceed == 'Y' or proceed == 'Yes' or proceed == 'yes' or proceed == 'YES' or proceed == 'Y' or proceed == 'y'):
#        break

    #Renames the first sheet incase of no edits.
#dest = 'S:\\Programs\\Communities In Schools\\'+year+'\\6. Program Director Tools\\2023 Weekly Report Log\\'
class programDetails():
    Status = ''
    endProgram = False
    def program(eventRun = None):
        file = os.path.split(backgroundInfo.path)[1]
        file = file.replace(".xlsx","")

        programDetails.Status = "Creating Pivot Document..." 
        files = backgroundInfo.path #+ filechange
        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(files)
        Services = target.Worksheets(1)
        try:
            Services.Name = file
        except:
            programDetails.endProgram = True
            programDetails.Status = "Program Complete!"
            target.Close(SaveChanges=True)
            xlApp.Quit()
            return
        target.Close(SaveChanges=True)
        xlApp.Quit()
                #Creates the two Registration sheets
        wb = load_workbook(filename = files)
        ws = wb.active
        wb.create_sheet("Reg_Time_Pivot_Table",0)
        wb.create_sheet("Reg_Status_Pivot_Table",0)
        wb.save(filename = files)
        wb.close()
                #Creates the pivot table for Reg_Time_Pivot_Table
        def clear_pts(wsw):
            for pt in wsw.PivotTables():
                pt.TableRange2.Clear()

        def insert_pt_field_set1(pt):
            field_rows = {}
            field_rows['Campus'] = pt.PivotFields("Campus")
            field_rows['CasefileProviders'] = pt.PivotFields("CasefileProviders")
            
            field_values = {}
            field_values['CaseFileStatus'] = pt.PivotFields("CaseFileStatus")
            field_values['ContractMetDate'] = pt.PivotFields("ContractMetDate")
            field_values['ServiceMinutes'] = pt.PivotFields("ServiceMinutes")

            field_columns = {}
            field_columns['values'] = pt.PivotFields("Values")


                    # Insert row fields
                    # https://docs.microsoft.com/en-us/office/vba/api/excel.xlpivotfieldorientation
                    
            field_rows['Campus'].Orientation = 1
            field_rows['Campus'].Position = 1

            field_rows['CasefileProviders'].Orientation = 1
            field_rows['CasefileProviders'].Position = 2


                    # Insert values field
                    # https://docs.microsoft.com/en-us/office/vba/api/excel.xlconsolidationfunction

            field_values['CaseFileStatus'].Orientation = 4
            field_values['CaseFileStatus'].Function = -4112 # value reference
            field_values['CaseFileStatus'].NumberFormat = "#,##0"

            field_values['ContractMetDate'].Orientation = 4
            field_values['ContractMetDate'].Function = -4112 # value reference
            field_values['ContractMetDate'].NumberFormat = "#,##0"

            field_values['ServiceMinutes'].Orientation = 4
            field_values['ServiceMinutes'].Function = -4157 # value reference
            field_values['ServiceMinutes'].NumberFormat = "#,##0"

            field_columns['values'].Orientation = 2
            field_columns['values'].Position = 1


                # Rquired to open excel in Win32Com
        xlApp = win32.Dispatch('Excel.Application')
        xlApp.Visible = False

                # Change the Path to Correct Folder if file is changed
        #dest = 'S:\\Programs\\Communities In Schools\\'+year+'\\6. Program Director Tools\\2023 Weekly Report Log\\'
        finalDest = backgroundInfo.path #+ filechange
            
                # Sheet Location and Name
        wbw = xlApp.Workbooks.Open(files)
        wsw_data = wbw.Worksheets(file)
        wsw_pivot = wbw.Worksheets("Reg_Time_Pivot_Table")


                # clear pivot tables on Report tab
        clear_pts(wsw_pivot)

                # create pt cache connection
        pt_cache = wbw.PivotCaches().Create(1, wsw_data.Range("A1").CurrentRegion)

                # insert pivot table designer/editor
        pt = pt_cache.CreatePivotTable(wsw_pivot.Range("B3"), "Service")

        insert_pt_field_set1(pt)

        wbw.Close(SaveChanges=True)
        xlApp.Quit()
            #Does the math to get the time for Reg_Time_Pivot_Table
        wb = load_workbook(filename = files)
        ws = wb.active
        pivotOne = wb['Reg_Time_Pivot_Table']
        num = 5
        while True:
            if(pivotOne['E'+str(num)].value == None):
                break
            else:
                part1='=SUM(E'
                part2=str(num)
                part3='/60)'
                pivotOne['F'+str(num)].value = part1+part2+part3
                num+= 1
        pivotOne['F4'].value = 'Sum of Service Minutes'
        wb.save(filename = files)
        wb.close()

                #-------------------Pivot 2-----------------
                #Create the 2nd pivot able for casefile status.
        def clear_pts(wsw):
            for pt in wsw.PivotTables():
                pt.TableRange2.Clear()

        def insert_pt_field_set1(pt):
            field_rows = {}
            field_rows['Campus'] = pt.PivotFields("Campus")
            field_rows['CasefileProviders'] = pt.PivotFields("CasefileProviders")
            
            field_values = {}
            field_values['CaseFileStatus'] = pt.PivotFields("CaseFileStatus")

            field_columns = {}
            field_columns['CCaseFileStatus'] = pt.PivotFields("CaseFileStatus")


                    # Insert row fields
                    # https://docs.microsoft.com/en-us/office/vba/api/excel.xlpivotfieldorientation
                    
            field_rows['Campus'].Orientation = 1
            field_rows['Campus'].Position = 1

            field_rows['CasefileProviders'].Orientation = 1
            field_rows['CasefileProviders'].Position = 2


                    # Insert values field
                    # https://docs.microsoft.com/en-us/office/vba/api/excel.xlconsolidationfunction

            field_values['CaseFileStatus'].Orientation = 4
            field_values['CaseFileStatus'].Function = -4112 # value reference
            field_values['CaseFileStatus'].NumberFormat = "#,##0"
                
            field_columns['CCaseFileStatus'].Orientation = 2
            field_columns['CCaseFileStatus'].Position = 1


                # Rquired to open excel in Win32Com
        xlApp = win32.Dispatch('Excel.Application')
        xlApp.Visible = False

                # Change the Path to Correct Folder if file is changed
        #dest = 'S:\\Programs\\Communities In Schools\\'+year+'\\6. Program Director Tools\\2023 Weekly Report Log\\'
        finalDest = backgroundInfo.path #+ filechange
            
                # Sheet Location and Name
        wbw = xlApp.Workbooks.Open(files)
        wsw_data = wbw.Worksheets(file)
        wsw_pivot = wbw.Worksheets("Reg_Status_Pivot_Table")


                # clear pivot tables on Report tab
        clear_pts(wsw_pivot)

                # create pt cache connection
        pt_cache = wbw.PivotCaches().Create(1, wsw_data.Range("A1").CurrentRegion)

                # insert pivot table designer/editor
        pt = pt_cache.CreatePivotTable(wsw_pivot.Range("B3"), "Service")

        insert_pt_field_set1(pt)

        wbw.Close(SaveChanges=True)
        xlApp.Quit()


                #----Close Out Tracking
        programDetails.Status = "Creating Outcomes sheet..." 
                #This creates the Close out section of the code
        wb = load_workbook(filename = files)
        ws = wb.active
        wb.create_sheet("CloseOut_Pivot_Table",0)
        wb.save(filename = files)
        wb.close()

        def clear_pts(wsw):
            for pt in wsw.PivotTables():
                pt.TableRange2.Clear()


        def insert_pt_field_set1(pt):
                
            field_rows = {}
            field_rows['Campus'] = pt.PivotFields("Campus")
            field_rows['CasefileProviders'] = pt.PivotFields("CasefileProviders")
            
            field_values = {}
            field_values['CaseFileStatus'] = pt.PivotFields("CaseFileStatus")

            field_filters = {}
            field_filters['NewCaseFileStatus'] = pt.PivotFields("CaseFileStatus")

                    # Insert row fields
                    # https://docs.microsoft.com/en-us/office/vba/api/excel.xlpivotfieldorientation
                    
            field_rows['Campus'].Orientation = 1
            field_rows['Campus'].Position = 1

            field_rows['CasefileProviders'].Orientation = 1
            field_rows['CasefileProviders'].Position = 2

            field_filters['NewCaseFileStatus'].Orientation = 3
            field_filters['NewCaseFileStatus'].Position = 1
                #These trys are used to check if the excel document has certain pivot items so it won't crash!
            try:
                pt.PivotFields("CaseFileStatus").PivotItems("Assessed").Visible = False
            except:
                None;
            else:
                pt.PivotFields("CaseFileStatus").PivotItems("Assessed").Visible = False
            
            try:
                pt.PivotFields("CaseFileStatus").PivotItems("Complete").Visible = True
            except:
                None;
            else:
                pt.PivotFields("CaseFileStatus").PivotItems("Complete").Visible = True
            
            try:
                pt.PivotFields("CaseFileStatus").PivotItems("Enrolled").Visible = False
            except:
                None;
            else:
                pt.PivotFields("CaseFileStatus").PivotItems("Enrolled").Visible = False
                
            try:
                pt.PivotFields("CaseFileStatus").PivotItems("In Program Transfer").Visible = True
            except:
                None;
            else:
                pt.PivotFields("CaseFileStatus").PivotItems("In Program Transfer").Visible = True
            
            try:
                pt.PivotFields("CaseFileStatus").PivotItems("Inactive").Visible = False
            except:
                None;
            else:
                pt.PivotFields("CaseFileStatus").PivotItems("Inactive").Visible = False
            
            try:
                pt.PivotFields("CaseFileStatus").PivotItems("Out of Program Transfer").Visible = True
            except:
                None;
            else:
                pt.PivotFields("CaseFileStatus").PivotItems("Out of Program Transfer").Visible = True
            
            try:
                pt.PivotFields("CaseFileStatus").PivotItems("Progress").Visible = False
            except:
                None;
            else:
                pt.PivotFields("CaseFileStatus").PivotItems("Progress").Visible = False
            
            try:
                pt.PivotFields("CaseFileStatus").PivotItems("Registered").Visible = False
            except:
                None;
            else:
                pt.PivotFields("CaseFileStatus").PivotItems("Registered").Visible = False


                    # Insert values field
                    # https://docs.microsoft.com/en-us/office/vba/api/excel.xlconsolidationfunction

            field_values['CaseFileStatus'].Orientation = 4
            field_values['CaseFileStatus'].Function = -4112 # value reference
            field_values['CaseFileStatus'].NumberFormat = "#,##0"


                # Rquired to open excel in Win32Com
        xlApp = win32.Dispatch('Excel.Application')
        xlApp.Visible = False

                # Change the Path to Correct Folder if file is changed
        #dest = 'S:\\Programs\\Communities In Schools\\'+year+'\\6. Program Director Tools\\2023 Weekly Report Log\\'
        finalDest = backgroundInfo.path #+ filechange
            
                # Sheet Location and Name
        wbw = xlApp.Workbooks.Open(files)
        wsw_data = wbw.Worksheets(file)
        wsw_pivot = wbw.Worksheets("CloseOut_Pivot_Table")
        wsw_pivot.Select()

                # clear pivot tables on Report tab
        clear_pts(wsw_pivot)

                # create pt cache connection
        pt_cache = wbw.PivotCaches().Create(1, wsw_data.Range("A1").CurrentRegion)

                # insert pivot table designer/editor
        pt = pt_cache.CreatePivotTable(wsw_pivot.Range("B3"), "Service")

        insert_pt_field_set1(pt)

        wbw.Close(SaveChanges=True)
        xlApp.Quit()
        programDetails.endProgram = True
        programDetails.Status = "Program Complete!"

while(backgroundInfo.previousTab == 0):
    root = mtTkinter.Tk()
    root.lift()
    root.attributes('-topmost', True)
    root.grab_set()
    root.grab_release()
    root.focus_force()
    root.update()
    root.iconbitmap("C:/Users/T Choat/Desktop/Python Code/MST Builder/cis.ico")
    root.title("Registration and Outcomes Builder")
    root.geometry("525x350")
    root.config(background = "white")
    root.minsize(525,350)
    root.maxsize(525,350)

    pathSelectLabel = Label(root, text="Select a R&O File:",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 10))

    pathEntry = Entry(width = 53)

    if(backgroundInfo.previousTab == 0):
        backgroundInfo.recallPath()
        pathBrowseButton = Button(root, text="Browse Folders", command= backgroundInfo.browse_button_One)
        
    nextButton = Button(root, text="Next",height = 2, width = 5, command = backgroundInfo.closeNameWindow)

    root.protocol("WM_DELETE_WINDOW", backgroundInfo.close_window)

    root.bind("<Control-Return>",backgroundInfo.closeNameWindow)
    root.bind("<Control-Escape>",backgroundInfo.close_window)
    root.bind("<Control-r>",backgroundInfo.browse_button_One)

    pathSelectLabel.place(x=20, y=30)
            
    pathEntry.place(x=155, y=30)
            
    pathBrowseButton.place(x=385, y=60)
          
    nextButton.place(x=450, y = 290)

    root.mainloop()
    while(backgroundInfo.previousTab == 1):
        loading = mtTkinter.Tk()
        loading.lift()
        loading.attributes('-topmost', True)
        loading.grab_set()
        loading.grab_release()
        loading.focus_force()
        loading.update()
        loading.iconbitmap("C:/Users/T Choat/Desktop/Python Code/MST Builder/cis.ico")
        loading.title("Registration and Outcomes Builder")
        loading.geometry("525x350")
        loading.config(background = "white")
        loading.minsize(525,350)
        loading.maxsize(525,350)
  

        def run_program(eventExecute = None):
            def progressBar():
                previousButton.config(state = "disabled")
                runButton.config(state = "disabled")
                i = 0
                progress_Bar.start()

                
                for i in range(10000000):
                    if(programDetails.endProgram == False):
                        #progress_Bar.update()
                        statusLabel.config(text = programDetails.Status)
                        loading.after(10)
                        loading.unbind("<Control-Return>")
                        loading.unbind("<Control-BackSpace>")
                        loading.bind("<Control-Escape>",backgroundInfo.close_window)
                    if(programDetails.endProgram == True):
                        statusLabel.config(text = programDetails.Status)
                        #progress_Bar["value"] = 300
                        loading.after(10)
                        #progress_Bar.update()
                        progress_Bar.stop()
                        closeButton.config(state = "active")
                        loading.unbind("<Control-Return>")
                        loading.unbind("<Control-BackSpace>")
                        loading.bind("<Control-Escape>",backgroundInfo.close_window)
                        
            threading.Thread(target=progressBar, daemon = True).start()
            threading.Thread(target=programDetails.program, daemon = True).start()


        progress_Bar = ttk.Progressbar(loading, orient = "horizontal", length = 385, mode = "indeterminate")
        progress_Bar.pack()

        statusLabel = Label(loading, text="Press Run to Continue...",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 10))
        statusLabel.pack()

        def closeButtonEnd():
            progress_Bar.stop()
            loading.destroy
            exit()

        closeButton = Button(loading, text="Close",height = 2, width = 5, command = backgroundInfo.close_window)
        closeButton.pack()
        if(programDetails.endProgram == False):
            progress_Bar.stop()
            closeButton.config(state = "disable")

        previousButton = Button(loading, text="Back",height = 2, width = 5, command = backgroundInfo.previous_Button)
        previousButton.pack()
        
        if(programDetails.endProgram == False):
            progress_Bar.stop()
            previousButton.config(state = "active")
            loading.bind("<Control-Return>",run_program)
            loading.bind("<Control-BackSpace>", backgroundInfo.previous_Button)
            loading.bind("<Control-Escape>",backgroundInfo.close_window)

        runButton = Button(loading, text="Run",height = 2, width = 5, command = run_program)

        loading.protocol("WM_DELETE_WINDOW", backgroundInfo.close_window)

        progress_Bar.place(x=65, y = 75)

        statusLabel.place(x=65, y=100)

        previousButton.place(x=330, y = 290)

        runButton.place(x=390, y = 290)

        closeButton.place(x=450, y = 290)
        
        loading.mainloop()

        loading.update()

    startOff = 1
exit()


#Goodluck!
