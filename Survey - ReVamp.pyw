#               Installing Pips Guide
#---------------------------------------------------------
#
#       Run the CIS PythonStartup Batch File
#
#               Last Update: 8/18/2023
#
#                 Server Version
#---------------------------------------------------------
# Excel-Python Program for MSTs only (Openpyxl, Win32com, Ensurezip, and Pandas)
# Developed by CISHOT Data Entry Team

#         Final Log : Chapter 3
#       The Ever      Lasting Task
#
#Hello Data Person,
#
#The Survey program is something that was made to scrap together information about campuses,
# services, and checks overall student activity. There are four files you need to run this program:
# Assessment Report, Service Mix Report, Campus Report, and Progress Report. You can
# do all campuses or just one campus to save time. 
#
#Goodbye,
#Data Dev.
#
#Disclaimer: Some numbers are off by a small margin due to formatting of code and triggers that
# are use to seperate information in the Service Mix file.
#
#Hope this helps and good luck!


#Anything red is either obsolete code or comments



# Required Information ------
from openpyxl import load_workbook
import openpyxl as opxl
import win32com.client as win32
import pywintypes
import os
from pywintypes import com_error
from win32com.client import constants
import pandas as pd
import numpy as np
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from mttkinter import *
import time
import threading

class backgroundDetails():
    pathA = ''
    pathM = ''
    pathC = ''
    pathP = ''
    previousTab = 0
    def browse_button_One(event1=None):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        pathEntryA.delete(0, END)
        pathEntryA.insert(END, filename)
        backgroundDetails.pathA = filename

    def browse_button_Two(event2=None):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        pathEntryM.delete(0, END)
        pathEntryM.insert(END, filename)
        backgroundDetails.pathM = filename

    def browse_button_Three(event3=None):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        pathEntryC.delete(0, END)
        pathEntryC.insert(END, filename)
        backgroundDetails.pathC = filename

    def browse_button_Four(event4=None):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        pathEntryP.delete(0, END)
        pathEntryP.insert(END, filename)
        backgroundDetails.pathP = filename

    def recallPath():
        pathEntryA.delete(0, END)
        pathEntryA.insert(END, backgroundDetails.pathA)
        pathEntryM.delete(0, END)
        pathEntryM.insert(END, backgroundDetails.pathM)
        pathEntryC.delete(0, END)
        pathEntryC.insert(END, backgroundDetails.pathC)
        pathEntryP.delete(0, END)
        pathEntryP.insert(END, backgroundDetails.pathP)

    def closeNameWindow(eventNext=None):
            backgroundDetails.pathA = pathEntryA.get()
            backgroundDetails.pathM = pathEntryM.get()
            backgroundDetails.pathC = pathEntryC.get()
            backgroundDetails.pathP = pathEntryP.get()
            if(backgroundDetails.pathA != "" and backgroundDetails.pathM != "" and backgroundDetails.pathC != "" and backgroundDetails.pathP != ""):
                backgroundDetails.previousTab = 1
                root.destroy()

    def close_window(eventClose=None):
        exit();

    def previous_Button(eventPrevious=None):
        backgroundDetails.previousTab = 0
        loading.destroy()

def dispatch(app_name:str):
    try:
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    except AttributeError:
        # Corner case dependencies.
        import os
        import re
        import sys
        import shutil
        # Remove cache and try again.
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    return app



#Phase 1
class programDetails():
    Status = ''
    endProgram = False
    def program(eventRun=None):
        programDetails.Status = ""
        programDetails.Status = "Merging Sheets..."
        breaker = True
        pathTwo = ""
        pathOne = ""
        pathFour = ""
        pathFive = ""
        while (breaker != False):
            fileM = os.path.split(backgroundDetails.pathM)[1]
            fileM = fileM.replace(".xlsx","")
            fileA = os.path.split(backgroundDetails.pathA)[1]
            fileA = fileA.replace(".xlsx","")
            fileC = os.path.split(backgroundDetails.pathC)[1]
            fileC = fileC.replace(".xlsx","")
            fileP = os.path.split(backgroundDetails.pathP)[1]
            fileP = fileP.replace(".xlsx","")
            #locationOne = input('Enter the file name of the Assessed Report\n')
            locationOne = fileA
            #fileChangeOne = locationOne + '.xlsx'
            #locationTwo = input('Enter the file name of the Service Mix Report\n')
            locationTwo = fileM
            #fileChangeTwo = locationTwo + '.xlsx'
            #locationFour = input('Enter the file name of the Campus Report\n')
            locationFour = fileC
            #fileChangeFour = locationFour + '.xlsx'
            #locationFive = input('Enter the file name of the Progress Report\n')
            locationFive = fileP
            #fileChangeFive = locationFive + '.xlsx'
            path = 'C:/Users/T Choat/Desktop/Python Code/S-U-R-V-E-Y L-I-S-T/'
            pathTwo = backgroundDetails.pathM
            pathOne = backgroundDetails.pathA
            pathFour = backgroundDetails.pathC
            pathFive = backgroundDetails.pathP
            #checkLocation = input('Are you sure you want to use files: '+ locationOne + ', '+ locationTwo+ ', '+ locationFour+', and ' + locationFive + '\n')
            #if (checkLocation == "yes" or checkLocation == "Yes" or checkLocation == "YES" or checkLocation == "Y" or checkLocation == "y"):
            #break
            breaker = False

        #For Assessed Report

        ws = load_workbook(filename = pathOne)
        wb = ws.active

        if (wb['F1'].value == 'Campus_Name__CDN_'):
            wb.delete_cols(1, 5);

        if (wb['G1'].value == 'CasefileId'):
            wb.delete_cols(2, 5);
            
        if (wb['I1'].value == 'NameFirst'):
            wb.delete_cols(3, 5);

        if (wb['S1'].value == 'AssessedAreas'):
            wb.delete_cols(7, 12);

        if (wb['O1'].value == 'AssessmentCompleteDate'):
            wb.delete_cols(10, 5);

        if (wb['J1'].value == 'AssessmentCompleteDate'):
            wb.delete_cols(11, 10);

        wb.auto_filter.ref = wb.dimensions

        ws.save(pathOne)
        ws.close()
        del ws
        del wb

        #For Mix Report

        wsMix = load_workbook(filename = pathTwo)
        wbMix = wsMix.active

        if (wbMix['F1'].value == 'CampusName_CDN'):
            wbMix.delete_cols(1, 5);

        if (wbMix['G1'].value == 'CASEFILE_ID'):
            wbMix.delete_cols(2, 5);

        if (wbMix['E1'].value == 'NAME_FIRST'):
            wbMix.delete_cols(3, 2);

        if (wbMix['G1'].value == 'GENDER_CD'):
            wbMix.delete_cols(8, 3);

        if (wbMix['H1'].value == 'ETHNICITY_CD'):
            wbMix.delete_cols(9, 6);

        if (wbMix['I1'].value == 'ASSESSED_AREAS'):
            wbMix.delete_cols(10, 10);

        if (wbMix['J1'].value == 'SERVICE_DATE'):
            wbMix.delete_cols(11, 3);

        if (wbMix['L1'].value == 'SERVICE_CODE_DESC'):
            wbMix.delete_cols(13, 4);

        if (wbMix['M1'].value == 'PRIMARY_PROVIDER'):
            wbMix.delete_cols(14, 3);

        if (wbMix['N1'].value == 'NOTES'):
            wbMix.delete_cols(15, 2);



        wbMix.auto_filter.ref = wbMix.dimensions

        wsMix.save(pathTwo)
        wsMix.close()
        del wsMix

        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        source = xlApp.Workbooks.Open(pathOne)
        sourceCampus = xlApp.Workbooks.Open(pathFour)
        sourceProgress = xlApp.Workbooks.Open(pathFive)
        MixSort = target.Worksheets(locationTwo)

        MixSort.Name = locationTwo

        source.Worksheets(1).Name = locationOne
        cutting = source.Worksheets(1)
        cutting.Copy(Before=MixSort)

        sourceCampus.Worksheets(1).Name = locationFour
        cutting = sourceCampus.Worksheets(1)
        cutting.Copy(After=MixSort)

        sourceProgress.Worksheets(1).Name = locationFive
        cutting = sourceProgress.Worksheets(1)
        cutting.Copy(After=MixSort)

        target.Close(SaveChanges=True)
        source.Close(SaveChanges=True)
        sourceCampus.Close(SaveChanges=True)
        sourceProgress.Close(SaveChanges=True)
        xlApp.Quit()

        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        source = xlApp.Workbooks.Open(pathOne)
        MixSort = target.Worksheets(locationOne)
        ranMixSort= MixSort.Range('A2:Z100000')
        ranMixSort.Sort(Key1=MixSort.Range('B1'), Order1=1, Orientation=1)

        AssessSort = target.Worksheets(locationOne)
        rangeAssessSort= AssessSort.Range('A2:Z100000')
        rangeAssessSort.Sort(Key1=AssessSort.Range('B1'), Order1=1, Orientation=1)
        MixSort.Columns.AutoFit()
        AssessSort.Columns.AutoFit()
        MixSort.Rows.AutoFit()
        AssessSort.Rows.AutoFit()

        target.Close(SaveChanges=True)
        source.Close(SaveChanges=True)

        xlApp.Quit()
        programDetails.Status = ""
        programDetails.Status = "Page Sorted..."

        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        Mix = target.Worksheets(locationTwo)
        Assess = target.Worksheets(locationOne)

        ranMixSort= Mix.Range('A2:O500000')
        ranMixSort.Sort(Key1=Mix.Range('B1'), Order1=1, Orientation=1)

        assessStep = 2
        mixStep = 2
        mixValue = ''
        assessValue = ''
        checkPoint = 2
        invalidPoint = 2
        check = 0

        while(True):
            mixValue = Mix.Range('B'+str(mixStep)).Value
            assessValue = Assess.Range('B'+str(assessStep)).Value
            if(mixValue == 'CASEFILE_ID'):
                mixStep += 1
            if(str(mixValue) in str(assessValue)):
                assessStep += 1
                checkPoint = mixStep
                invalidPoint = assessStep
            elif(str(assessValue) == None or str(mixValue) == None or str(mixValue) == 'None' or str(assessValue) == 'None'):
                break;
            elif(int(assessValue) < int(mixValue) and (str(assessValue) != None or str(mixValue) != None)):
                Assess.Range("A"+str(assessStep)+":O"+str(assessStep)).Delete(Shift=-4162)
                mixStep = checkPoint
                assessStep = invalidPoint
                assessStep += 1
            mixStep += 1


    #    while(True):
    #        mixValue = Mix.Range('B'+str(mixStep)).Value
    #        assessValue = Assess.Range('B'+str(assessStep)).Value
    #        if(mixValue == 'CASEFILE_ID'):
    #            mixStep += 1
    #        if(str(mixValue) in str(assessValue)):
    #            assessStep += 1
    #            checkPoint = mixStep
    #            invalidPoint = assessStep
    #        elif(assessValue == None):
    #            break;
    #        elif(int(assessValue) < int(mixValue)):
    #            Mix.Range("A"+str(mixStep)+":O"+str(mixStep)).Delete(Shift=-4162)
    #            mixStep = checkPoint
    #            assessStep = invalidPoint
    #        mixStep += 1

        target.Close(SaveChanges=True)
        xlApp.Quit()
        programDetails.Status = ""
        programDetails.Status = "List compiled..."

        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        target2 = xlApp.Workbooks.Open(pathOne)
        Mix = target.Worksheets(locationTwo)
        Assess = target.Worksheets(locationOne)

        assessStep = 2
        mixStep = 2
        mixValue =0
        assessValue =0
        checkMixValue = 0
        checkAssessValue = 0
        flip = 1
        
        target.Close(SaveChanges=True)
        xlApp.Quit()

        programDetails.Status = ""
        programDetails.Status = "Page Styled..."
        pathListSurvey = 'C:/Users/T Choat/Desktop/Python Code/S-U-R-V-E-Y L-I-S-T/Survey Codes'
        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        target2 = xlApp.Workbooks.Open(pathOne)
        target3 = xlApp.Workbooks.Open(pathListSurvey)
        Mix = target.Worksheets(locationTwo)
        Assess = target.Worksheets(locationOne)
        Survey = target3.Worksheets('List')

        assessStep = 2
        mixStep = 2
        mixValue = ''
        assessValue = ''
        service = str(Assess.Range('I'+str(assessStep)).Value)
        service = service.replace(".0","")
        service = service.replace("703218CI","")
        service = service.replace("1019","")
        service = service.replace("1009","")
        serviceMix = str(Mix.Range('K'+str(mixStep)).Value)
        serviceMix = serviceMix.replace(".0","")
        mixValue = Mix.Range('B'+str(mixStep)).Value
        assessValue = Assess.Range('B'+str(assessStep)).Value
        countOverall = 0
        countService = 0
        countMonthly = 0
        countIndirect = 0
        countAllOther = 0

        countServiceTracker = 0
        countMonthlyTracker = 0
        countIndirectTracker = 0
        countAllOtherTracker = 0

        total = 0

        sumServiceTracker = 0
        sumMonthlyTracker = 0
        sumIndirectTracker = 0
        sumAllOtherTracker = 0
        percentLimit = .50

        highDateMax = str(Mix.Range('J'+str(mixStep)).Value);
        highDate = str(Mix.Range('J'+str(mixStep)).Value);
        compareDate = str(Mix.Range('J'+str(mixStep)).Value);
        lowDate = str(Mix.Range('J'+str(mixStep)).Value);
        lowDateMax = str(Mix.Range('J'+str(mixStep)).Value);
        highMixMax = ""
        highMix = ""
        lowMix = ""
        lowMixMax = ""
        datePart1 = "=DATEDIF(J"
        datePart1Alt = "=DATEDIF(\'" + str(locationTwo) + "\'!J"
        datePart2 = ",J"
        datePart2Alt = ",\'" + str(locationTwo) + "\'!J"
        datePart3 = ',\"m\")+1' #Plus one at end
        flip = 1
        highMix = 2
        lowMix = 2

        maleTracker = 0
        femaleTracker = 0
        switchSex = 0
        sumServiceTrackerMale = 0
        sumServiceTrackerFemale = 0
        sumMonthlyTrackerMale = 0
        sumMonthlyTrackerFemale = 0


        switchEth = 0
        whiteEth = 0
        blackEth = 0
        hispanicEth = 0
        asianEth = 0
        nativeEth = 0
        pacificEth = 0

        #Total services for ethicity/race
        sumAsianServiceTacker = 0
        sumBlackServiceTacker = 0
        sumHispanicnServiceTacker = 0
        sumNativeServiceTacker = 0
        sumWhiteServiceTacker = 0
        sumPacificServiceTacker = 0


        #Gender/Sex total Services
        maleServiced = 0
        femaleServiced = 0

        #Total services for ethicity/race
        asianServiced = 0
        blackServiced = 0
        hispanicServiced = 0
        nativeServiced = 0
        whiteServiced = 0
        pacificServiced = 0

        asianMaleAllService = 0
        asianFemaleAllService = 0
        blackMaleAllService = 0
        blackFemaleAllService = 0
        hispanicMaleAllService = 0
        hispanicFemaleAllService = 0
        nativeMaleAllService = 0
        nativeFemaleAllService = 0
        whiteMaleAllService = 0
        whiteFemaleAllService = 0
        pacificMaleAllService = 0
        pacificFemaleAllService = 0

        #To get total in specifics
        asianMale = 0
        asianFemale = 0
        blackMale = 0
        blackFemale = 0
        hispanicMale = 0
        hispanicFemale = 0
        nativeMale = 0
        nativeFemale = 0
        whiteMale = 0
        whiteFemale = 0
        pacificMale = 0
        pacificFemale = 0

        #Monthly service for specifics under race/gender catagories (Can be used to total race as well)
        asianMaleMonthly = 0
        asianFemaleMonthly = 0
        blackMaleMonthly = 0
        blackFemaleMonthly = 0
        hispanicMaleMonthly = 0
        hispanicFemaleMonthly = 0
        nativeMaleMonthly = 0
        nativeFemaleMonthly = 0
        whiteMaleMonthly = 0
        whiteFemaleMonthly = 0
        pacificMaleMonthly = 0
        pacificFemaleMonthly = 0

        #Assessed service for specifics under race/gender catagories (Can be used to total race as well)
        asianMaleService = 0
        asianFemaleService = 0
        blackMaleService = 0
        blackFemaleService = 0
        hispanicMaleService = 0
        hispanicFemaleService = 0
        nativeMaleService = 0
        nativeFemaleService = 0
        whiteMaleService = 0
        whiteFemaleService = 0
        pacificMaleService = 0
        pacificFemaleService = 0

        #Indirect service for specifics under race/gender catagories (Can be used to total race as well)
        asianMaleIndirect = 0
        asianFemaleIndirect = 0
        blackMaleIndirect = 0
        blackFemaleIndirect = 0
        hispanicMaleIndirect = 0
        hispanicFemaleIndirect = 0
        nativeMaleIndirect = 0
        nativeFemaleIndirect = 0
        whiteMaleIndirect = 0
        whiteFemaleIndirect = 0
        pacificMaleIndirect = 0
        pacificFemaleIndirect = 0

        #All Other service for specifics under race/gender catagories (Can be used to total race as well)
        asianMaleOther = 0
        asianFemaleOther = 0
        blackMaleOther = 0
        blackFemaleOther = 0
        hispanicMaleOther = 0
        hispanicFemaleOther = 0
        nativeMaleOther = 0
        nativeFemaleOther = 0
        whiteMaleOther = 0
        whiteFemaleOther = 0
        pacificMaleOther = 0
        pacificFemaleOther = 0

        #All service for specifics under race/gender catagories (Can be used to total race as well)
        asianMaleOverall = 0
        asianFemaleOverall = 0
        blackMaleOverall = 0
        blackFemaleOverall = 0
        hispanicMaleOverall = 0
        hispanicFemaleOverall = 0
        nativeMaleOverall = 0
        nativeFemaleOverall = 0
        whiteMaleOverall = 0
        whiteFemaleOverall = 0
        pacificMaleOverall = 0
        pacificFemaleOverall = 0

        #Completed service counter for ethnicity and gender/sex
        asianMaleServiced = 0
        asianFemaleServiced = 0
        blackMaleServiced = 0
        blackFemaleServiced = 0
        hispanicMaleServiced = 0
        hispanicFemaleServiced = 0
        nativeMaleServiced = 0
        nativeFemaleServiced = 0
        whiteMaleServiced = 0
        whiteFemaleServiced = 0
        pacificMaleServiced = 0
        pacificFemaleServiced = 0

        #To get total of Assessed Areas
        countAcademics = 0
        countBehavior = 0
        countSocialService = 0
        countAttendance = 0

        #Assessed Areas for specifics under race/Ethicity catagories
        asianBehavioral = 0
        asianAcademics = 0
        asianSocial = 0
        asianAttendance = 0

        blackBehavioral = 0
        blackAcademics = 0
        blackSocial = 0
        blackAttendance = 0

        hispanicBehavioral = 0
        hispanicAcademics = 0
        hispanicSocial = 0
        hispanicAttendance = 0

        nativeBehavioral = 0
        nativeAcademics = 0
        nativeSocial = 0
        nativeAttendance = 0

        whiteBehavioral = 0
        whiteAcademics = 0
        whiteSocial = 0
        whiteAttendance = 0

        pacificBehavioral = 0
        pacificAcademics = 0
        pacificSocial = 0
        pacificAttendance = 0

        #Assessed Areas for specifics under race/gender catagories
        asianMaleBehavioral = 0
        asianMaleAcademics = 0
        asianMaleSocial = 0
        asianMaleAttendance = 0
        asianFemaleBehavioral = 0
        asianFemaleAcademics = 0
        asianFemaleSocial = 0
        asianFemaleAttendance = 0

        blackMaleBehavioral = 0
        blackMaleAcademics = 0
        blackMaleSocial = 0
        blackMaleAttendance = 0
        blackFemaleBehavioral = 0
        blackFemaleAcademics = 0
        blackFemaleSocial = 0
        blackFemaleAttendance = 0

        hispanicMaleBehavioral = 0
        hispanicMaleAcademics = 0
        hispanicMaleSocial = 0
        hispanicMaleAttendance = 0
        hispanicFemaleBehavioral = 0
        hispanicFemaleAcademics = 0
        hispanicFemaleSocial = 0
        hispanicFemaleAttendance = 0

        nativeMaleBehavioral = 0
        nativeMaleAcademics = 0
        nativeMaleSocial = 0
        nativeMaleAttendance = 0
        nativeFemaleBehavioral = 0
        nativeFemaleAcademics = 0
        nativeFemaleSocial = 0
        nativeFemaleAttendance = 0

        whiteMaleBehavioral = 0
        whiteMaleAcademics = 0
        whiteMaleSocial = 0
        whiteMaleAttendance = 0
        whiteFemaleBehavioral = 0
        whiteFemaleAcademics = 0
        whiteFemaleSocial = 0
        whiteFemaleAttendance = 0

        pacificMaleBehavioral = 0
        pacificMaleAcademics = 0
        pacificMaleSocial = 0
        pacificMaleAttendance = 0
        pacificFemaleBehavioral = 0
        pacificFemaleAcademics = 0
        pacificFemaleSocial = 0
        pacificFemaleAttendance = 0

        behaviorServiced = 0
        academicServiced = 0
        socialServiced = 0
        attendanceServiced = 0

        female = 0
        male = 0

        #Outcome Reserves (Outcome ) (Race/Eth) (Male or Female)

        AAF = 0
        AAM = 0
        ABF = 0
        ABM = 0
        AHF = 0
        AHM = 0
        AIF = 0
        AIM = 0
        AWF = 0
        AWM = 0
        APF = 0
        APM = 0

        ARAF = 0
        ARAM = 0
        ARBF = 0
        ARBM = 0
        ARHF = 0
        ARHM = 0
        ARIF = 0
        ARIM = 0
        ARWF = 0
        ARWM = 0
        ARPF = 0
        ARPM = 0

        BNAF = 0
        BNAM = 0
        BNBF = 0
        BNBM = 0
        BNHF = 0
        BNHM = 0
        BNIF = 0
        BNIM = 0
        BNWF = 0
        BNWM = 0
        BNPF = 0
        BNPM = 0

        CCAF = 0
        CCAM = 0
        CCBF = 0
        CCBM = 0
        CCHF = 0
        CCHM = 0
        CCIF = 0
        CCIM = 0
        CCWF = 0
        CCWM = 0
        CCPF = 0
        CCPM = 0

        CPAF = 0
        CPAM = 0
        CPBF = 0
        CPBM = 0
        CPHF = 0
        CPHM = 0
        CPIF = 0
        CPIM = 0
        CPWF = 0
        CPWM = 0
        CPPF = 0
        CPPM = 0

        CCRAF = 0
        CCRAM = 0
        CCRBF = 0
        CCRBM = 0
        CCRHF = 0
        CCRHM = 0
        CCRIF = 0
        CCRIM = 0
        CCRWF = 0
        CCRWM = 0
        CCRPF = 0
        CCRPM = 0

        DCAF = 0
        DCAM = 0
        DCBF = 0
        DCBM = 0
        DCHF = 0
        DCHM = 0
        DCIF = 0
        DCIM = 0
        DCWF = 0
        DCWM = 0
        DCPF = 0
        DCPM = 0

        FCAF = 0
        FCAM = 0
        FCBF = 0
        FCBM = 0
        FCHF = 0
        FCHM = 0
        FCIF = 0
        FCIM = 0
        FCWF = 0
        FCWM = 0
        FCPF = 0
        FCPM = 0

        GAF = 0
        GAM = 0
        GBF = 0
        GBM = 0
        GHF = 0
        GHM = 0
        GIF = 0
        GIM = 0
        GWF = 0
        GWM = 0
        GPF = 0
        GPM = 0

        GLAF = 0
        GLAM = 0
        GLBF = 0
        GLBM = 0
        GLHF = 0
        GLHM = 0
        GLIF = 0
        GLIM = 0
        GLWF = 0
        GLWM = 0
        GLPF = 0
        GLPM = 0

        HCAF = 0
        HCAM = 0
        HCBF = 0
        HCBM = 0
        HCHF = 0
        HCHM = 0
        HCIF = 0
        HCIM = 0
        HCWF = 0
        HCWM = 0
        HCPF = 0
        HCPM = 0

        LDAF = 0
        LDAM = 0
        LDBF = 0
        LDBM = 0
        LDHF = 0
        LDHM = 0
        LDIF = 0
        LDIM = 0
        LDWF = 0
        LDWM = 0
        LDPF = 0
        LDPM = 0

        LSAF = 0
        LSAM = 0
        LSBF = 0
        LSBM = 0
        LSHF = 0
        LSHM = 0
        LSIF = 0
        LSIM = 0
        LSWF = 0
        LSWM = 0
        LSPF = 0
        LSPM = 0

        MHCAF = 0
        MHCAM = 0
        MHCBF = 0
        MHCBM = 0
        MHCHF = 0
        MHCHM = 0
        MHCIF = 0
        MHCIM = 0
        MHCWF = 0
        MHCWM = 0
        MHCPF = 0
        MHCPM = 0

        MHWAF = 0
        MHWAM = 0
        MHWBF = 0
        MHWBM = 0
        MHWHF = 0
        MHWHM = 0
        MHWIF = 0
        MHWIM = 0
        MHWWF = 0
        MHWWM = 0
        MHWPF = 0
        MHWPM = 0

        RISAF = 0
        RISAM = 0
        RISBF = 0
        RISBM = 0
        RISHF = 0
        RISHM = 0
        RISIF = 0
        RISIM = 0
        RISWF = 0
        RISWM = 0
        RISPF = 0
        RISPM = 0

        RDMAF = 0
        RDMAM = 0
        RDMBF = 0
        RDMBM = 0
        RDMHF = 0
        RDMHM = 0
        RDMIF = 0
        RDMIM = 0
        RDMWF = 0
        RDMWM = 0
        RDMPF = 0
        RDMPM = 0

        SEMAF = 0
        SEMAM = 0
        SEMBF = 0
        SEMBM = 0
        SEMHF = 0
        SEMHM = 0
        SEMIF = 0
        SEMIM = 0
        SEMWF = 0
        SEMWM = 0
        SEMPF = 0
        SEMPM = 0

        SEAF = 0
        SEAM = 0
        SEBF = 0
        SEBM = 0
        SEHF = 0
        SEHM = 0
        SEIF = 0
        SEIM = 0
        SEWF = 0
        SEWM = 0
        SEPF = 0
        SEPM = 0

        SRAF = 0
        SRAM = 0
        SRBF = 0
        SRBM = 0
        SRHF = 0
        SRHM = 0
        SRIF = 0
        SRIM = 0
        SRWF = 0
        SRWM = 0
        SRPF = 0
        SRPM = 0

        SSAF = 0
        SSAM = 0
        SSBF = 0
        SSBM = 0
        SSHF = 0
        SSHM = 0
        SSIF = 0
        SSIM = 0
        SSWF = 0
        SSWM = 0
        SSPF = 0
        SSPM = 0

        TAF = 0
        TAM = 0
        TBF = 0
        TBM = 0
        THF = 0
        THM = 0
        TIF = 0
        TIM = 0
        TWF = 0
        TWM = 0
        TPF = 0
        TPM = 0

        TRAF = 0
        TRAM = 0
        TRBF = 0
        TRBM = 0
        TRHF = 0
        TRHM = 0
        TRIF = 0
        TRIM = 0
        TRWF = 0
        TRWM = 0
        TRPF = 0
        TRPM = 0




        AMSecondary = 0
        AMPreparation = 0
        AMExploration = 0
        AMEmployment = 0
        AMFAFSA = 0

        AFSecondary = 0
        AFPreparation = 0
        AFExploration = 0
        AFEmployment = 0
        AFFAFSA = 0

        BMSecondary = 0
        BMPreparation = 0
        BMExploration = 0
        BMEmployment = 0
        BMFAFSA = 0

        BFSecondary = 0
        BFPreparation = 0
        BFExploration = 0
        BFEmployment = 0
        BFFAFSA = 0

        HMSecondary = 0
        HMPreparation = 0
        HMExploration = 0
        HMEmployment = 0
        HMFAFSA = 0

        HFSecondary = 0
        HFPreparation = 0
        HFExploration = 0
        HFEmployment = 0
        HFFAFSA = 0

        IMSecondary = 0
        IMPreparation = 0
        IMExploration = 0
        IMEmployment = 0
        IMFAFSA = 0

        IFSecondary = 0
        IFPreparation = 0
        IFExploration = 0
        IFEmployment = 0
        IFFAFSA = 0

        WMSecondary = 0
        WMPreparation = 0
        WMExploration = 0
        WMEmployment = 0
        WMFAFSA = 0

        WFSecondary = 0
        WFPreparation = 0
        WFExploration = 0
        WFEmployment = 0
        WFFAFSA = 0

        PMSecondary = 0
        PMPreparation = 0
        PMExploration = 0
        PMEmployment = 0
        PMFAFSA = 0

        PFSecondary = 0
        PFPreparation = 0
        PFExploration = 0
        PFEmployment = 0
        PFFAFSA = 0


        AMSecondaryFull = 0
        AMPreparationFull = 0
        AMExplorationFull = 0
        AMEmploymentFull = 0
        AMFAFSAFull = 0

        AFSecondaryFull = 0
        AFPreparationFull = 0
        AFExplorationFull = 0
        AFEmploymentFull = 0
        AFFAFSAFull = 0

        BMSecondaryFull = 0
        BMPreparationFull = 0
        BMExplorationFull = 0
        BMEmploymentFull = 0
        BMFAFSAFull = 0

        BFSecondaryFull = 0
        BFPreparationFull = 0
        BFExplorationFull = 0
        BFEmploymentFull = 0
        BFFAFSAFull = 0

        HMSecondaryFull = 0
        HMPreparationFull = 0
        HMExplorationFull = 0
        HMEmploymentFull = 0
        HMFAFSAFull = 0

        HFSecondaryFull = 0
        HFPreparationFull = 0
        HFExplorationFull = 0
        HFEmploymentFull = 0
        HFFAFSAFull = 0

        IMSecondaryFull = 0
        IMPreparationFull = 0
        IMExplorationFull = 0
        IMEmploymentFull = 0
        IMFAFSAFull = 0

        IFSecondaryFull = 0
        IFPreparationFull = 0
        IFExplorationFull = 0
        IFEmploymentFull = 0
        IFFAFSAFull = 0

        WMSecondaryFull = 0
        WMPreparationFull = 0
        WMExplorationFull = 0
        WMEmploymentFull = 0
        WMFAFSAFull = 0

        WFSecondaryFull = 0
        WFPreparationFull = 0
        WFExplorationFull = 0
        WFEmploymentFull = 0
        WFFAFSAFull = 0

        PMSecondaryFull = 0
        PMPreparationFull = 0
        PMExplorationFull = 0
        PMEmploymentFull = 0
        PMFAFSAFull = 0

        PFSecondaryFull = 0
        PFPreparationFull = 0
        PFExplorationFull = 0
        PFEmploymentFull = 0
        PFFAFSAFull = 0

        surveyStep = 2
        shortList = []

        while(True):
            mixValue = Mix.Range('B'+str(mixStep)).Value
            assessValue = Assess.Range('B'+str(assessStep)).Value
            pastmixValue = Mix.Range('B'+str(mixStep+1)).Value
            #if((mixStep-1) == 1):
            #    pastmixValue = Mix.Range('B'+str(mixStep)).Value
            while(mixValue != assessValue):
                if(assessValue != mixValue):
                    assessStep += 1
                    assessValue = Assess.Range('B'+str(assessStep)).Value
            if(mixValue == assessValue):
                countOverall += 1
                mixValue = Mix.Range('B'+str(mixStep)).Value
                assessValue = Assess.Range('B'+str(assessStep)).Value
                #print(str(mixValue) + " " + str(assessValue))
                service = str(Assess.Range('I'+str(assessStep)).Value)
                service = service.replace(".0","")
                service = service.replace("703218CI","")
                service = service.replace("1019","")
                service = service.replace("1009","")
                service = service.replace(","," ,")
                serviceMix = str(Mix.Range('K'+str(mixStep)).Value)
                serviceMix = serviceMix.replace(".0","")
                #print("Tick " + str(serviceMix))
                mixValue = Mix.Range('B'+str(mixStep)).Value
                assessValue = Assess.Range('B'+str(assessStep)).Value
                compareDate = str(Mix.Range('J'+str(mixStep)).Value)
                if(mixValue == assessValue and switchSex == 0):
                    if(Mix.Range('G'+str(mixStep)).Value == "M" and switchSex == 0):
                        maleTracker = maleTracker +1
                        switchSex = switchSex +1
                    elif(Mix.Range('G'+str(mixStep)).Value == "F" and switchSex == 0):
                        femaleTracker = femaleTracker +1
                        switchSex = switchSex +1
                if(mixValue != pastmixValue):
                    switchEth = 0
                if(switchEth == 0 and mixValue != pastmixValue):
                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and switchEth == 0): #Reports have white space so have to add space to get report to calculate
                        whiteEth = whiteEth + 1
                        switchEth = switchEth + 1
                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and switchEth == 0):
                        blackEth = blackEth + 1
                        switchEth = switchEth + 1
                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and switchEth == 0):
                        hispanicEth = hispanicEth + 1
                        switchEth = switchEth + 1
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and switchEth == 0):
                        asianEth = asianEth + 1
                        switchEth = switchEth + 1
                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and switchEth == 0):
                        nativeEth = nativeEth + 1
                        switchEth = switchEth + 1
                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and switchEth == 0):
                        pacificEth = pacificEth + 1
                        switchEth = switchEth + 1
                if(compareDate != 'None' and mixValue == assessValue):
                    if(highDate < compareDate and compareDate != None and highDate != None):
                        highDate = compareDate
                        highMix = mixStep
                        if (highDateMax < compareDate and compareDate !=None):
                            highDateMax = compareDate
                            highMixMax = mixStep
                    elif(highDate == compareDate and compareDate != None):
                        highDate = highDate
                        highMix = mixStep
                    if(lowDate > compareDate and compareDate != None and lowDate != None):
                        lowDate = compareDate
                        lowMix = mixStep
                        if (lowDateMax > compareDate and compareDate !=None):
                            lowDateMax = compareDate
                            lowMixMax = mixStep
                    elif(lowDate == compareDate and compareDate != None):
                        lowDate = lowDate
                        lowMix = mixStep
                #print(highDateMax)
                serviceMix = serviceMix + " "
                service = service + " "
                if(str(mixValue) in str(assessValue)):
                    if(str(serviceMix) in str(service)):   # 8/30/23 New Line
                        countService = countService + 1
                        #print("Mix " + str(mixValue) + " has " + str(serviceMix) + " and Assess " + str(assessValue) + " has " + str(service) + "  total = " + str(countService)) # Check 1570967
                        
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AMSecondary += 1
                        elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AMPreparation += 1
                        elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AMExploration += 1
                        elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AMEmployment += 1
                        elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AMFAFSA += 1

                        if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AFSecondary += 1
                        elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AFPreparation += 1
                        elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AFExploration += 1
                        elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AFEmployment += 1
                        elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AFFAFSA += 1

                        if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BMSecondary += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BMPreparation += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BMExploration += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BMEmployment += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BMFAFSA += 1

                        if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BFSecondary += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BFPreparation += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BFExploration += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BFEmployment += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BFFAFSA += 1

                        if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HMSecondary += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HMPreparation += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HMExploration += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HMEmployment += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HMFAFSA += 1

                        if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HFSecondary += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HFPreparation += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HFExploration += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HFEmployment += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HFFAFSA += 1
                                
                        if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IMSecondary += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IMPreparation += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IMExploration += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IMEmployment += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IMFAFSA += 1

                        if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IFSecondary += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IFPreparation += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IFExploration += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IFEmployment += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IFFAFSA += 1

                        if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WMSecondary += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WMPreparation += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WMExploration += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WMEmployment += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WMFAFSA += 1

                        if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WFSecondary += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WFPreparation += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WFExploration += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WFEmployment += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                WFFAFSA += 1

                        if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PMSecondary += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PMPreparation += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PMExploration += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PMEmployment += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PMFAFSA += 1

                        if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Post Secondary" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PFSecondary += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "4038" in str(Mix.Range('K'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PFPreparation += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Career Exploration" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PFExploration += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Employment Skills" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PFEmployment += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "FAFSA" in str(Mix.Range('L'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                PFFAFSA += 1

                        if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianFemaleService += 1
                        elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianMaleService += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackFemaleService += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackMaleService += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicFemaleService += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicMaleService += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeFemaleService += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeMaleService += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):    
                            whiteFemaleService += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):    
                            whiteMaleService += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificFemaleService += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificMaleService += 1

                    surveyStep = 2
                    fixSurvey = str(Survey.Range('A'+str(surveyStep)).Value)
                    fixSurvey = fixSurvey.replace(".0","")
                    while(fixSurvey != 'None'):
                        fixSurvey = str(Survey.Range('A'+str(surveyStep)).Value)
                        fixSurvey = fixSurvey.replace(".0","")
                        if(str(fixSurvey) in str(serviceMix)):
                            if('Indirect' == (Survey.Range('C'+str(surveyStep)).Value)):
                                countIndirect = countIndirect + 1
                                if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianFemaleIndirect += 1
                                elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianMaleIndirect += 1
                                elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackFemaleIndirect += 1
                                elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackMaleIndirect += 1
                                elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicFemaleIndirect += 1
                                elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicMaleIndirect += 1
                                elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeFemaleIndirect += 1
                                elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeMaleIndirect += 1
                                elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):    
                                    whiteFemaleIndirect += 1
                                elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):    
                                    whiteMaleIndirect += 1
                                elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificFemaleIndirect += 1
                                elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificMaleIndirect += 1
                            if(fixSurvey == 'None'):
                                print(str(serviceMix))
                            if(str(fixSurvey) in str(serviceMix)):
                                if('Monthly' == Survey.Range('C'+str(surveyStep)).Value):
                                    countMonthly = 1 + countMonthly
                                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        asianFemaleMonthly += 1
                                    elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        asianMaleMonthly += 1
                                    elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        blackFemaleMonthly += 1
                                    elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        blackMaleMonthly += 1
                                    elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        hispanicFemaleMonthly += 1
                                    elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        hispanicMaleMonthly += 1
                                    elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        nativeFemaleMonthly += 1
                                    elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        nativeMaleMonthly += 1
                                    elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):    
                                        whiteFemaleMonthly += 1
                                    elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):    
                                        whiteMaleMonthly += 1
                                    elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        pacificFemaleMonthly += 1
                                    elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                        pacificMaleMonthly += 1
                                if(fixSurvey == 'None'):
                                    print(str(serviceMix))
                        surveyStep += 1
                if(mixValue != assessValue or pastmixValue != mixValue):
                    countAllOther = (countOverall) - (countMonthly + countService + countIndirect)
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        asianFemaleOther = (asianFemaleOverall) - (asianFemaleMonthly + asianFemaleService + asianFemaleIndirect)
                        if(asianFemaleOther <= 0):
                            asianFemaleOther = 0
                    elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        asianMaleOther = (asianMaleOverall) - (asianMaleMonthly + asianMaleService + asianMaleIndirect)
                        if(asianMaleOther <= 0):
                            asianMaleOther = 0
                    elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        blackFemaleOther = (blackFemaleOverall) - (blackFemaleMonthly + blackFemaleService + blackFemaleIndirect)
                        if(blackFemaleOther <= 0):
                            blackFemaleOther = 0
                    elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        blackMaleOther = (blackMaleOverall) - (blackMaleMonthly + blackMaleService + blackMaleIndirect)
                        if(blackMaleOther <= 0):
                            blackMaleOther = 0
                    elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        hispanicFemaleOther = (hispanicFemaleOverall) - (hispanicFemaleMonthly + hispanicFemaleService + hispanicFemaleIndirect)
                        if(hispanicFemaleOther <= 0):
                            hispanicFemaleOther = 0
                    elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        hispanicMaleOther = (hispanicMaleOverall) - (hispanicMaleMonthly + hispanicMaleService + hispanicMaleIndirect)
                        if(hispanicMaleOther <= 0):
                            hispanicMaleOther = 0
                    elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        nativeFemaleOther = (nativeFemaleOverall) - (nativeFemaleMonthly + nativeFemaleService + nativeFemaleIndirect)
                        if(nativeFemaleOther <= 0):
                            nativeFemaleOther = 0
                    elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        nativeMaleOther = (nativeMaleOverall) - (nativeMaleMonthly + nativeMaleService + nativeMaleIndirect)
                        if(nativeMaleOther <= 0):
                            nativeMaleOther = 0
                    elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):    
                        whiteFemaleOther = (whiteFemaleOverall) - (whiteFemaleMonthly + whiteFemaleService + whiteFemaleIndirect)
                        if(hispanicMaleOther <= 0):
                            hispanicMaleOther = 0
                    elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):    
                        whiteMaleOther = (whiteMaleOverall) - (whiteMaleMonthly + whiteMaleService + whiteMaleIndirect)
                        if(hispanicMaleOther <= 0):
                            hispanicMaleOther = 0
                    elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        pacificFemaleOther = (pacificFemaleOverall) - (pacificFemaleMonthly + pacificFemaleService + pacificFemaleIndirect)
                        if(pacificFemaleOther <= 0):
                            pacificFemaleOther = 0
                    elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        pacificMaleOther = (pacificMaleOverall) - (pacificMaleMonthly + pacificMaleService + pacificMaleIndirect)
                        if(pacificMaleOther <= 0):
                            pacificMaleOther = 0

                    if("Absences" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ABF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ABM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                AWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                APF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                APM += 1
                    if("Academic Readiness" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                ARPM += 1
                    if("Basic Needs" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                BNPM += 1
                    if("Classroom Conduct" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCPM += 1
                    if("Classroom Participation" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CPPM += 1
                    if("College and Career Readiness" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                CCRPM += 1
                    if("Delinquent Conduct" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCPM += 1
                    if("Family Conflict" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                DCPM += 1
                    if("Grades" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GPM += 1
                    if("Grief/Loss" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                GPM += 1
                    if("Homework Completion" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                HCPM += 1
                    if("Lanuage Development" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LDPM += 1
                    if("Life Skills" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                LSPM += 1
                    if("Mental Health Crisis" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHCPM += 1
                    if("Mental Health and Wellness" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                MHWPM += 1
                    if("Re-Engagement In School" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                IM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RISPM += 1
                    if("Responsible Decision Making" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                RDMPM += 1
                    if("School Engagement and Mtvn" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEMPM += 1
                    if("Self Esteem" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SEPM += 1
                    if("Self Regulation" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SRPM += 1
                    if("Social Skills" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                SSPM += 1
                    if("Tardies" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                THF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                THM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TPM += 1
                    if("Test Readiness" in str(Assess.Range('H'+str(assessStep)).Value)):
                        if("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRAF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRAM += 1
                        elif("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRBF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRBM += 1
                        elif("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRHF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRHM += 1
                        elif("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRIF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRIM += 1
                        elif("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRWF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRWM += 1
                        elif("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                            if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRPF += 1
                            elif("M" in str(Mix.Range('G'+str(mixStep)).Value)):
                                TRPM += 1
                        
                if(service == None):
                    if(flip == 0):
                        SaveChanges=True
                        Mix.Range("A"+str(mixStep+1)+":K"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("L"+str(mixStep+1)+":N"+str(mixStep+1)).Borders(8).LineStyle = 8
                        Mix.Range("O"+str(mixStep+1)+":R"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("O"+str(mixStep+1)+":O"+str(mixStep+6)).Borders(7).LineStyle = 1
                        Mix.Range("R"+str(mixStep+1)+":R"+str(mixStep+6)).Borders(10).LineStyle = 1
                        Mix.Range("O"+str(mixStep+6)+":R"+str(mixStep+6)).Borders(9).LineStyle = 1
                        serviceList = str(Assess.Range('I'+str(assessStep)).Value)
                        serviceList = serviceList.replace(".0","")
                        serviceList = serviceList.replace("10191801","")
                        serviceList = serviceList.replace("1019","")
                        serviceList = serviceList.replace("1009","")
                        Mix.Range('O'+str(mixStep+1)).Value = serviceList
                        Mix.Range('O'+str(mixStep+2)).Value = "Monthly Check In"
                        Mix.Range('O'+str(mixStep+3)).Value = "Planned Service"
                        Mix.Range('O'+str(mixStep+4)).Value = "Indirect"
                        Mix.Range('O'+str(mixStep+5)).Value = "All Other"
                        Mix.Range('O'+str(mixStep+6)).Value = "Percent"
                        Mix.Range('Q'+str(mixStep+1)).Value = "Calculated Data"
                        Mix.Range('Q'+str(mixStep+2)).Value = "Months"
                        Mix.Range('Q'+str(mixStep+3)).Value = "Weight"
                        Mix.Range('Q'+str(mixStep+4)).Value = "Planned Service"
                        Mix.Range('Q'+str(mixStep+5)).Value = "Monthly Check In"
                        Mix.Range('Q'+str(mixStep+6)).Value = "All Other"
                        Mix.Range('P'+str(mixStep+2)).Value = "="+str(countMonthly)+"/"+str(countOverall)
                        Mix.Range('P'+str(mixStep+2)).NumberFormat = "0.00%"
                        Mix.Range('P'+str(mixStep+3)).Value = "="+str(countService)+"/"+str(countOverall)
                        Mix.Range('P'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('P'+str(mixStep+4)).Value = "="+str(countIndirect)+"/"+str(countOverall)
                        Mix.Range('P'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('P'+str(mixStep+5)).Value = "="+str(countAllOther)+"/"+str(countOverall)
                        Mix.Range('P'+str(mixStep+5)).NumberFormat = "0.00%"
                        if(countOverall != 0):
                            total = (countMonthly+countService+countIndirect+countAllOther)/(countOverall)
                        Mix.Range('P'+str(mixStep+6)).Value = total
                        Mix.Range('P'+str(mixStep+6)).NumberFormat = "0.00%"
                        Mix.Range('R'+str(mixStep+2)).Value = datePart1 + str(lowMix) + datePart2 + str(highMix) + datePart3
                        Mix.Range('R'+str(mixStep+3)).Value = "=SUM("+str(countMonthly)+"/"+str(Mix.Range('R'+str(mixStep+2)).Value)+")"
                        Mix.Range('R'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('R'+str(mixStep+4)).Value = "=("+str(countService)+"/"+str(Mix.Range('R'+str(mixStep+2)).Value)+")"#+"+str(Mix.Range('O'+str(mixStep+2)).Value)
                        percentCheck1 = (Mix.Range('R'+str(mixStep+4)).Value)
                        if (percentLimit <= percentCheck1):                                                                                                                                                                                            #Percent Location
                            sumServiceTracker= sumServiceTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumServiceTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumServiceTrackerFemale += 1
                        Mix.Range('R'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('R'+str(mixStep+5)).Value = "=("+str(countMonthly)+"/"+str(Mix.Range('R'+str(mixStep+2)).Value)+")";
                        percentCheck2= (Mix.Range('R'+str(mixStep+4)).Value)
                        if (percentLimit <= percentCheck2):
                            sumMonthlyTracker= sumMonthlyTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumMonthlyTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumMonthlyTrackerFemale += 1

                        if(percentCheck1 >= percentLimit and percentCheck2 >= percentLimit):
                            if('4037' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMSecondaryFull += 1

                            if('4038' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMPreparationFull += 1

                            if('4031' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMExplorationFull += 1

                            if('4012' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMEmploymentFull += 1

                            if('4040' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMFAFSAFull += 1

                            if('4037' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFSecondaryFull += 1

                            if('4038' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFPreparationFull += 1

                            if('4031' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFExplorationFull += 1

                            if('4012' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFEmploymentFull += 1

                            if('4040' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFFAFSAFull += 1
                                    
                            if ("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                                asianServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    asianMaleServiced += 1
                                    maleServiced += 1
                            if ("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                                blackServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    blackMaleServiced += 1
                                    maleServiced += 1
                            if ("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                                hispanicServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    hispanicMaleServiced += 1
                                    maleServiced += 1
                            if ("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                                nativeServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    nativeMaleServiced += 1
                                    maleServiced += 1
                            if ("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                                whiteServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    whiteFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    whiteMaleServiced += 1
                                    maleServiced += 1
                            if ("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                                pacificServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    pacificMaleServiced += 1
                                    maleServiced += 1
                                    
                            if ("Academics" in str(Mix.Range('I'+str(mixStep)).Value)):
                                academicServiced += 1
                            if ("Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                                attendanceServiced += 1
                            if ("Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                                behaviorServiced += 1
                            if ("Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                                socialServiced += 1

                        if (str(Mix.Range('G'+str(mixStep)).Value) == "F"):
                                female += 1
                        if (str(Mix.Range('G'+str(mixStep)).Value) == "M"):
                                male += 1
                                
                        Mix.Range('R'+str(mixStep+5)).NumberFormat = "0.00%"
                        Mix.Range('R'+str(mixStep+6)).Value = "=("+str(countAllOther)+"/"+str(Mix.Range('R'+str(mixStep+2)).Value)+")";
                        Mix.Range('R'+str(mixStep+6)).NumberFormat = "0.00%"
                        flip = 1
                    elif(flip == 1):
                        SaveChanges=True
                        Mix.Range("A"+str(mixStep+1)+":K"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("L"+str(mixStep+1)+":S"+str(mixStep+1)).Borders(8).LineStyle = 8
                        Mix.Range("W"+str(mixStep+1)+":T"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("T"+str(mixStep+1)+":T"+str(mixStep+6)).Borders(7).LineStyle = 1
                        Mix.Range("W"+str(mixStep+1)+":W"+str(mixStep+6)).Borders(10).LineStyle = 1
                        Mix.Range("T"+str(mixStep+6)+":W"+str(mixStep+6)).Borders(9).LineStyle = 1
                        serviceList = str(Assess.Range('I'+str(assessStep)).Value)
                        serviceList = serviceList.replace(".0","")
                        serviceList = serviceList.replace("10191801","")
                        serviceList = serviceList.replace("1019","")
                        serviceList = serviceList.replace("1009","")
                        Mix.Range('T'+str(mixStep+1)).Value = serviceList
                        Mix.Range('T'+str(mixStep+2)).Value = "Monthly Check In"
                        Mix.Range('T'+str(mixStep+3)).Value = "Planned Service"
                        Mix.Range('T'+str(mixStep+4)).Value = "Indirect"
                        Mix.Range('T'+str(mixStep+5)).Value = "All Other"
                        Mix.Range('T'+str(mixStep+6)).Value = "Percent"
                        Mix.Range('V'+str(mixStep+1)).Value = "Calculated Data"
                        Mix.Range('V'+str(mixStep+2)).Value = "Months"
                        Mix.Range('V'+str(mixStep+3)).Value = "Weight"
                        Mix.Range('V'+str(mixStep+4)).Value = "Planned Service"
                        Mix.Range('V'+str(mixStep+5)).Value = "Monthly Check In"
                        Mix.Range('V'+str(mixStep+6)).Value = "All Other"
                        Mix.Range('U'+str(mixStep+2)).Value = "="+str(countMonthly)+"/"+str(countOverall)
                        Mix.Range('U'+str(mixStep+2)).NumberFormat = "0.00%"
                        Mix.Range('U'+str(mixStep+3)).Value = "="+str(countService)+"/"+str(countOverall)
                        Mix.Range('U'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('U'+str(mixStep+4)).Value = "="+str(countIndirect)+"/"+str(countOverall)
                        Mix.Range('U'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('U'+str(mixStep+5)).Value = "="+str(countAllOther)+"/"+str(countOverall)
                        Mix.Range('U'+str(mixStep+5)).NumberFormat = "0.00%"
                        if(countOverall != 0):
                            total = (countMonthly+countService+countIndirect+countAllOther)/(countOverall)
                        Mix.Range('U'+str(mixStep+6)).Value = total
                        Mix.Range('U'+str(mixStep+6)).NumberFormat = "0.00%"
                        Mix.Range('W'+str(mixStep+2)).Value = datePart1 + str(lowMix) + datePart2 + str(highMix) + datePart3
                        Mix.Range('W'+str(mixStep+3)).Value = "=SUM("+str(countMonthly)+"/"+str(Mix.Range('W'+str(mixStep+2)).Value)+")"
                        Mix.Range('W'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('W'+str(mixStep+4)).Value = "=("+str(countService)+"/"+str(Mix.Range('W'+str(mixStep+2)).Value)+")"#+"+str(Mix.Range('T'+str(mixStep+2)).Value)
                        percentCheck1 = int(Mix.Range('W'+str(mixStep+4)).Value)
                        if (percentLimit <= percentCheck1):                                                                                                                                                                                            #Percent Location
                            sumServiceTracker= sumServiceTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumServiceTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumServiceTrackerFemale += 1
                        Mix.Range('W'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('W'+str(mixStep+5)).Value = "=("+str(countMonthly)+"/"+str(Mix.Range('W'+str(mixStep+2)).Value)+")";
                        percentCheck2= int(Mix.Range('W'+str(mixStep+5)).Value)
                        if (percentLimit <= percentCheck2):
                            sumMonthlyTracker= sumMonthlyTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumMonthlyTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumMonthlyTrackerFemale += 1

                        if(percentCheck1 >= percentLimit and percentCheck2 >= percentLimit):
                            if('4037' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMSecondaryFull += 1

                            if('4038' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMPreparationFull += 1

                            if('4031' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMExplorationFull += 1

                            if('4012' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMEmploymentFull += 1

                            if('4040' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMFAFSAFull += 1

                            if('4037' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFSecondaryFull += 1

                            if('4038' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFPreparationFull += 1

                            if('4031' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFExplorationFull += 1

                            if('4012' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFEmploymentFull += 1

                            if('4040' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFFAFSAFull += 1
                                    
                            if ("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                                asianServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    asianMaleServiced += 1
                                    maleServiced += 1
                            if ("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                                blackServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    blackMaleServiced += 1
                                    maleServiced += 1
                            if ("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                                hispanicServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    hispanicMaleServiced += 1
                                    maleServiced += 1
                            if ("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                                nativeServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    nativeMaleServiced += 1
                                    maleServiced += 1
                            if ("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                                whiteServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    whiteFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    whiteMaleServiced += 1
                                    maleServiced += 1
                            if ("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                                pacificServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    pacificMaleServiced += 1
                                    maleServiced += 1
                                    
                            if ("Academics" in str(Mix.Range('I'+str(mixStep)).Value)):
                                academicServiced += 1
                            if ("Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                                attendanceServiced += 1
                            if ("Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                                behaviorServiced += 1
                            if ("Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                                socialServiced += 1

                        if (str(Mix.Range('G'+str(mixStep)).Value) == "F"):
                                female += 1
                        if (str(Mix.Range('G'+str(mixStep)).Value) == "M"):
                                male += 1
                                
                        Mix.Range('W'+str(mixStep+5)).NumberFormat = "0.00%"
                        Mix.Range('W'+str(mixStep+6)).Value = "=("+str(countAllOther)+"/"+str(Mix.Range('W'+str(mixStep+2)).Value)+")";
                        Mix.Range('W'+str(mixStep+6)).NumberFormat = "0.00%"
                        flip = 2
                    elif(flip == 2):
                        SaveChanges=True
                        Mix.Range("A"+str(mixStep+1)+":K"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("L"+str(mixStep+1)+":X"+str(mixStep+1)).Borders(8).LineStyle = 8
                        Mix.Range("Y"+str(mixStep+1)+":AB"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("Y"+str(mixStep+1)+":Y"+str(mixStep+6)).Borders(7).LineStyle = 1
                        Mix.Range("AB"+str(mixStep+1)+":AB"+str(mixStep+6)).Borders(10).LineStyle = 1
                        Mix.Range("Y"+str(mixStep+6)+":AB"+str(mixStep+6)).Borders(9).LineStyle = 1
                        serviceList = str(Assess.Range('I'+str(assessStep)).Value)
                        serviceList = serviceList.replace(".0","")
                        serviceList = serviceList.replace("10191801","")
                        serviceList = serviceList.replace("1019","")
                        serviceList = serviceList.replace("1009","")
                        Mix.Range('Y'+str(mixStep+1)).Value = serviceList
                        Mix.Range('Y'+str(mixStep+2)).Value = "Monthly Check In"
                        Mix.Range('Y'+str(mixStep+3)).Value = "Planned Service"
                        Mix.Range('Y'+str(mixStep+4)).Value = "Indirect"
                        Mix.Range('Y'+str(mixStep+5)).Value = "All Other"
                        Mix.Range('Y'+str(mixStep+6)).Value = "Percent"
                        Mix.Range('AA'+str(mixStep+1)).Value = "Calculated Data"
                        Mix.Range('AA'+str(mixStep+2)).Value = "Months"
                        Mix.Range('AA'+str(mixStep+3)).Value = "Weight"
                        Mix.Range('AA'+str(mixStep+4)).Value = "Planned Service"
                        Mix.Range('AA'+str(mixStep+5)).Value = "Monthly Check In"
                        Mix.Range('AA'+str(mixStep+6)).Value = "All Other"
                        Mix.Range('Z'+str(mixStep+2)).Value = "="+str(countMonthly)+"/"+str(countOverall)
                        Mix.Range('Z'+str(mixStep+2)).NumberFormat = "0.00%"
                        Mix.Range('Z'+str(mixStep+3)).Value = "="+str(countService)+"/"+str(countOverall)
                        Mix.Range('Z'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('Z'+str(mixStep+4)).Value = "="+str(countIndirect)+"/"+str(countOverall)
                        Mix.Range('Z'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('Z'+str(mixStep+5)).Value = "="+str(countAllOther)+"/"+str(countOverall)
                        Mix.Range('Z'+str(mixStep+5)).NumberFormat = "0.00%"
                        if(countOverall != 0):
                            total = (countMonthly+countService+countIndirect+countAllOther)/(countOverall)
                        Mix.Range('Z'+str(mixStep+6)).Value = total
                        Mix.Range('Z'+str(mixStep+6)).NumberFormat = "0.00%"
                        Mix.Range('AB'+str(mixStep+2)).Value = datePart1 + str(lowMix) + datePart2 + str(highMix) + datePart3
                        Mix.Range('AB'+str(mixStep+3)).Value = "=SUM("+str(countMonthly)+"/"+str(Mix.Range('AB'+str(mixStep+2)).Value)+")"
                        Mix.Range('AB'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('AB'+str(mixStep+4)).Value = "=("+str(countService)+"/"+str(Mix.Range('AB'+str(mixStep+2)).Value)+")"#+"+str(Mix.Range('Y'+str(mixStep+2)).Value)
                        percentCheck1 = int(Mix.Range('AB'+str(mixStep+4)).Value)
                        if (percentLimit <= percentCheck1):                                                                                                                                                                                            #Percent Location
                            sumServiceTracker= sumServiceTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumServiceTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumServiceTrackerFemale += 1
                        Mix.Range('AB'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('AB'+str(mixStep+5)).Value = "=("+str(countMonthly)+"/"+str(Mix.Range('AB'+str(mixStep+2)).Value)+")";
                        percentCheck2= int(Mix.Range('AB'+str(mixStep+5)).Value)
                        if (percentLimit <= percentCheck2):
                            sumMonthlyTracker= sumMonthlyTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumMonthlyTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumMonthlyTrackerFemale += 1

                        if(percentCheck1 >= percentLimit and percentCheck2 >= percentLimit):
                            if('4037' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMSecondaryFull += 1

                            if('4038' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMPreparationFull += 1

                            if('4031' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMExplorationFull += 1

                            if('4012' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMEmploymentFull += 1

                            if('4040' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMFAFSAFull += 1

                            if('4037' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFSecondaryFull += 1

                            if('4038' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFPreparationFull += 1

                            if('4031' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFExplorationFull += 1

                            if('4012' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFEmploymentFull += 1

                            if('4040' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFFAFSAFull += 1
                                
                            if ("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                                asianServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    asianMaleServiced += 1
                                    maleServiced += 1
                            if ("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                                blackServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    blackMaleServiced += 1
                                    maleServiced += 1
                            if ("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                                hispanicServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    hispanicMaleServiced += 1
                                    maleServiced += 1
                            if ("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                                nativeServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    nativeMaleServiced += 1
                                    maleServiced += 1
                            if ("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                                whiteServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    whiteFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    whiteMaleServiced += 1
                                    maleServiced += 1
                            if ("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                                pacificServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    pacificMaleServiced += 1
                                    maleServiced += 1
                                    
                            if ("Academics" in str(Mix.Range('I'+str(mixStep)).Value)):
                                academicServiced += 1
                            if ("Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                                attendanceServiced += 1
                            if ("Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                                behaviorServiced += 1
                            if ("Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                                socialServiced += 1

                        if (str(Mix.Range('G'+str(mixStep)).Value) == "F"):
                                female += 1
                        if (str(Mix.Range('G'+str(mixStep)).Value) == "M"):
                                male += 1
                                
                        Mix.Range('AB'+str(mixStep+5)).NumberFormat = "0.00%"
                        Mix.Range('AB'+str(mixStep+6)).Value = "=("+str(countAllOther)+"/"+str(Mix.Range('AB'+str(mixStep+2)).Value)+")";
                        Mix.Range('AB'+str(mixStep+6)).NumberFormat = "0.00%"
                        flip = 3
                    elif(flip == 3):
                        SaveChanges=True
                        Mix.Range("A"+str(mixStep+1)+":K"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("L"+str(mixStep+1)+":AC"+str(mixStep+1)).Borders(8).LineStyle = 8
                        Mix.Range("AD"+str(mixStep+1)+":AG"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("AD"+str(mixStep+1)+":AD"+str(mixStep+6)).Borders(7).LineStyle = 1
                        Mix.Range("AG"+str(mixStep+1)+":AG"+str(mixStep+6)).Borders(10).LineStyle = 1
                        Mix.Range("AD"+str(mixStep+6)+":AG"+str(mixStep+6)).Borders(9).LineStyle = 1
                        serviceList = str(Assess.Range('I'+str(assessStep)).Value)
                        serviceList = serviceList.replace(".0","")
                        serviceList = serviceList.replace("10191801","")
                        serviceList = serviceList.replace("1019","")
                        serviceList = serviceList.replace("1009","")
                        Mix.Range('AD'+str(mixStep+1)).Value = serviceList
                        Mix.Range('AD'+str(mixStep+2)).Value = "Monthly Check In"
                        Mix.Range('AD'+str(mixStep+3)).Value = "Planned Service"
                        Mix.Range('AD'+str(mixStep+4)).Value = "Indirect"
                        Mix.Range('AD'+str(mixStep+5)).Value = "All Other"
                        Mix.Range('AD'+str(mixStep+6)).Value = "Percent"
                        Mix.Range('AF'+str(mixStep+1)).Value = "Calculated Data"
                        Mix.Range('AF'+str(mixStep+2)).Value = "Months"
                        Mix.Range('AF'+str(mixStep+3)).Value = "Weight"
                        Mix.Range('AF'+str(mixStep+4)).Value = "Planned Service"
                        Mix.Range('AF'+str(mixStep+5)).Value = "Monthly Check In"
                        Mix.Range('AF'+str(mixStep+6)).Value = "All Other"
                        Mix.Range('AE'+str(mixStep+2)).Value = "="+str(countMonthly)+"/"+str(countOverall)
                        Mix.Range('AE'+str(mixStep+2)).NumberFormat = "0.00%"
                        Mix.Range('AE'+str(mixStep+3)).Value = "="+str(countService)+"/"+str(countOverall)
                        Mix.Range('AE'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('AE'+str(mixStep+4)).Value = "="+str(countIndirect)+"/"+str(countOverall)
                        Mix.Range('AE'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('AE'+str(mixStep+5)).Value = "="+str(countAllOther)+"/"+str(countOverall)
                        Mix.Range('AE'+str(mixStep+5)).NumberFormat = "0.00%"
                        if(countOverall != 0):
                            total = (countMonthly+countService+countIndirect+countAllOther)/(countOverall)
                        Mix.Range('AE'+str(mixStep+6)).Value = total
                        Mix.Range('AE'+str(mixStep+6)).NumberFormat = "0.00%"
                        Mix.Range('AG'+str(mixStep+2)).Value = datePart1 + str(lowMix) + datePart2 + str(highMix) + datePart3
                        Mix.Range('AG'+str(mixStep+3)).Value = "=SUM("+str(countMonthly)+"/"+str(Mix.Range('AG'+str(mixStep+2)).Value)+")"
                        Mix.Range('AG'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('AG'+str(mixStep+4)).Value = "=("+str(countService)+"/"+str(Mix.Range('AG'+str(mixStep+2)).Value)+")"#+"+str(Mix.Range('Y'+str(mixStep+2)).Value)
                        percentCheck1= int(Mix.Range('AG'+str(mixStep+4)).Value)
                        if (percentLimit <= percentCheck1):                                                                                                                                                                                            #Percent Location
                            sumServiceTracker= sumServiceTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumServiceTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumServiceTrackerFemale += 1
                        Mix.Range('AG'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('AG'+str(mixStep+5)).Value = "=("+str(countMonthly)+"/"+str(Mix.Range('AG'+str(mixStep+2)).Value)+")";
                        percentCheck2= int(Mix.Range('AG'+str(mixStep+5)).Value)
                        if (percentLimit <= percentCheck2):
                            sumMonthlyTracker= sumMonthlyTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumMonthlyTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumMonthlyTrackerFemale += 1
                                
                        if(percentCheck1 >= percentLimit and percentCheck2 >= percentLimit):
                            if('4037' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMSecondaryFull += 1

                            if('4038' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMPreparationFull += 1

                            if('4031' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMExplorationFull += 1

                            if('4012' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMEmploymentFull += 1

                            if('4040' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMFAFSAFull += 1

                            if('4037' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFSecondaryFull += 1

                            if('4038' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFPreparationFull += 1

                            if('4031' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFExplorationFull += 1

                            if('4012' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFEmploymentFull += 1

                            if('4040' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFFAFSAFull += 1
                                    
                            if ("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                                asianServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    asianMaleServiced += 1
                                    maleServiced += 1
                            if ("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                                blackServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    blackMaleServiced += 1
                                    maleServiced += 1
                            if ("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                                hispanicServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    hispanicMaleServiced += 1
                                    maleServiced += 1
                            if ("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                                nativeServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    nativeMaleServiced += 1
                                    maleServiced += 1
                            if ("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                                whiteServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    whiteFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    whiteMaleServiced += 1
                                    maleServiced += 1
                            if ("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                                pacificServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    pacificMaleServiced += 1
                                    maleServiced += 1
                                    
                            if ("Academics" in str(Mix.Range('I'+str(mixStep)).Value)):
                                academicServiced += 1
                            if ("Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                                attendanceServiced += 1
                            if ("Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                                behaviorServiced += 1
                            if ("Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                                socialServiced += 1

                        if (str(Mix.Range('G'+str(mixStep)).Value) == "F"):
                                female += 1
                        if (str(Mix.Range('G'+str(mixStep)).Value) == "M"):
                                male += 1
                                
                        Mix.Range('AG'+str(mixStep+5)).NumberFormat = "0.00%"
                        Mix.Range('AG'+str(mixStep+6)).Value = "=("+str(countAllOther)+"/"+str(Mix.Range('AG'+str(mixStep+2)).Value)+")";
                        Mix.Range('AG'+str(mixStep+6)).NumberFormat = "0.00%"
                        flip = 0
                        

                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        asianAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianFemaleAcademics += 1
                        else:
                            asianMaleAcademics += 1
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        asianAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianFemaleAttendance += 1
                        else:
                            asianMaleAttendance += 1
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        asianBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianFemaleBehavioral += 1
                        else:
                            asianMaleBehavioral += 1
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        asianSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianFemaleSocial += 1
                        else:
                            asianMaleSocial += 1
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        asianFemale += 1
                    elif ("A" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        asianMale += 1



                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        blackAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackFemaleAcademics += 1
                        else:
                            blackMaleAcademics += 1
                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        blackAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackFemaleAttendance += 1
                        else:
                            blackMaleAttendance += 1
                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        blackBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackFemaleBehavioral += 1
                        else:
                            blackMaleBehavioral += 1
                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        blackSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackFemaleSocial += 1
                        else:
                            blackMaleSocial += 1
                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        blackFemale += 1
                    elif ("B" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        blackMale += 1


                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        hispanicAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicFemaleAcademics += 1
                        else:
                            hispanicMaleAcademics += 1
                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        hispanicAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicFemaleAttendance += 1
                        else:
                            hispanicMaleAttendance += 1
                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        hispanicBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicFemaleBehavioral += 1
                        else:
                            hispanicMaleBehavioral += 1
                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        hispanicSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicFemaleSocial += 1
                        else:
                            hispanicMaleSocial += 1
                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        hispanicFemale += 1
                    elif ("H" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        hispanicMale += 1


                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        nativeAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeFemaleAcademics += 1
                        else:
                            nativeMaleAcademics += 1
                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        nativeAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeFemaleAttendance += 1
                        else:
                            nativeMaleAttendance += 1
                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        nativeBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeFemaleBehavioral += 1
                        else:
                            nativeMaleBehavioral += 1
                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        nativeSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeFemaleSocial += 1
                        else:
                            nativeMaleSocial += 1
                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        nativeFemale += 1
                    elif ("I" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        nativeMale += 1


                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        whiteAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            whiteFemaleAcademics += 1
                        else:
                            whiteMaleAcademics += 1
                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        whiteAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            whiteFemaleAttendance += 1
                        else:
                            whiteMaleAttendance += 1
                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        whiteBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            whiteFemaleBehavioral += 1
                        else:
                            whiteMaleBehavioral += 1
                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        whiteSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            whiteFemaleSocial += 1
                        else:
                            whiteMaleSocial += 1
                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        whiteFemale += 1
                    elif ("W" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        whiteMale += 1

                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        pacificAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificFemaleAcademics += 1
                        else:
                            pacificMaleAcademics += 1
                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        pacificAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificFemaleAttendance += 1
                        else:
                            pacificMaleAttendance += 1
                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        pacificBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificFemaleBehavioral += 1
                        else:
                            pacificMaleBehavioral += 1
                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        pacificSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificFemaleSocial += 1
                        else:
                            pacificMaleSocial += 1
                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        pacificFemale += 1
                    elif ("P" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        pacificMale += 1
                            
                    countServiceTracker += countService
                    countMonthlyTracker += countMonthly
                    countIndirectTracker += countIndirect
                    countAllOtherTracker += countAllOther
                    countMonthly = 0
                    countService = 0
                    countIndirect = 0
                    countAllOther = 0
                    countOverall = 0
                    switchSex = 0
                    highDate = str(Mix.Range('J'+str(mixStep)).Value);
                    lowDate = str(Mix.Range('J'+str(mixStep)).Value);
                    compareDate = str(Mix.Range('J'+str(mixStep)).Value);
                elif(mixValue != assessValue or pastmixValue != mixValue):
                    #print("---New Student---")
                    if(flip == 0):
                        SaveChanges=True
                        Mix.Range("A"+str(mixStep+1)+":K"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("L"+str(mixStep+1)+":N"+str(mixStep+1)).Borders(8).LineStyle = 8
                        Mix.Range("O"+str(mixStep+1)+":R"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("O"+str(mixStep+1)+":O"+str(mixStep+6)).Borders(7).LineStyle = 1
                        Mix.Range("R"+str(mixStep+1)+":R"+str(mixStep+6)).Borders(10).LineStyle = 1
                        Mix.Range("O"+str(mixStep+6)+":R"+str(mixStep+6)).Borders(9).LineStyle = 1
                        serviceList = str(Assess.Range('I'+str(assessStep)).Value)
                        serviceList = serviceList.replace(".0","")
                        serviceList = serviceList.replace("10191801","")
                        serviceList = serviceList.replace("1019","")
                        serviceList = serviceList.replace("1009","")
                        Mix.Range('O'+str(mixStep+1)).Value = serviceList
                        Mix.Range('O'+str(mixStep+2)).Value = "Monthly Check In"
                        Mix.Range('O'+str(mixStep+3)).Value = "Planned Service"
                        Mix.Range('O'+str(mixStep+4)).Value = "Indirect"
                        Mix.Range('O'+str(mixStep+5)).Value = "All Other"
                        Mix.Range('O'+str(mixStep+6)).Value = "Percent"
                        Mix.Range('Q'+str(mixStep+1)).Value = "Calculated Data"
                        Mix.Range('Q'+str(mixStep+2)).Value = "Months"
                        Mix.Range('Q'+str(mixStep+3)).Value = "Weight"
                        Mix.Range('Q'+str(mixStep+4)).Value = "Planned Service"
                        Mix.Range('Q'+str(mixStep+5)).Value = "Monthly Check In"
                        Mix.Range('Q'+str(mixStep+6)).Value = "All Other"
                        Mix.Range('P'+str(mixStep+2)).Value = "="+str(countMonthly)+"/"+str(countOverall)
                        Mix.Range('P'+str(mixStep+2)).NumberFormat = "0.00%"
                        Mix.Range('P'+str(mixStep+3)).Value = "="+str(countService)+"/"+str(countOverall)
                        Mix.Range('P'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('P'+str(mixStep+4)).Value = "="+str(countIndirect)+"/"+str(countOverall)
                        Mix.Range('P'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('P'+str(mixStep+5)).Value = "="+str(countAllOther)+"/"+str(countOverall)
                        Mix.Range('P'+str(mixStep+5)).NumberFormat = "0.00%"
                        if(countOverall != 0):
                            total = (countMonthly+countService+countIndirect+countAllOther)/(countOverall)
                        Mix.Range('P'+str(mixStep+6)).Value = total
                        Mix.Range('P'+str(mixStep+6)).NumberFormat = "0.00%"
                        Mix.Range('R'+str(mixStep+2)).Value = datePart1 + str(lowMix) + datePart2 + str(highMix) + datePart3
                        Mix.Range('R'+str(mixStep+3)).Value = "=SUM("+str(countMonthly)+"/"+str(Mix.Range('R'+str(mixStep+2)).Value)+")"
                        Mix.Range('R'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('R'+str(mixStep+4)).Value = "=("+str(countService)+"/"+str(Mix.Range('R'+str(mixStep+2)).Value)+")"#+"+str(Mix.Range('O'+str(mixStep+2)).Value)
                        percentCheck1 = (Mix.Range('R'+str(mixStep+4)).Value)
                        if (percentLimit <= percentCheck1):                                                                                                                                                                                            #Percent Location
                            sumServiceTracker= sumServiceTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumServiceTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumServiceTrackerFemale += 1
                        Mix.Range('R'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('R'+str(mixStep+5)).Value = "=("+str(countMonthly)+"/"+str(Mix.Range('R'+str(mixStep+2)).Value)+")";
                        percentCheck2= (Mix.Range('R'+str(mixStep+5)).Value)
                        if (percentLimit <= percentCheck2):
                            sumMonthlyTracker= sumMonthlyTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumMonthlyTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumMonthlyTrackerFemale += 1

                        if(percentCheck1 >= percentLimit and percentCheck2 >= percentLimit):
                            if('4037' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMSecondaryFull += 1

                            if('4038' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMPreparationFull += 1

                            if('4031' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMExplorationFull += 1

                            if('4012' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMEmploymentFull += 1

                            if('4040' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMFAFSAFull += 1

                            if('4037' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFSecondaryFull += 1

                            if('4038' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFPreparationFull += 1

                            if('4031' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFExplorationFull += 1

                            if('4012' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFEmploymentFull += 1

                            if('4040' in str(Mix.Range('O'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFFAFSAFull += 1
                                    
                            if ("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                                asianServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    asianMaleServiced += 1
                                    maleServiced += 1
                            if ("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                                blackServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    blackMaleServiced += 1
                                    maleServiced += 1
                            if ("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                                hispanicServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    hispanicMaleServiced += 1
                                    maleServiced += 1
                            if ("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                                nativeServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    nativeMaleServiced += 1
                                    maleServiced += 1
                            if ("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                                whiteServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    whiteFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    whiteMaleServiced += 1
                                    maleServiced += 1
                            if ("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                                pacificServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    pacificMaleServiced += 1
                                    maleServiced += 1
                                    
                            if ("Academics" in str(Mix.Range('I'+str(mixStep)).Value)):
                                academicServiced += 1
                            if ("Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                                attendanceServiced += 1
                            if ("Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                                behaviorServiced += 1
                            if ("Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                                socialServiced += 1

                        if (str(Mix.Range('G'+str(mixStep)).Value) == "F"):
                                female += 1
                        if (str(Mix.Range('G'+str(mixStep)).Value) == "M"):
                                male += 1
                                
                        Mix.Range('R'+str(mixStep+5)).NumberFormat = "0.00%"
                        Mix.Range('R'+str(mixStep+6)).Value = "=("+str(countAllOther)+"/"+str(Mix.Range('R'+str(mixStep+2)).Value)+")";
                        Mix.Range('R'+str(mixStep+6)).NumberFormat = "0.00%"
                        flip = 1
                    elif(flip == 1):
                        SaveChanges=True
                        Mix.Range("A"+str(mixStep+1)+":K"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("L"+str(mixStep+1)+":S"+str(mixStep+1)).Borders(8).LineStyle = 8
                        Mix.Range("W"+str(mixStep+1)+":T"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("T"+str(mixStep+1)+":T"+str(mixStep+6)).Borders(7).LineStyle = 1
                        Mix.Range("W"+str(mixStep+1)+":W"+str(mixStep+6)).Borders(10).LineStyle = 1
                        Mix.Range("T"+str(mixStep+6)+":W"+str(mixStep+6)).Borders(9).LineStyle = 1
                        serviceList = str(Assess.Range('I'+str(assessStep)).Value)
                        serviceList = serviceList.replace(".0","")
                        serviceList = serviceList.replace("10191801","")
                        serviceList = serviceList.replace("1019","")
                        serviceList = serviceList.replace("1009","")
                        Mix.Range('T'+str(mixStep+1)).Value = serviceList
                        Mix.Range('T'+str(mixStep+2)).Value = "Monthly Check In"
                        Mix.Range('T'+str(mixStep+3)).Value = "Planned Service"
                        Mix.Range('T'+str(mixStep+4)).Value = "Indirect"
                        Mix.Range('T'+str(mixStep+5)).Value = "All Other"
                        Mix.Range('T'+str(mixStep+6)).Value = "Percent"
                        Mix.Range('V'+str(mixStep+1)).Value = "Calculated Data"
                        Mix.Range('V'+str(mixStep+2)).Value = "Months"
                        Mix.Range('V'+str(mixStep+3)).Value = "Weight"
                        Mix.Range('V'+str(mixStep+4)).Value = "Planned Service"
                        Mix.Range('V'+str(mixStep+5)).Value = "Monthly Check In"
                        Mix.Range('V'+str(mixStep+6)).Value = "All Other"
                        Mix.Range('U'+str(mixStep+2)).Value = "="+str(countMonthly)+"/"+str(countOverall)
                        Mix.Range('U'+str(mixStep+2)).NumberFormat = "0.00%"
                        Mix.Range('U'+str(mixStep+3)).Value = "="+str(countService)+"/"+str(countOverall)
                        Mix.Range('U'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('U'+str(mixStep+4)).Value = "="+str(countIndirect)+"/"+str(countOverall)
                        Mix.Range('U'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('U'+str(mixStep+5)).Value = "="+str(countAllOther)+"/"+str(countOverall)
                        Mix.Range('U'+str(mixStep+5)).NumberFormat = "0.00%"
                        if(countOverall != 0):
                            total = (countMonthly+countService+countIndirect+countAllOther)/(countOverall)
                        Mix.Range('U'+str(mixStep+6)).Value = total
                        Mix.Range('U'+str(mixStep+6)).NumberFormat = "0.00%"
                        Mix.Range('W'+str(mixStep+2)).Value = datePart1 + str(lowMix) + datePart2 + str(highMix) + datePart3
                        Mix.Range('W'+str(mixStep+3)).Value = "=SUM("+str(countMonthly)+"/"+str(Mix.Range('W'+str(mixStep+2)).Value)+")"
                        Mix.Range('W'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('W'+str(mixStep+4)).Value = "=("+str(countService)+"/"+str(Mix.Range('W'+str(mixStep+2)).Value)+")"#+"+str(Mix.Range('T'+str(mixStep+2)).Value)
                        percentCheck1 = int(Mix.Range('W'+str(mixStep+4)).Value)
                        if (percentLimit <= percentCheck1):                                                                                                                                                                                            #Percent Location
                            sumServiceTracker= sumServiceTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumServiceTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumServiceTrackerFemale += 1
                        Mix.Range('W'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('W'+str(mixStep+5)).Value = "=("+str(countMonthly)+"/"+str(Mix.Range('W'+str(mixStep+2)).Value)+")";
                        percentCheck2= int(Mix.Range('W'+str(mixStep+5)).Value)
                        if (percentLimit <= percentCheck2):
                            sumMonthlyTracker= sumMonthlyTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumMonthlyTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumMonthlyTrackerFemale += 1

                        if(percentCheck1 >= percentLimit and percentCheck2 >= percentLimit):
                            if('4037' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMSecondaryFull += 1

                            if('4038' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMPreparationFull += 1

                            if('4031' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMExplorationFull += 1

                            if('4012' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMEmploymentFull += 1

                            if('4040' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMFAFSAFull += 1

                            if('4037' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFSecondaryFull += 1

                            if('4038' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFPreparationFull += 1

                            if('4031' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFExplorationFull += 1

                            if('4012' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFEmploymentFull += 1

                            if('4040' in str(Mix.Range('T'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFFAFSAFull += 1
                                    
                            if ("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                                asianServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    asianMaleServiced += 1
                                    maleServiced += 1
                            if ("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                                blackServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    blackMaleServiced += 1
                                    maleServiced += 1
                            if ("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                                hispanicServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    hispanicMaleServiced += 1
                                    maleServiced += 1
                            if ("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                                nativeServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    nativeMaleServiced += 1
                                    maleServiced += 1
                            if ("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                                whiteServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    whiteFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    whiteMaleServiced += 1
                                    maleServiced += 1
                            if ("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                                pacificServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    pacificMaleServiced += 1
                                    maleServiced += 1
                                    
                            if ("Academics" in str(Mix.Range('I'+str(mixStep)).Value)):
                                academicServiced += 1
                            if ("Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                                attendanceServiced += 1
                            if ("Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                                behaviorServiced += 1
                            if ("Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                                socialServiced += 1

                        if (str(Mix.Range('G'+str(mixStep)).Value) == "F"):
                                female += 1
                        if (str(Mix.Range('G'+str(mixStep)).Value) == "M"):
                                male += 1
                                
                        Mix.Range('W'+str(mixStep+5)).NumberFormat = "0.00%"
                        Mix.Range('W'+str(mixStep+6)).Value = "=("+str(countAllOther)+"/"+str(Mix.Range('W'+str(mixStep+2)).Value)+")";
                        Mix.Range('W'+str(mixStep+6)).NumberFormat = "0.00%"
                        flip = 2
                    elif(flip == 2):
                        SaveChanges=True
                        Mix.Range("A"+str(mixStep+1)+":K"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("L"+str(mixStep+1)+":X"+str(mixStep+1)).Borders(8).LineStyle = 8
                        Mix.Range("Y"+str(mixStep+1)+":AB"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("Y"+str(mixStep+1)+":Y"+str(mixStep+6)).Borders(7).LineStyle = 1
                        Mix.Range("AB"+str(mixStep+1)+":AB"+str(mixStep+6)).Borders(10).LineStyle = 1
                        Mix.Range("Y"+str(mixStep+6)+":AB"+str(mixStep+6)).Borders(9).LineStyle = 1
                        serviceList = str(Assess.Range('I'+str(assessStep)).Value)
                        serviceList = serviceList.replace(".0","")
                        serviceList = serviceList.replace("10191801","")
                        serviceList = serviceList.replace("1019","")
                        serviceList = serviceList.replace("1009","")
                        Mix.Range('Y'+str(mixStep+1)).Value = serviceList
                        Mix.Range('Y'+str(mixStep+2)).Value = "Monthly Check In"
                        Mix.Range('Y'+str(mixStep+3)).Value = "Planned Service"
                        Mix.Range('Y'+str(mixStep+4)).Value = "Indirect"
                        Mix.Range('Y'+str(mixStep+5)).Value = "All Other"
                        Mix.Range('Y'+str(mixStep+6)).Value = "Percent"
                        Mix.Range('AA'+str(mixStep+1)).Value = "Calculated Data"
                        Mix.Range('AA'+str(mixStep+2)).Value = "Months"
                        Mix.Range('AA'+str(mixStep+3)).Value = "Weight"
                        Mix.Range('AA'+str(mixStep+4)).Value = "Planned Service"
                        Mix.Range('AA'+str(mixStep+5)).Value = "Monthly Check In"
                        Mix.Range('AA'+str(mixStep+6)).Value = "All Other"
                        Mix.Range('Z'+str(mixStep+2)).Value = "="+str(countMonthly)+"/"+str(countOverall)
                        Mix.Range('Z'+str(mixStep+2)).NumberFormat = "0.00%"
                        Mix.Range('Z'+str(mixStep+3)).Value = "="+str(countService)+"/"+str(countOverall)
                        Mix.Range('Z'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('Z'+str(mixStep+4)).Value = "="+str(countIndirect)+"/"+str(countOverall)
                        Mix.Range('Z'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('Z'+str(mixStep+5)).Value = "="+str(countAllOther)+"/"+str(countOverall)
                        Mix.Range('Z'+str(mixStep+5)).NumberFormat = "0.00%"
                        if(countOverall != 0):
                            total = (countMonthly+countService+countIndirect+countAllOther)/(countOverall)
                        Mix.Range('Z'+str(mixStep+6)).Value = total
                        Mix.Range('Z'+str(mixStep+6)).NumberFormat = "0.00%"
                        Mix.Range('AB'+str(mixStep+2)).Value = datePart1 + str(lowMix) + datePart2 + str(highMix) + datePart3
                        Mix.Range('AB'+str(mixStep+3)).Value = "=SUM("+str(countMonthly)+"/"+str(Mix.Range('AB'+str(mixStep+2)).Value)+")"
                        Mix.Range('AB'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('AB'+str(mixStep+4)).Value = "=("+str(countService)+"/"+str(Mix.Range('AB'+str(mixStep+2)).Value)+")"#+"+str(Mix.Range('Y'+str(mixStep+2)).Value)
                        percentCheck1 = int(Mix.Range('AB'+str(mixStep+4)).Value)
                        if (percentLimit <= percentCheck1):                                                                                                                                                                                            #Percent Location
                            sumServiceTracker= sumServiceTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumServiceTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumServiceTrackerFemale += 1
                        Mix.Range('AB'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('AB'+str(mixStep+5)).Value = "=("+str(countMonthly)+"/"+str(Mix.Range('AB'+str(mixStep+2)).Value)+")";
                        percentCheck2= int(Mix.Range('AB'+str(mixStep+5)).Value)
                        if (percentLimit <= percentCheck2):
                            sumMonthlyTracker= sumMonthlyTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumMonthlyTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumMonthlyTrackerFemale += 1

                        if(percentCheck1 >= percentLimit and percentCheck2 >= percentLimit):
                            if('4037' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMSecondaryFull += 1

                            if('4038' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMPreparationFull += 1

                            if('4031' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMExplorationFull += 1

                            if('4012' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMEmploymentFull += 1

                            if('4040' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMFAFSAFull += 1

                            if('4037' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFSecondaryFull += 1

                            if('4038' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFPreparationFull += 1

                            if('4031' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFExplorationFull += 1

                            if('4012' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFEmploymentFull += 1

                            if('4040' in str(Mix.Range('Y'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFFAFSAFull += 1
                                
                            if ("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                                asianServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    asianMaleServiced += 1
                                    maleServiced += 1
                            if ("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                                blackServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    blackMaleServiced += 1
                                    maleServiced += 1
                            if ("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                                hispanicServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    hispanicMaleServiced += 1
                                    maleServiced += 1
                            if ("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                                nativeServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    nativeMaleServiced += 1
                                    maleServiced += 1
                            if ("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                                whiteServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    whiteFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    whiteMaleServiced += 1
                                    maleServiced += 1
                            if ("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                                pacificServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    pacificMaleServiced += 1
                                    maleServiced += 1
                                    
                            if ("Academics" in str(Mix.Range('I'+str(mixStep)).Value)):
                                academicServiced += 1
                            if ("Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                                attendanceServiced += 1
                            if ("Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                                behaviorServiced += 1
                            if ("Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                                socialServiced += 1

                        if (str(Mix.Range('G'+str(mixStep)).Value) == "F"):
                                female += 1
                        if (str(Mix.Range('G'+str(mixStep)).Value) == "M"):
                                male += 1
                                
                        Mix.Range('AB'+str(mixStep+5)).NumberFormat = "0.00%"
                        Mix.Range('AB'+str(mixStep+6)).Value = "=("+str(countAllOther)+"/"+str(Mix.Range('AB'+str(mixStep+2)).Value)+")";
                        Mix.Range('AB'+str(mixStep+6)).NumberFormat = "0.00%"
                        flip = 3
                    elif(flip == 3):
                        SaveChanges=True
                        Mix.Range("A"+str(mixStep+1)+":K"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("L"+str(mixStep+1)+":AC"+str(mixStep+1)).Borders(8).LineStyle = 8
                        Mix.Range("AD"+str(mixStep+1)+":AG"+str(mixStep+1)).Borders(8).LineStyle = 1
                        Mix.Range("AD"+str(mixStep+1)+":AD"+str(mixStep+6)).Borders(7).LineStyle = 1
                        Mix.Range("AG"+str(mixStep+1)+":AG"+str(mixStep+6)).Borders(10).LineStyle = 1
                        Mix.Range("AD"+str(mixStep+6)+":AG"+str(mixStep+6)).Borders(9).LineStyle = 1
                        serviceList = str(Assess.Range('I'+str(assessStep)).Value)
                        serviceList = serviceList.replace(".0","")
                        serviceList = serviceList.replace("10191801","")
                        serviceList = serviceList.replace("1019","")
                        serviceList = serviceList.replace("1009","")
                        Mix.Range('AD'+str(mixStep+1)).Value = serviceList
                        Mix.Range('AD'+str(mixStep+2)).Value = "Monthly Check In"
                        Mix.Range('AD'+str(mixStep+3)).Value = "Planned Service"
                        Mix.Range('AD'+str(mixStep+4)).Value = "Indirect"
                        Mix.Range('AD'+str(mixStep+5)).Value = "All Other"
                        Mix.Range('AD'+str(mixStep+6)).Value = "Percent"
                        Mix.Range('AF'+str(mixStep+1)).Value = "Calculated Data"
                        Mix.Range('AF'+str(mixStep+2)).Value = "Months"
                        Mix.Range('AF'+str(mixStep+3)).Value = "Weight"
                        Mix.Range('AF'+str(mixStep+4)).Value = "Planned Service"
                        Mix.Range('AF'+str(mixStep+5)).Value = "Monthly Check In"
                        Mix.Range('AF'+str(mixStep+6)).Value = "All Other"
                        Mix.Range('AE'+str(mixStep+2)).Value = "="+str(countMonthly)+"/"+str(countOverall)
                        Mix.Range('AE'+str(mixStep+2)).NumberFormat = "0.00%"
                        Mix.Range('AE'+str(mixStep+3)).Value = "="+str(countService)+"/"+str(countOverall)
                        Mix.Range('AE'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('AE'+str(mixStep+4)).Value = "="+str(countIndirect)+"/"+str(countOverall)
                        Mix.Range('AE'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('AE'+str(mixStep+5)).Value = "="+str(countAllOther)+"/"+str(countOverall)
                        Mix.Range('AE'+str(mixStep+5)).NumberFormat = "0.00%"
                        if(countOverall != 0):
                            total = (countMonthly+countService+countIndirect+countAllOther)/(countOverall)
                        Mix.Range('AE'+str(mixStep+6)).Value = total
                        Mix.Range('AE'+str(mixStep+6)).NumberFormat = "0.00%"
                        Mix.Range('AG'+str(mixStep+2)).Value = datePart1 + str(lowMix) + datePart2 + str(highMix) + datePart3
                        Mix.Range('AG'+str(mixStep+3)).Value = "=SUM("+str(countMonthly)+"/"+str(Mix.Range('AG'+str(mixStep+2)).Value)+")"
                        Mix.Range('AG'+str(mixStep+3)).NumberFormat = "0.00%"
                        Mix.Range('AG'+str(mixStep+4)).Value = "=("+str(countService)+"/"+str(Mix.Range('AG'+str(mixStep+2)).Value)+")"#+"+str(Mix.Range('Y'+str(mixStep+2)).Value)
                        percentCheck1= int(Mix.Range('AG'+str(mixStep+4)).Value)
                        if (percentLimit <= percentCheck1):                                                                                                                                                                                            #Percent Location
                            sumServiceTracker= sumServiceTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumServiceTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumServiceTrackerFemale += 1
                        Mix.Range('AG'+str(mixStep+4)).NumberFormat = "0.00%"
                        Mix.Range('AG'+str(mixStep+5)).Value = "=("+str(countMonthly)+"/"+str(Mix.Range('AG'+str(mixStep+2)).Value)+")";
                        percentCheck2= int(Mix.Range('AG'+str(mixStep+5)).Value)
                        if (percentLimit <= percentCheck2):
                            sumMonthlyTracker= sumMonthlyTracker + 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "M"):
                                sumMonthlyTrackerMale += 1
                            if (str(Mix.Range('G'+str(mixStep+3)).Value) == "F"):
                                sumMonthlyTrackerFemale += 1
                                
                        if(percentCheck1 >= percentLimit and percentCheck2 >= percentLimit):
                            if('4037' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMSecondaryFull += 1

                            if('4038' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMPreparationFull += 1

                            if('4031' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMExplorationFull += 1

                            if('4012' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMEmploymentFull += 1

                            if('4040' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AMFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BMFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HMFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IMFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WMFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PMFAFSAFull += 1

                            if('4037' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFSecondaryFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFSecondaryFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFSecondaryFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFSecondaryFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFSecondaryFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFSecondaryFull += 1

                            if('4038' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFPreparationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFPreparationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFPreparationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFPreparationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFPreparationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFPreparationFull += 1

                            if('4031' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFExplorationFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFExplorationFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFExplorationFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFExplorationFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFExplorationFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFExplorationFull += 1

                            if('4012' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFEmploymentFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFEmploymentFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFEmploymentFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFEmploymentFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFEmploymentFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFEmploymentFull += 1

                            if('4040' in str(Mix.Range('AD'+str(mixStep)).Value)):
                                if('A' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    AFFAFSAFull += 1
                                if('B' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    BFFAFSAFull += 1
                                if('H' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    HFFAFSAFull += 1
                                if('I' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    IFFAFSAFull += 1
                                if('W' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    WFFAFSAFull += 1
                                if('P' in str(Mix.Range('H'+str(mixStep)).Value)):
                                    PFFAFSAFull += 1
                                    
                            if ("A" in str(Mix.Range('H'+str(mixStep)).Value)):
                                asianServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    asianFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    asianMaleServiced += 1
                                    maleServiced += 1
                            if ("B" in str(Mix.Range('H'+str(mixStep)).Value)):
                                blackServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    blackFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    blackMaleServiced += 1
                                    maleServiced += 1
                            if ("H" in str(Mix.Range('H'+str(mixStep)).Value)):
                                hispanicServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    hispanicFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    hispanicMaleServiced += 1
                                    maleServiced += 1
                            if ("I" in str(Mix.Range('H'+str(mixStep)).Value)):
                                nativeServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    nativeFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    nativeMaleServiced += 1
                                    maleServiced += 1
                            if ("W" in str(Mix.Range('H'+str(mixStep)).Value)):
                                whiteServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    whiteFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    whiteMaleServiced += 1
                                    maleServiced += 1
                            if ("P" in str(Mix.Range('H'+str(mixStep)).Value)):
                                pacificServiced += 1
                                if ("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                                    pacificFemaleServiced += 1
                                    femaleServiced += 1
                                else:
                                    pacificMaleServiced += 1
                                    maleServiced += 1
                                    
                            if ("Academics" in str(Mix.Range('I'+str(mixStep)).Value)):
                                academicServiced += 1
                            if ("Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                                attendanceServiced += 1
                            if ("Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                                behaviorServiced += 1
                            if ("Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                                socialServiced += 1

                        if (str(Mix.Range('G'+str(mixStep)).Value) == "F"):
                                female += 1
                        if (str(Mix.Range('G'+str(mixStep)).Value) == "M"):
                                male += 1
                                
                        Mix.Range('AG'+str(mixStep+5)).NumberFormat = "0.00%"
                        Mix.Range('AG'+str(mixStep+6)).Value = "=("+str(countAllOther)+"/"+str(Mix.Range('AG'+str(mixStep+2)).Value)+")";
                        Mix.Range('AG'+str(mixStep+6)).NumberFormat = "0.00%"
                        flip = 0
                        

                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        asianAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianFemaleAcademics += 1
                        else:
                            asianMaleAcademics += 1
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        asianAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianFemaleAttendance += 1
                        else:
                            asianMaleAttendance += 1
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        asianBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianFemaleBehavioral += 1
                        else:
                            asianMaleBehavioral += 1
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        asianSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            asianFemaleSocial += 1
                        else:
                            asianMaleSocial += 1
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        asianFemale += 1
                    elif ("A" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        asianMale += 1



                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        blackAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackFemaleAcademics += 1
                        else:
                            blackMaleAcademics += 1
                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        blackAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackFemaleAttendance += 1
                        else:
                            blackMaleAttendance += 1
                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        blackBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackFemaleBehavioral += 1
                        else:
                            blackMaleBehavioral += 1
                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        blackSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            blackFemaleSocial += 1
                        else:
                            blackMaleSocial += 1
                    if("B" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        blackFemale += 1
                    elif ("B" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        blackMale += 1


                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        hispanicAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicFemaleAcademics += 1
                        else:
                            hispanicMaleAcademics += 1
                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        hispanicAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicFemaleAttendance += 1
                        else:
                            hispanicMaleAttendance += 1
                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        hispanicBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicFemaleBehavioral += 1
                        else:
                            hispanicMaleBehavioral += 1
                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        hispanicSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            hispanicFemaleSocial += 1
                        else:
                            hispanicMaleSocial += 1
                    if("H" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        hispanicFemale += 1
                    elif ("H" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        hispanicMale += 1


                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        nativeAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeFemaleAcademics += 1
                        else:
                            nativeMaleAcademics += 1
                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        nativeAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeFemaleAttendance += 1
                        else:
                            nativeMaleAttendance += 1
                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        nativeBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeFemaleBehavioral += 1
                        else:
                            nativeMaleBehavioral += 1
                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        nativeSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            nativeFemaleSocial += 1
                        else:
                            nativeMaleSocial += 1
                    if("I" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        nativeFemale += 1
                    elif ("I" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        nativeMale += 1


                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        whiteAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            whiteFemaleAcademics += 1
                        else:
                            whiteMaleAcademics += 1
                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        whiteAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            whiteFemaleAttendance += 1
                        else:
                            whiteMaleAttendance += 1
                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        whiteBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            whiteFemaleBehavioral += 1
                        else:
                            whiteMaleBehavioral += 1
                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        whiteSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            whiteFemaleSocial += 1
                        else:
                            whiteMaleSocial += 1
                    if("W" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        whiteFemale += 1
                    elif ("W" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        whiteMale += 1

                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Academic" in str(Mix.Range('I'+str(mixStep)).Value)):
                        pacificAcademics += 1
                        countAcademics +=1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificFemaleAcademics += 1
                        else:
                            pacificMaleAcademics += 1
                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Attendance" in str(Mix.Range('I'+str(mixStep)).Value)):
                        pacificAttendance += 1
                        countAttendance += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificFemaleAttendance += 1
                        else:
                            pacificMaleAttendance += 1
                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Behavior" in str(Mix.Range('I'+str(mixStep)).Value)):
                        pacificBehavioral += 1
                        countBehavior += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificFemaleBehavioral += 1
                        else:
                            pacificMaleBehavioral += 1
                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "Social Services" in str(Mix.Range('I'+str(mixStep)).Value)):
                        pacificSocial += 1
                        countSocialService += 1
                        if("F" in str(Mix.Range('G'+str(mixStep)).Value)):
                            pacificFemaleSocial += 1
                        else:
                            pacificMaleSocial += 1
                    if("P" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        pacificFemale += 1
                    elif ("P" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        pacificMale += 1
                            
                    countServiceTracker += countService
                    countMonthlyTracker += countMonthly
                    countIndirectTracker += countIndirect
                    countAllOtherTracker += countAllOther
                    countMonthly = 0
                    countService = 0
                    countIndirect = 0
                    countAllOther = 0
                    countOverall = 0
                    switchSex = 0
                    highDate = str(Mix.Range('J'+str(mixStep)).Value);
                    lowDate = str(Mix.Range('J'+str(mixStep)).Value);
                    compareDate = str(Mix.Range('J'+str(mixStep)).Value);
                if(mixValue == assessValue):
                    serviceMix = str(Mix.Range('K'+str(mixStep)).Value)
                    if("A" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        asianFemaleOverall += 1
                    elif("A" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        asianMaleOverall += 1
                    elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        blackFemaleOverall += 1
                    elif("B" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        blackMaleOverall += 1
                    elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        hispanicFemaleOverall += 1
                    elif("H" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        blackMaleOverall += 1
                    elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        nativeFemaleOverall += 1
                    elif("I" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        nativeMaleOverall += 1
                    elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):    
                        whiteFemaleOverall += 1
                    elif("W" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):    
                        whiteMaleOverall += 1
                    elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "F" in str(Mix.Range('G'+str(mixStep)).Value)):
                        pacificFemaleOverall += 1
                    elif("P" in str(Mix.Range('H'+str(mixStep)).Value) and "M" in str(Mix.Range('G'+str(mixStep)).Value)):
                        pacificMaleOverall += 1
            #if(assessValue != mixValue):
            #    assessStep += 1
            mixStep = 1 + mixStep
            if(assessValue == None or mixValue == None):
                break
            target.Save()
            target2.Save()

        programDetails.Status = ""
        programDetails.Status = "Page Calculated..."

        '''
        Mix.Range("B"+str(mixStep+4)+":D"+str(mixStep+5)).Borders(8).LineStyle = 1
        Mix.Range("B"+str(mixStep+4)+":B"+str(mixStep+9)).Borders(7).LineStyle = 1
        Mix.Range("D"+str(mixStep+4)+":D"+str(mixStep+9)).Borders(10).LineStyle = 1
        Mix.Range("B"+str(mixStep+10)+":D"+str(mixStep+10)).Borders(8).LineStyle = 1

        Mix.Range('B'+str(mixStep+5)).Value = "Asian/Pacific Island"
        Mix.Range('C'+str(mixStep+5)).Value = asianEth
        Mix.Range('B'+str(mixStep+6)).Value = "Black/African American"
        Mix.Range('C'+str(mixStep+6)).Value = blackEth
        Mix.Range('B'+str(mixStep+7)).Value = "Hispanic/Latino"
        Mix.Range('C'+str(mixStep+7)).Value = hispanicEth
        Mix.Range('B'+str(mixStep+8)).Value = "Native American"
        Mix.Range('C'+str(mixStep+8)).Value = nativeEth
        Mix.Range('B'+str(mixStep+9)).Value = "White"
        Mix.Range('C'+str(mixStep+9)).Value = whiteEth

        Mix.Range("F"+str(mixStep+4)+":H"+str(mixStep+5)).Borders(8).LineStyle = 1
        Mix.Range("F"+str(mixStep+4)+":F"+str(mixStep+6)).Borders(7).LineStyle = 1
        Mix.Range("H"+str(mixStep+4)+":H"+str(mixStep+6)).Borders(10).LineStyle = 1
        Mix.Range("F"+str(mixStep+7)+":H"+str(mixStep+7)).Borders(8).LineStyle = 1

        Mix.Range('F'+str(mixStep+5)).Value = "Female"
        Mix.Range('G'+str(mixStep+4)).Value = "Sex/Monthly"
        Mix.Range('G'+str(mixStep+5)).Value = '=(' + str(sumMonthlyTrackerFemale) + '/' + str(femaleTracker) + ')'
        Mix.Range('G'+str(mixStep+5)).NumberFormat = "0.00%"
        Mix.Range('H'+str(mixStep+5)).Value = '=(' + str(sumServiceTrackerFemale) + '/' + str(femaleTracker) + ')'
        Mix.Range('H'+str(mixStep+5)).NumberFormat = "0.00%"
        Mix.Range('F'+str(mixStep+6)).Value = "Male"
        Mix.Range('H'+str(mixStep+4)).Value = "Sex/Service"
        Mix.Range('G'+str(mixStep+6)).Value = '=(' + str(sumMonthlyTrackerMale) + '/' + str(maleTracker) + ')'
        Mix.Range('G'+str(mixStep+6)).NumberFormat = "0.00%"
        Mix.Range('H'+str(mixStep+6)).Value = '=(' + str(sumServiceTrackerMale) + '/' + str(maleTracker) + ')'
        Mix.Range('H'+str(mixStep+6)).NumberFormat = "0.00%"

        Mix.Range("F"+str(mixStep+9)+":H"+str(mixStep+10)).Borders(8).LineStyle = 1
        Mix.Range("F"+str(mixStep+9)+":F"+str(mixStep+11)).Borders(7).LineStyle = 1
        Mix.Range("H"+str(mixStep+9)+":H"+str(mixStep+11)).Borders(10).LineStyle = 1
        Mix.Range("F"+str(mixStep+12)+":H"+str(mixStep+12)).Borders(8).LineStyle = 1

        Mix.Range('G'+str(mixStep+9)).Value = datePart1 + str(lowMixMax) + datePart2 + str(highMixMax) + datePart3
        Mix.Range('F'+str(mixStep+10)).Value = "Monthly Check-In"
        Mix.Range('G'+str(mixStep+10)).Value = '=(' + str(countMonthlyTracker) + '/' + str(assessStep) + ')/' +str(Mix.Range('G'+str(mixStep+9)).Value)
        Mix.Range('G'+str(mixStep+10)).NumberFormat = "0.00%"
        Mix.Range('F'+str(mixStep+11)).Value = "Services"
        Mix.Range('G'+str(mixStep+11)).Value = '=(' + str(countServiceTracker) + '/' + str(assessStep) + ')/' +str(Mix.Range('G'+str(mixStep+9)).Value)
        Mix.Range('G'+str(mixStep+11)).NumberFormat = "0.00%"
        Mix.Range('H'+str(mixStep+10)).Value = '=(' + str(sumMonthlyTracker) + '/' + str(assessStep) + ')'
        Mix.Range('H'+str(mixStep+10)).NumberFormat = "0.00%"
        Mix.Range('H'+str(mixStep+11)).Value = '=(' + str(sumServiceTracker) + '/' + str(assessStep) + ')'
        Mix.Range('H'+str(mixStep+11)).NumberFormat = "0.00%"
        '''
        target.Close(SaveChanges=True)
        target2.Close(SaveChanges=True)
        xlApp.Quit()

        xlApp = win32.Dispatch('Excel.Application')
        pathThree = 'C:/Users/T Choat/Desktop/Python Code/S-U-R-V-E-Y L-I-S-T/Service Dash'
        target = xlApp.Workbooks.Open(pathTwo)
        target3 = xlApp.Workbooks.Open(pathThree)
        MixSort = target.Worksheets(locationFour)

        cutting = target3.Worksheets(1)
        cutting.Copy(After=MixSort)

        target.Close(SaveChanges=True)
        target3.Close(SaveChanges=True)

        xlApp.Quit()

        programDetails.Status = ""
        programDetails.Status = "Dashboard Calculating..."

        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        target2 = xlApp.Workbooks.Open(pathOne)
        target3 = xlApp.Workbooks.Open(pathThree)
        Mix = target.Worksheets(locationTwo)
        Assess = target.Worksheets(locationOne)
        Dash = target.Worksheets("Dashboard")

        Dash.Range('L6').Value = datePart1Alt + str(lowMixMax) + datePart2Alt + str(highMixMax) + datePart3
        Dash.Range('L6').NumberFormat = "0"
        Dash.Range('D9').Value = '=(' + str(countMonthlyTracker) + ')/(' + str(Dash.Range('L6').Value) +  ")"
        Dash.Range('D15').Value = '=(' + str(countAllOtherTracker) + ')/(' + str(Dash.Range('L6').Value) +  ")"
        Dash.Range('F9').Value = '=(((' + str(countMonthlyTracker) + ')/' + str(Dash.Range('L6').Value) +  ')/((' + str(countAllOtherTracker+countIndirectTracker+countMonthlyTracker+countServiceTracker) + ")))"
        Dash.Range('F9').NumberFormat = "0.00%"
        Dash.Range('H9').Value = '=(' + str(countMonthlyTracker) +  ")"
        Dash.Range('J9').Value = '=((' + str(countMonthlyTracker) + '/' + str(countAllOtherTracker+countIndirectTracker+countMonthlyTracker+countServiceTracker) + '))'
        Dash.Range('J9').NumberFormat = "0.00%"

        Dash.Range('D11').Value = '=(' + str(countServiceTracker) + ')/(' + str(Dash.Range('L6').Value) +  ")"
        Dash.Range('F11').Value = '=(((' + str(countServiceTracker) + ')/' + str(Dash.Range('L6').Value) +  ')/((' + str(countAllOtherTracker+countIndirectTracker+countMonthlyTracker+countServiceTracker) + ")))"
        Dash.Range('F11').NumberFormat = "0.00%"
        Dash.Range('H11').Value = '=(' + str(countServiceTracker) +  ")"
        Dash.Range('J11').Value = '=((' + str(countServiceTracker) + '/' + str(countAllOtherTracker+countIndirectTracker+countMonthlyTracker+countServiceTracker) + '))'
        Dash.Range('J11').NumberFormat = "0.00%"

        Dash.Range('D13').Value = '=(' + str(countIndirectTracker) + ')/(' + str(Dash.Range('L6').Value) +  ")"
        Dash.Range('F13').Value = '=(((' + str(countIndirectTracker) + ')/' + str(Dash.Range('L6').Value) +  ')/((' + str(countAllOtherTracker+countIndirectTracker+countMonthlyTracker+countServiceTracker) + ")))"
        Dash.Range('F13').NumberFormat = "0.00%"
        Dash.Range('H13').Value = '=(' + str(countIndirectTracker) +  ")"
        Dash.Range('J13').Value = '=((' + str(countIndirectTracker) + '/' + str(countAllOtherTracker+countIndirectTracker+countMonthlyTracker+countServiceTracker) + '))'
        Dash.Range('J13').NumberFormat = "0.00%"

        Dash.Range('D15').Value = '=(' + str(countAllOtherTracker) + ')/(' + str(Dash.Range('L6').Value) +  ")"
        Dash.Range('F15').Value = '=(((' + str(countAllOtherTracker) + ')/' + str(Dash.Range('L6').Value) +  ')/((' + str(countAllOtherTracker+countIndirectTracker+countMonthlyTracker+countServiceTracker) + ")))"
        Dash.Range('F15').NumberFormat = "0.00%"
        Dash.Range('H15').Value = '=(' + str(countAllOtherTracker) +  ")"
        Dash.Range('J15').Value = '=((' + str(countAllOtherTracker) + '/' + str(countAllOtherTracker+countIndirectTracker+countMonthlyTracker+countServiceTracker) + '))'
        Dash.Range('J15').NumberFormat = "0.00%"



        #G2
        Dash.Range('D20').Value = "=(" + str(asianServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D20').NumberFormat = "0.00%"
        Dash.Range('D22').Value = "=(" + str(blackServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D22').NumberFormat = "0.00%"
        Dash.Range('D24').Value = "=(" + str(hispanicServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D24').NumberFormat = "0.00%"
        Dash.Range('D26').Value = "=(" + str(nativeServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D26').NumberFormat = "0.00%"
        Dash.Range('D28').Value = "=(" + str(whiteServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D28').NumberFormat = "0.00%"
        Dash.Range('D30').Value = "=(" + str(pacificServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D30').NumberFormat = "0.00%"

        Dash.Range('F20').Value = asianServiced
        Dash.Range('F22').Value = blackServiced
        Dash.Range('F24').Value = hispanicServiced
        Dash.Range('F26').Value = nativeServiced
        Dash.Range('F28').Value = whiteServiced
        Dash.Range('F30').Value = pacificServiced

        Dash.Range('H20').Value = asianEth
        Dash.Range('H22').Value = blackEth
        Dash.Range('H24').Value = hispanicEth
        Dash.Range('H26').Value = nativeEth
        Dash.Range('H28').Value = whiteEth
        Dash.Range('H30').Value = pacificEth

        Dash.Range('J20').Value = "=(" + str(asianEth) + "/(" + str(asianEth) + "+" + str(blackEth) + "+" + str(hispanicEth) + "+" + str(nativeEth) + "+" + str(whiteEth) + "+" + str(pacificEth) + "))"
        Dash.Range('J20').NumberFormat = "0.00%"
        Dash.Range('J22').Value = "=(" + str(blackEth) + "/(" + str(asianEth) + "+" + str(blackEth) + "+" + str(hispanicEth) + "+" + str(nativeEth) + "+" + str(whiteEth) + "+" +  str(pacificEth) + "))"
        Dash.Range('J22').NumberFormat = "0.00%"
        Dash.Range('J24').Value = "=(" + str(hispanicEth) + "/(" + str(asianEth) + "+" + str(blackEth) + "+" + str(hispanicEth) + "+" + str(nativeEth) + "+" + str(whiteEth) + "+" +  str(pacificEth) + "))"
        Dash.Range('J24').NumberFormat = "0.00%"
        Dash.Range('J26').Value = "=(" + str(nativeEth) + "/(" + str(asianEth) + "+" + str(blackEth) + "+" + str(hispanicEth) + "+" + str(nativeEth) + "+" + str(whiteEth) + "+" +  str(pacificEth) + "))"
        Dash.Range('J26').NumberFormat = "0.00%"
        Dash.Range('J28').Value = "=(" + str(whiteEth) + "/(" + str(asianEth) + "+" + str(blackEth) + "+" + str(hispanicEth) + "+" + str(nativeEth) + "+" + str(whiteEth) + "+" +  str(pacificEth) + "))"
        Dash.Range('J28').NumberFormat = "0.00%"
        Dash.Range('J30').Value = "=(" + str(pacificEth) + "/(" + str(asianEth) + "+" + str(blackEth) + "+" + str(hispanicEth) + "+" + str(nativeEth) + "+" + str(whiteEth) + "+" +  str(pacificEth) + "))"
        Dash.Range('J30').NumberFormat = "0.00%"

        #G3
        Dash.Range('D35').Value = "=(" + str(behaviorServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D35').NumberFormat = "0.00%"
        Dash.Range('D37').Value = "=(" + str(academicServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D37').NumberFormat = "0.00%"
        Dash.Range('D39').Value = "=(" + str(socialServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D39').NumberFormat = "0.00%"
        Dash.Range('D41').Value = "=(" + str(attendanceServiced) + "/" + str(assessStep) + ")"
        Dash.Range('D41').NumberFormat = "0.00%"

        Dash.Range('F35').Value = behaviorServiced
        Dash.Range('F37').Value = academicServiced
        Dash.Range('F39').Value = socialServiced
        Dash.Range('F41').Value = attendanceServiced

        Dash.Range('H35').Value = countBehavior
        if(countBehavior != 0):
            Dash.Range('J35').Value = countBehavior/(countBehavior+countAcademics+countSocialService+countAttendance)
        else:
            Dash.Range('J35').Value = 0
        Dash.Range('J35').NumberFormat = "0.00%"
        Dash.Range('H37').Value = countAcademics
        if(countAcademics != 0):
            Dash.Range('J37').Value = countAcademics/(countBehavior+countAcademics+countSocialService+countAttendance)
        else:
            Dash.Range('J37').Value = 0
        Dash.Range('J37').NumberFormat = "0.00%"
        Dash.Range('H39').Value = countSocialService
        if(countSocialService != 0):
            Dash.Range('J39').Value = countSocialService/(countBehavior+countAcademics+countSocialService+countAttendance)
        else:
            Dash.Range('J39').Value = 0
        Dash.Range('J39').NumberFormat = "0.00%"
        Dash.Range('H41').Value = countAttendance
        if(countAttendance != 0):
            Dash.Range('J41').Value = countAttendance/(countBehavior+countAcademics+countSocialService+countAttendance)
        else:
            Dash.Range('J41').Value = 0
        Dash.Range('J41').NumberFormat = "0.00%"

        #R/E 1
        Dash.Range('O9').Value = "=((" + str(asianMaleService) + "+" + str(asianFemaleService) + "+" + str(asianMaleMonthly) + "+" + str(asianFemaleMonthly) + "+" + str(asianMaleIndirect) + "+" + str(asianFemaleIndirect) + "+" + str(asianMaleOther) + "+" + str(asianFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")"
        Dash.Range('O11').Value = "=((" + str(blackMaleService) + "+" + str(blackFemaleService) + "+" + str(blackMaleMonthly) + "+" + str(blackFemaleMonthly) + "+" + str(blackMaleIndirect) + "+" + str(blackFemaleIndirect) + "+" + str(blackMaleOther) + "+" + str(blackFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")"
        Dash.Range('O13').Value = "=((" + str(hispanicMaleService) + "+" + str(hispanicFemaleService) + "+" + str(hispanicMaleMonthly) + "+" + str(hispanicFemaleMonthly) + "+" + str(hispanicMaleIndirect) + "+" + str(hispanicFemaleIndirect) + "+" + str(hispanicMaleOther) + "+" + str(hispanicFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")"
        Dash.Range('O15').Value = "=((" + str(nativeMaleService) + "+" + str(nativeFemaleService) + "+" + str(nativeMaleMonthly) + "+" + str(nativeFemaleMonthly) + "+" + str(nativeMaleIndirect) + "+" + str(nativeFemaleIndirect) + "+" + str(nativeMaleOther) + "+" + str(nativeFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")"
        Dash.Range('O17').Value = "=((" + str(whiteMaleService) + "+" + str(whiteFemaleService) + "+" + str(whiteMaleMonthly) + "+" + str(whiteFemaleMonthly) + "+" + str(whiteMaleIndirect) + "+" + str(whiteFemaleIndirect) + "+" + str(whiteMaleOther) + "+" + str(whiteFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")"
        Dash.Range('O19').Value = "=((" + str(pacificMaleService) + "+" + str(pacificFemaleService) + "+" + str(pacificMaleMonthly) + "+" + str(pacificFemaleMonthly) + "+" + str(pacificMaleIndirect) + "+" + str(pacificFemaleIndirect) + "+" + str(pacificMaleOther) + "+" + str(pacificFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")"

        totalSumRE = asianMaleService + asianFemaleService + asianMaleMonthly + asianFemaleMonthly + asianMaleIndirect + asianFemaleIndirect + asianMaleOther + asianFemaleOther + blackMaleService + blackFemaleService + blackMaleMonthly + blackFemaleMonthly + blackMaleIndirect + blackFemaleIndirect + blackMaleOther + blackFemaleOther + hispanicMaleService + hispanicFemaleService + hispanicMaleMonthly + hispanicFemaleMonthly + hispanicMaleIndirect + hispanicFemaleIndirect + hispanicMaleOther + hispanicFemaleOther + nativeMaleService + nativeFemaleService + nativeMaleMonthly + nativeFemaleMonthly + nativeMaleIndirect + nativeFemaleIndirect + nativeMaleOther + nativeFemaleOther + whiteMaleService + whiteFemaleService + whiteMaleMonthly + whiteFemaleMonthly + whiteMaleIndirect + whiteFemaleIndirect + whiteMaleOther + whiteFemaleOther + pacificMaleService + pacificFemaleService + pacificMaleMonthly + pacificFemaleMonthly + pacificMaleIndirect + pacificFemaleIndirect + pacificMaleOther + pacificFemaleOther
        Dash.Range('Q9').Value = "=(((" + str(asianMaleService) + "+" + str(asianFemaleService) + "+" + str(asianMaleMonthly) + "+" + str(asianFemaleMonthly) + "+" + str(asianMaleIndirect) + "+" + str(asianFemaleIndirect) + "+" + str(asianMaleOther) + "+" + str(asianFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")/((" +  str(totalSumRE) + ")/" + str(Dash.Range('L6').Value) +"))"
        Dash.Range('Q9').NumberFormat = "0.00%"
        Dash.Range('Q11').Value = "=(((" + str(blackMaleService) + "+" + str(blackFemaleService) + "+" + str(blackMaleMonthly) + "+" + str(blackFemaleMonthly) + "+" + str(blackMaleIndirect) + "+" + str(blackFemaleIndirect) + "+" + str(blackMaleOther) + "+" + str(blackFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")/((" +  str(totalSumRE) + ")/" + str(Dash.Range('L6').Value) +"))"
        Dash.Range('Q11').NumberFormat = "0.00%"
        Dash.Range('Q13').Value = "=(((" + str(hispanicMaleService) + "+" + str(hispanicFemaleService) + "+" + str(hispanicMaleMonthly) + "+" + str(hispanicFemaleMonthly) + "+" + str(hispanicMaleIndirect) + "+" + str(hispanicFemaleIndirect) + "+" + str(hispanicMaleOther) + "+" + str(hispanicFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")/((" +  str(totalSumRE) + ")/" + str(Dash.Range('L6').Value) +"))"
        Dash.Range('Q13').NumberFormat = "0.00%"
        Dash.Range('Q15').Value = "=(((" + str(nativeMaleService) + "+" + str(nativeFemaleService) + "+" + str(nativeMaleMonthly) + "+" + str(nativeFemaleMonthly) + "+" + str(nativeMaleIndirect) + "+" + str(nativeFemaleIndirect) + "+" + str(nativeMaleOther) + "+" + str(nativeFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")/((" +  str(totalSumRE) + ")/" + str(Dash.Range('L6').Value) +"))"
        Dash.Range('Q15').NumberFormat = "0.00%"
        Dash.Range('Q17').Value = "=(((" + str(whiteMaleService) + "+" + str(whiteFemaleService) + "+" + str(whiteMaleMonthly) + "+" + str(whiteFemaleMonthly) + "+" + str(whiteMaleIndirect) + "+" + str(whiteFemaleIndirect) + "+" + str(whiteMaleOther) + "+" + str(whiteFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")/((" +  str(totalSumRE) + ")/" + str(Dash.Range('L6').Value) +"))"
        Dash.Range('Q17').NumberFormat = "0.00%"
        Dash.Range('Q19').Value = "=(((" + str(pacificMaleService) + "+" + str(pacificFemaleService) + "+" + str(pacificMaleMonthly) + "+" + str(pacificFemaleMonthly) + "+" + str(pacificMaleIndirect) + "+" + str(pacificFemaleIndirect) + "+" + str(pacificMaleOther) + "+" + str(pacificFemaleOther) + ")/" + str(Dash.Range('L6').Value) + ")/((" +  str(totalSumRE) + ")/" + str(Dash.Range('L6').Value) +"))"
        Dash.Range('Q19').NumberFormat = "0.00%"


        Dash.Range('S9').Value = "=(" + str(asianMaleService) + "+" + str(asianFemaleService) + "+" + str(asianMaleMonthly) + "+" + str(asianFemaleMonthly) + "+" + str(asianMaleIndirect) + "+" + str(asianFemaleIndirect) + "+" + str(asianMaleOther) + "+" + str(asianFemaleOther) + ")"
        Dash.Range('S11').Value = "=(" + str(blackMaleService) + "+" + str(blackFemaleService) + "+" + str(blackMaleMonthly) + "+" + str(blackFemaleMonthly) + "+" + str(blackMaleIndirect) + "+" + str(blackFemaleIndirect) + "+" + str(blackMaleOther) + "+" + str(blackFemaleOther) + ")"
        Dash.Range('S13').Value = "=(" + str(hispanicMaleService) + "+" + str(hispanicFemaleService) + "+" + str(hispanicMaleMonthly) + "+" + str(hispanicFemaleMonthly) + "+" + str(hispanicMaleIndirect) + "+" + str(hispanicFemaleIndirect) + "+" + str(hispanicMaleOther) + "+" + str(hispanicFemaleOther) + ")"
        Dash.Range('S15').Value = "=(" + str(nativeMaleService) + "+" + str(nativeFemaleService) + "+" + str(nativeMaleMonthly) + "+" + str(nativeFemaleMonthly) + "+" + str(nativeMaleIndirect) + "+" + str(nativeFemaleIndirect) + "+" + str(nativeMaleOther) + "+" + str(nativeFemaleOther) + ")"
        Dash.Range('S17').Value = "=(" + str(whiteMaleService) + "+" + str(whiteFemaleService) + "+" + str(whiteMaleMonthly) + "+" + str(whiteFemaleMonthly) + "+" + str(whiteMaleIndirect) + "+" + str(whiteFemaleIndirect) + "+" + str(whiteMaleOther) + "+" + str(whiteFemaleOther) + ")"
        Dash.Range('S19').Value = "=(" + str(pacificMaleService) + "+" + str(pacificFemaleService) + "+" + str(pacificMaleMonthly) + "+" + str(pacificFemaleMonthly) + "+" + str(pacificMaleIndirect) + "+" + str(pacificFemaleIndirect) + "+" + str(pacificMaleOther) + "+" + str(pacificFemaleOther) + ")"

        Dash.Range('U9').Value = "=((" + str(asianMaleService) + "+" + str(asianFemaleService) + "+" + str(asianMaleMonthly) + "+" + str(asianFemaleMonthly) + "+" + str(asianMaleIndirect) + "+" + str(asianFemaleIndirect) + "+" + str(asianMaleOther) + "+" + str(asianFemaleOther) + ")/(" + str(totalSumRE) + "))"
        Dash.Range('U9').NumberFormat = "0.00%"
        Dash.Range('U11').Value = "=((" + str(blackMaleService) + "+" + str(blackFemaleService) + "+" + str(blackMaleMonthly) + "+" + str(blackFemaleMonthly) + "+" + str(blackMaleIndirect) + "+" + str(blackFemaleIndirect) + "+" + str(blackMaleOther) + "+" + str(blackFemaleOther) + ")/(" + str(totalSumRE) + "))"
        Dash.Range('U11').NumberFormat = "0.00%"
        Dash.Range('U13').Value = "=((" + str(hispanicMaleService) + "+" + str(hispanicFemaleService) + "+" + str(hispanicMaleMonthly) + "+" + str(hispanicFemaleMonthly) + "+" + str(hispanicMaleIndirect) + "+" + str(hispanicFemaleIndirect) + "+" + str(hispanicMaleOther) + "+" + str(hispanicFemaleOther) + ")/(" + str(totalSumRE) + "))"
        Dash.Range('U13').NumberFormat = "0.00%"
        Dash.Range('U15').Value = "=((" + str(nativeMaleService) + "+" + str(nativeFemaleService) + "+" + str(nativeMaleMonthly) + "+" + str(nativeFemaleMonthly) + "+" + str(nativeMaleIndirect) + "+" + str(nativeFemaleIndirect) + "+" + str(nativeMaleOther) + "+" + str(nativeFemaleOther) + ")/(" + str(totalSumRE) + "))"
        Dash.Range('U15').NumberFormat = "0.00%"
        Dash.Range('U17').Value = "=((" + str(whiteMaleService) + "+" + str(whiteFemaleService) + "+" + str(whiteMaleMonthly) + "+" + str(whiteFemaleMonthly) + "+" + str(whiteMaleIndirect) + "+" + str(whiteFemaleIndirect) + "+" + str(whiteMaleOther) + "+" + str(whiteFemaleOther) + ")/(" + str(totalSumRE) + "))"
        Dash.Range('U17').NumberFormat = "0.00%"
        Dash.Range('U19').Value = "=((" + str(pacificMaleService) + "+" + str(pacificFemaleService) + "+" + str(pacificMaleMonthly) + "+" + str(pacificFemaleMonthly) + "+" + str(pacificMaleIndirect) + "+" + str(pacificFemaleIndirect) + "+" + str(pacificMaleOther) + "+" + str(pacificFemaleOther) + ")/(" + str(totalSumRE) + "))"
        Dash.Range('U19').NumberFormat = "0.00%"


        #R/E 2
        Dash.Range('O24').Value =  asianBehavioral
        Dash.Range('Q24').Value =  asianAcademics
        Dash.Range('S24').Value =  asianSocial
        Dash.Range('U24').Value =  asianAttendance
        Dash.Range('O26').Value =  blackBehavioral
        Dash.Range('Q26').Value =  blackAcademics
        Dash.Range('S26').Value =  blackSocial
        Dash.Range('U26').Value =  blackAttendance
        Dash.Range('O28').Value =  hispanicBehavioral
        Dash.Range('Q28').Value =  hispanicAcademics
        Dash.Range('S28').Value =  hispanicSocial
        Dash.Range('U28').Value =  hispanicAttendance
        Dash.Range('O30').Value =  nativeBehavioral
        Dash.Range('Q30').Value =  nativeAcademics
        Dash.Range('S30').Value =  nativeSocial
        Dash.Range('U30').Value =  nativeAttendance
        Dash.Range('O32').Value =  whiteBehavioral
        Dash.Range('Q32').Value =  whiteAcademics
        Dash.Range('S32').Value =  whiteSocial
        Dash.Range('U32').Value =  whiteAttendance
        Dash.Range('O34').Value =  pacificBehavioral
        Dash.Range('Q34').Value =  pacificAcademics
        Dash.Range('S34').Value =  pacificSocial
        Dash.Range('U34').Value =  pacificAttendance

        #G/S 1

        Dash.Range('Z9').Value = "=((" + str(asianMaleOverall) + "+" + str(blackMaleOverall) + "+" + str(hispanicMaleOverall) + "+" + str(nativeMaleOverall) + "+" + str(whiteMaleOverall) + "+" + str(pacificMaleOverall) + ")/(" + str(Dash.Range('L6').Value) + "))"
        Dash.Range('Z11').Value = "=((" + str(asianFemaleOverall) + "+" + str(blackFemaleOverall) + "+" + str(hispanicFemaleOverall) + "+" + str(nativeFemaleOverall) + "+" + str(whiteFemaleOverall) + "+" + str(pacificFemaleOverall) + ")/(" + str(Dash.Range('L6').Value) + "))"

        Dash.Range('AB9').Value = "=((" + str(asianMaleOverall) + "+" + str(blackMaleOverall) + "+" + str(hispanicMaleOverall) + "+" + str(nativeMaleOverall) + "+" + str(whiteMaleOverall) + "+" + str(pacificMaleOverall) + ")/(" + str(asianMaleOverall) + "+" + str(blackMaleOverall) + "+" + str(hispanicMaleOverall) + "+" + str(nativeMaleOverall) + "+" + str(whiteMaleOverall) + "+" + str(pacificMaleOverall) + "+" + str(asianFemaleOverall) + "+" + str(blackFemaleOverall) + "+" + str(hispanicFemaleOverall) + "+" + str(nativeFemaleOverall) + "+" + str(whiteFemaleOverall) + "+" + str(pacificFemaleOverall) + "))/" + str(Dash.Range('L6').Value)
        Dash.Range('AB9').NumberFormat = "0.00%"
        Dash.Range('AB11').Value = "=((" + str(asianFemaleOverall) + "+" + str(blackFemaleOverall) + "+" + str(hispanicFemaleOverall) + "+" + str(nativeFemaleOverall) + "+" + str(whiteFemaleOverall) + "+" + str(pacificFemaleOverall) + ")/(" + str(asianMaleOverall) + "+" + str(blackMaleOverall) + "+" + str(hispanicMaleOverall) + "+" + str(nativeMaleOverall) + "+" + str(whiteMaleOverall) + "+" + str(pacificMaleOverall)  + str(asianFemaleOverall) + "+" + str(blackFemaleOverall) + "+" + str(hispanicFemaleOverall) + "+" + str(nativeFemaleOverall) + "+" + str(whiteFemaleOverall) + "+" + str(pacificFemaleOverall) + "))/" + str(Dash.Range('L6').Value)
        Dash.Range('AB11').NumberFormat = "0.00%"

        Dash.Range('AD9').Value = "=(" + str(asianMaleOverall) + "+" + str(blackMaleOverall) + "+" + str(hispanicMaleOverall) + "+" + str(nativeMaleOverall) + "+" + str(whiteMaleOverall) + "+" + str(pacificMaleOverall) + ")"
        Dash.Range('AD11').Value = "=(" + str(asianFemaleOverall) + "+" + str(blackFemaleOverall) + "+" + str(hispanicFemaleOverall) + "+" + str(nativeFemaleOverall) + "+" + str(whiteFemaleOverall) + "+" + str(pacificFemaleOverall) + ")"

        Dash.Range('AF9').Value = "=((" + str(asianMaleOverall) + "+" + str(blackMaleOverall) + "+" + str(hispanicMaleOverall) + "+" + str(nativeMaleOverall) + "+" + str(whiteMaleOverall) + "+" + str(pacificMaleOverall) + ")/(" + str(asianMaleOverall) + "+" + str(blackMaleOverall) + "+" + str(hispanicMaleOverall) + "+" + str(nativeMaleOverall) + "+" + str(whiteMaleOverall) + "+" + str(pacificMaleOverall) + "+" + str(asianFemaleOverall) + "+" + str(blackFemaleOverall) + "+" + str(hispanicFemaleOverall) + "+" + str(nativeFemaleOverall) + "+" + str(whiteFemaleOverall) + "+" + str(pacificFemaleOverall) +"))"
        Dash.Range('AF9').NumberFormat = "0.00%"

        Dash.Range('AF11').Value = "=((" + str(asianFemaleOverall) + "+" + str(blackFemaleOverall) + "+" + str(hispanicFemaleOverall) + "+" + str(nativeFemaleOverall) + "+" + str(whiteFemaleOverall) + "+" + str(pacificFemaleOverall) + ")/(" + str(asianMaleOverall) + "+" + str(blackMaleOverall) + "+" + str(hispanicMaleOverall) + "+" + str(nativeMaleOverall) + "+" + str(whiteMaleOverall) + "+" + str(pacificMaleOverall)  + "+" + str(asianFemaleOverall) + "+" + str(blackFemaleOverall) + "+" + str(hispanicFemaleOverall) + "+" + str(nativeFemaleOverall) + "+" + str(whiteFemaleOverall) + "+" + str(pacificFemaleOverall) +"))"
        Dash.Range('AF11').NumberFormat = "0.00%"

        #G/S 2
        Dash.Range('Z16').Value = "=(" + str(pacificMaleBehavioral) + "+" + str(whiteMaleBehavioral) + "+" + str(nativeMaleBehavioral) + "+" + str(hispanicMaleBehavioral) + "+" + str(blackMaleBehavioral) + "+" + str(asianMaleBehavioral) + ")"
        Dash.Range('AB16').Value = "=(" + str(pacificMaleAcademics) + "+" + str(whiteMaleAcademics) + "+" + str(nativeMaleAcademics) + "+" + str(hispanicMaleAcademics) + "+" + str(blackMaleAcademics) + "+" + str(asianMaleAcademics) + ")"
        Dash.Range('AD16').Value = "=(" + str(pacificMaleSocial) + "+" + str(whiteMaleSocial) + "+" + str(nativeMaleSocial) + "+" + str(hispanicMaleSocial) + "+" + str(blackMaleSocial) + "+" + str(asianMaleSocial) + ")"
        Dash.Range('AF16').Value = "=(" + str(pacificMaleAttendance) + "+" + str(whiteMaleAttendance) + "+" + str(nativeMaleAttendance) + "+" + str(hispanicMaleAttendance) + "+" + str(blackMaleAttendance) + "+" + str(asianMaleAttendance) + ")"

        Dash.Range('Z18').Value = "=(" + str(pacificFemaleBehavioral) + "+" + str(whiteFemaleBehavioral) + "+" + str(nativeFemaleBehavioral) + "+" + str(hispanicFemaleBehavioral) + "+" + str(blackFemaleBehavioral) + "+" + str(asianFemaleBehavioral) + ")"
        Dash.Range('AB18').Value = "=(" + str(pacificFemaleAcademics) + "+" + str(whiteFemaleAcademics) + "+" + str(nativeFemaleAcademics) + "+" + str(hispanicFemaleAcademics) + "+" + str(blackFemaleAcademics) + "+" + str(asianFemaleAcademics) + ")"
        Dash.Range('AD18').Value = "=(" + str(pacificFemaleSocial) + "+" + str(whiteFemaleSocial) + "+" + str(nativeFemaleSocial) + "+" + str(hispanicFemaleSocial) + "+" + str(blackFemaleSocial) + "+" + str(asianFemaleSocial) + ")"
        Dash.Range('AF18').Value = "=(" + str(pacificFemaleAttendance) + "+" + str(whiteFemaleAttendance) + "+" + str(nativeFemaleAttendance) + "+" + str(hispanicFemaleAttendance) + "+" + str(blackFemaleAttendance) + "+" + str(asianFemaleAttendance) + ")"

        #G/S 3
        Dash.Range('Z23').Value = "=(" + str(maleServiced) + "/(" + str(male) + "+" + str(female) + "))"
        Dash.Range('AB23').Value = maleServiced
        Dash.Range('Z23').NumberFormat = "0.00%"
        Dash.Range('Z25').Value = "=(" + str(femaleServiced) + "/(" + str(male) + "+" + str(female) + "))"
        Dash.Range('AB25').Value = femaleServiced
        Dash.Range('Z25').NumberFormat = "0.00%"

        Dash.Range('AD23').Value = male
        Dash.Range('AD25').Value = female

        Dash.Range('AF23').Value = "=((" + str(male) + ")/(" + str(male) + "+" + str(female) + "))"
        Dash.Range('AF23').NumberFormat = "0.00%"
        Dash.Range('AF25').Value = "=((" + str(female) + ")/(" + str(male) + "+" + str(female) + "))"
        Dash.Range('AF25').NumberFormat = "0.00%"

        #G/S 4
        Dash.Range('Z30').Value = asianMale
        Dash.Range('AB30').Value = "=(" + str(asianMale) + ")/(" + str(asianMale) + "+" + str(blackMale) + "+" + str(hispanicMale) + "+" + str(nativeMale) + "+" + str(whiteMale) + "+" + str(pacificMale) + ")"
        Dash.Range('AB30').NumberFormat = "0.00%"
        Dash.Range('Z32').Value = blackMale
        Dash.Range('AB32').Value = "=(" + str(blackMale) + ")/(" + str(asianMale) + "+" + str(blackMale) + "+" + str(hispanicMale) + "+" + str(nativeMale) + "+" + str(whiteMale) + "+" + str(pacificMale) + ")"
        Dash.Range('AB32').NumberFormat = "0.00%"
        Dash.Range('Z34').Value = hispanicMale
        Dash.Range('AB34').Value = "=(" + str(hispanicMale) + ")/(" + str(asianMale) + "+" + str(blackMale) + "+" + str(hispanicMale) + "+" + str(nativeMale) + "+" + str(whiteMale) + "+" + str(pacificMale) + ")"
        Dash.Range('AB34').NumberFormat = "0.00%"
        Dash.Range('Z36').Value = nativeMale
        Dash.Range('AB36').Value = "=(" + str(nativeMale) + ")/(" + str(asianMale) + "+" + str(blackMale) + "+" + str(hispanicMale) + "+" + str(nativeMale) + "+" + str(whiteMale) + "+" + str(pacificMale) + ")"
        Dash.Range('AB36').NumberFormat = "0.00%"
        Dash.Range('Z38').Value = whiteMale
        Dash.Range('AB38').Value = "=(" + str(whiteMale) + ")/(" + str(asianMale) + "+" + str(blackMale) + "+" + str(hispanicMale) + "+" + str(nativeMale) + "+" + str(whiteMale) + "+" + str(pacificMale) + ")"
        Dash.Range('AB38').NumberFormat = "0.00%"
        Dash.Range('Z40').Value = pacificMale
        Dash.Range('AB40').Value = "=(" + str(pacificMale) + ")/(" + str(asianMale) + "+" + str(blackMale) + "+" + str(hispanicMale) + "+" + str(nativeMale) + "+" + str(whiteMale) + "+" + str(pacificMale) + ")"
        Dash.Range('AB40').NumberFormat = "0.00%"
        Dash.Range('AD30').Value = asianFemale
        Dash.Range('AF30').Value = "=(" + str(asianFemale) + ")/(" + str(asianFemale) + "+" + str(blackFemale) + "+" + str(hispanicFemale) + "+" + str(nativeFemale) + "+" + str(whiteFemale) + "+" + str(pacificFemale) + ")"
        Dash.Range('AF30').NumberFormat = "0.00%"
        Dash.Range('AD32').Value = blackFemale
        Dash.Range('AF32').Value = "=(" + str(blackFemale) + ")/(" + str(asianFemale) + "+" + str(blackFemale) + "+" + str(hispanicFemale) + "+" + str(nativeFemale) + "+" + str(whiteFemale) + "+" + str(pacificFemale) + ")"
        Dash.Range('AF32').NumberFormat = "0.00%"
        Dash.Range('AD34').Value = hispanicFemale
        Dash.Range('AF34').Value = "=(" + str(hispanicFemale) + ")/(" + str(asianFemale) + "+" + str(blackFemale) + "+" + str(hispanicFemale) + "+" + str(nativeFemale) + "+" + str(whiteFemale) + "+" + str(pacificFemale) + ")"
        Dash.Range('AF34').NumberFormat = "0.00%"
        Dash.Range('AD36').Value = nativeFemale
        Dash.Range('AF36').Value = "=(" + str(nativeFemale) + ")/(" + str(asianFemale) + "+" + str(blackFemale) + "+" + str(hispanicFemale) + "+" + str(nativeFemale) + "+" + str(whiteFemale) + "+" + str(pacificFemale) + ")"
        Dash.Range('AF36').NumberFormat = "0.00%"
        Dash.Range('AD38').Value = whiteFemale
        Dash.Range('AF38').Value = "=(" + str(whiteFemale) + ")/(" + str(asianFemale) + "+" + str(blackFemale) + "+" + str(hispanicFemale) + "+" + str(nativeFemale) + "+" + str(whiteFemale) + "+" + str(pacificFemale) + ")"
        Dash.Range('AF38').NumberFormat = "0.00%"
        Dash.Range('AD40').Value = pacificFemale
        Dash.Range('AF40').Value = "=(" + str(pacificFemale) + ")/(" + str(asianFemale) + "+" + str(blackFemale) + "+" + str(hispanicFemale) + "+" + str(nativeFemale) + "+" + str(whiteFemale) + "+" + str(pacificFemale) + ")"
        Dash.Range('AF40').NumberFormat = "0.00%"

        #Outcome Dashboard

        Dash.Range('Q50').Value = AAF
        Dash.Range('Q48').Value = AAM
        Dash.Range('Q54').Value = ABF
        Dash.Range('Q52').Value = ABM
        Dash.Range('Q58').Value = AHF
        Dash.Range('Q56').Value = AHM
        Dash.Range('Q62').Value = AIF
        Dash.Range('Q60').Value = AIM
        Dash.Range('Q66').Value = AWF
        Dash.Range('Q64').Value = AWM
        Dash.Range('Q70').Value = APF
        Dash.Range('Q68').Value = APM

        Dash.Range('D50').Value = ARAF
        Dash.Range('D48').Value = ARAM
        Dash.Range('D54').Value = ARBF
        Dash.Range('D52').Value = ARBM
        Dash.Range('D58').Value = ARHF
        Dash.Range('D56').Value = ARHM
        Dash.Range('D62').Value = ARIF
        Dash.Range('D60').Value = ARIM
        Dash.Range('D66').Value = ARWF
        Dash.Range('D64').Value = ARWM
        Dash.Range('D70').Value = ARPF
        Dash.Range('D68').Value = ARPM

        Dash.Range('X50').Value = BNAF
        Dash.Range('X48').Value = BNAM
        Dash.Range('X54').Value = BNBF
        Dash.Range('X52').Value = BNBM
        Dash.Range('X58').Value = BNHF
        Dash.Range('X56').Value = BNHM
        Dash.Range('X62').Value = BNIF
        Dash.Range('X60').Value = BNIM
        Dash.Range('X66').Value = BNWF
        Dash.Range('X64').Value = BNWM
        Dash.Range('X70').Value = BNPF
        Dash.Range('X68').Value = BNPM

        Dash.Range('D77').Value = CCAF
        Dash.Range('D75').Value = CCAM
        Dash.Range('D81').Value = CCBF
        Dash.Range('D79').Value = CCBM
        Dash.Range('D85').Value = CCHF
        Dash.Range('D83').Value = CCHM
        Dash.Range('D89').Value = CCIF
        Dash.Range('D87').Value = CCIM
        Dash.Range('D93').Value = CCWF
        Dash.Range('D91').Value = CCWM
        Dash.Range('D97').Value = CCPF
        Dash.Range('D95').Value = CCPM

        Dash.Range('J50').Value = CPAF
        Dash.Range('J48').Value = CPAM
        Dash.Range('J54').Value = CPBF
        Dash.Range('J52').Value = CPBM
        Dash.Range('J58').Value = CPHF
        Dash.Range('J56').Value = CPHM
        Dash.Range('J62').Value = CPIF
        Dash.Range('J60').Value = CPIM
        Dash.Range('J66').Value = CPWF
        Dash.Range('J64').Value = CPWM
        Dash.Range('J70').Value = CPPF
        Dash.Range('J68').Value = CPPM

        Dash.Range('AD50').Value = CCRAF
        Dash.Range('AD48').Value = CCRAM
        Dash.Range('AD54').Value = CCRBF
        Dash.Range('AD52').Value = CCRBM
        Dash.Range('AD58').Value = CCRHF
        Dash.Range('AD56').Value = CCRHM
        Dash.Range('AD62').Value = CCRIF
        Dash.Range('AD60').Value = CCRIM
        Dash.Range('AD66').Value = CCRWF
        Dash.Range('AD64').Value = CCRWM
        Dash.Range('AD70').Value = CCRPF
        Dash.Range('AD68').Value = CCRPM

        Dash.Range('Q77').Value = DCAF
        Dash.Range('Q75').Value = DCAM
        Dash.Range('Q81').Value = DCBF
        Dash.Range('Q79').Value = DCBM
        Dash.Range('Q85').Value = DCHF
        Dash.Range('Q83').Value = DCHM
        Dash.Range('Q89').Value = DCIF
        Dash.Range('Q87').Value = DCIM
        Dash.Range('Q93').Value = DCWF
        Dash.Range('Q91').Value = DCWM
        Dash.Range('Q97').Value = DCPF
        Dash.Range('Q95').Value = DCPM

        Dash.Range('J77').Value = FCAF
        Dash.Range('J75').Value = FCAM
        Dash.Range('J81').Value = FCBF
        Dash.Range('J79').Value = FCBM
        Dash.Range('J85').Value = FCHF
        Dash.Range('J83').Value = FCHM
        Dash.Range('J89').Value = FCIF
        Dash.Range('J87').Value = FCIM
        Dash.Range('J93').Value = FCWF
        Dash.Range('J91').Value = FCWM
        Dash.Range('J97').Value = FCPF
        Dash.Range('J95').Value = FCPM

        Dash.Range('F50').Value = GAF
        Dash.Range('F48').Value = GAM
        Dash.Range('F54').Value = GBF
        Dash.Range('F52').Value = GBM
        Dash.Range('F56').Value = GHF
        Dash.Range('F58').Value = GHM
        Dash.Range('F62').Value = GIF
        Dash.Range('F60').Value = GIM
        Dash.Range('F66').Value = GWF
        Dash.Range('F64').Value = GWM
        Dash.Range('F70').Value = GPF
        Dash.Range('F68').Value = GPM

        Dash.Range('M77').Value = GLAF
        Dash.Range('M75').Value = GLAM
        Dash.Range('M81').Value = GLBF
        Dash.Range('M79').Value = GLBM
        Dash.Range('M85').Value = GLHF
        Dash.Range('M83').Value = GLHM
        Dash.Range('M89').Value = GLIF
        Dash.Range('M87').Value = GLIM
        Dash.Range('M93').Value = GLWF
        Dash.Range('M91').Value = GLWM
        Dash.Range('M97').Value = GLPF
        Dash.Range('M95').Value = GLPM

        Dash.Range('O50').Value = HCAF
        Dash.Range('O48').Value = HCAM
        Dash.Range('O54').Value = HCBF
        Dash.Range('O52').Value = HCBM
        Dash.Range('O58').Value = HCHF
        Dash.Range('O56').Value = HCHM
        Dash.Range('O62').Value = HCIF
        Dash.Range('O60').Value = HCIM
        Dash.Range('O66').Value = HCWF
        Dash.Range('O64').Value = HCWM
        Dash.Range('O70').Value = HCPF
        Dash.Range('O68').Value = HCPM

        Dash.Range('M50').Value = LDAF
        Dash.Range('M48').Value = LDAM
        Dash.Range('M54').Value = LDBF
        Dash.Range('M52').Value = LDBM
        Dash.Range('M58').Value = LDHF
        Dash.Range('M56').Value = LDHM
        Dash.Range('M62').Value = LDIF
        Dash.Range('M60').Value = LDIM
        Dash.Range('M66').Value = LDWF
        Dash.Range('M64').Value = LDWM
        Dash.Range('M70').Value = LDPF
        Dash.Range('M68').Value = LDPM

        Dash.Range('Z50').Value = LSAF
        Dash.Range('Z48').Value = LSAM
        Dash.Range('Z54').Value = LSBF
        Dash.Range('Z52').Value = LSBM
        Dash.Range('Z58').Value = LSHF
        Dash.Range('Z56').Value = LSHM
        Dash.Range('Z62').Value = LSIF
        Dash.Range('Z60').Value = LSIM
        Dash.Range('Z66').Value = LSWF
        Dash.Range('Z64').Value = LSWM
        Dash.Range('Z70').Value = LSPF
        Dash.Range('Z68').Value = LSPM

        Dash.Range('U77').Value = MHCAF
        Dash.Range('U75').Value = MHCAM
        Dash.Range('U81').Value = MHCBF
        Dash.Range('U79').Value = MHCBM
        Dash.Range('U85').Value = MHCHF
        Dash.Range('U83').Value = MHCHM
        Dash.Range('U89').Value = MHCIF
        Dash.Range('U87').Value = MHCIM
        Dash.Range('U93').Value = MHCWF
        Dash.Range('U91').Value = MHCWM
        Dash.Range('U97').Value = MHCPF
        Dash.Range('U95').Value = MHCPM

        Dash.Range('H77').Value = MHWAF
        Dash.Range('H75').Value = MHWAM
        Dash.Range('H81').Value = MHWBF
        Dash.Range('H79').Value = MHWBM
        Dash.Range('H85').Value = MHWHF
        Dash.Range('H83').Value = MHWHM
        Dash.Range('H89').Value = MHWIF
        Dash.Range('H87').Value = MHWIM
        Dash.Range('H93').Value = MHWWF
        Dash.Range('H91').Value = MHWWM
        Dash.Range('H97').Value = MHWPF
        Dash.Range('H95').Value = MHWPM

        Dash.Range('U50').Value = RISAF
        Dash.Range('U48').Value = RISAM
        Dash.Range('U54').Value = RISBF
        Dash.Range('U52').Value = RISBM
        Dash.Range('U58').Value = RISHF
        Dash.Range('U56').Value = RISHM
        Dash.Range('U62').Value = RISIF
        Dash.Range('U60').Value = RISIM
        Dash.Range('U66').Value = RISWF
        Dash.Range('U64').Value = RISWM
        Dash.Range('U70').Value = RISPF
        Dash.Range('U68').Value = RISPM

        Dash.Range('S77').Value = RDMAF
        Dash.Range('S75').Value = RDMAM
        Dash.Range('S81').Value = RDMBF
        Dash.Range('S79').Value = RDMBM
        Dash.Range('S85').Value = RDMHF
        Dash.Range('S83').Value = RDMHM
        Dash.Range('S89').Value = RDMIF
        Dash.Range('S87').Value = RDMIM
        Dash.Range('S93').Value = RDMWF
        Dash.Range('S91').Value = RDMWM
        Dash.Range('S97').Value = RDMPF
        Dash.Range('S95').Value = RDMPM

        Dash.Range('O77').Value = SEMAF
        Dash.Range('O75').Value = SEMAM
        Dash.Range('O81').Value = SEMBF
        Dash.Range('O79').Value = SEMBM
        Dash.Range('O85').Value = SEMHF
        Dash.Range('O83').Value = SEMHM
        Dash.Range('O89').Value = SEMIF
        Dash.Range('O87').Value = SEMIM
        Dash.Range('O93').Value = SEMWF
        Dash.Range('O91').Value = SEMWM
        Dash.Range('O97').Value = SEMPF
        Dash.Range('O95').Value = SEMPM

        Dash.Range('AF50').Value = SEAF
        Dash.Range('AF48').Value = SEAM
        Dash.Range('AF54').Value = SEBF
        Dash.Range('AF52').Value = SEBM
        Dash.Range('AF58').Value = SEHF
        Dash.Range('AF56').Value = SEHM
        Dash.Range('AF62').Value = SEIF
        Dash.Range('AF60').Value = SEIM
        Dash.Range('AF66').Value = SEWF
        Dash.Range('AF64').Value = SEWM
        Dash.Range('AF70').Value = SEPF
        Dash.Range('AF68').Value = SEPM

        Dash.Range('F77').Value = SRAF
        Dash.Range('F75').Value = SRAM
        Dash.Range('F81').Value = SRBF
        Dash.Range('F79').Value = SRBM
        Dash.Range('F85').Value = SRHF
        Dash.Range('F83').Value = SRHM
        Dash.Range('F89').Value = SRIF
        Dash.Range('F87').Value = SRIM
        Dash.Range('F93').Value = SRWF
        Dash.Range('F91').Value = SRWM
        Dash.Range('F97').Value = SRPF
        Dash.Range('F95').Value = SRPM

        Dash.Range('AB50').Value = SSAF
        Dash.Range('AB48').Value = SSAM
        Dash.Range('AB54').Value = SSBF
        Dash.Range('AB52').Value = SSBM
        Dash.Range('AB58').Value = SSHF
        Dash.Range('AB56').Value = SSHM
        Dash.Range('AB62').Value = SSIF
        Dash.Range('AB60').Value = SSIM
        Dash.Range('AB66').Value = SSWF
        Dash.Range('AB64').Value = SSWM
        Dash.Range('AB70').Value = SSPF
        Dash.Range('AB68').Value = SSPM

        Dash.Range('S50').Value = TAF
        Dash.Range('S48').Value = TAM
        Dash.Range('S54').Value = TBF
        Dash.Range('S52').Value = TBM
        Dash.Range('S58').Value = THF
        Dash.Range('S56').Value = THM
        Dash.Range('S62').Value = TIF
        Dash.Range('S60').Value = TIM
        Dash.Range('S66').Value = TWF
        Dash.Range('S64').Value = TWM
        Dash.Range('S70').Value = TPF
        Dash.Range('S68').Value = TPM

        Dash.Range('H50').Value = TRAF
        Dash.Range('H48').Value = TRAM
        Dash.Range('H54').Value = TRBF
        Dash.Range('H52').Value = TRBM
        Dash.Range('H58').Value = TRHF
        Dash.Range('H56').Value = TRHM
        Dash.Range('H62').Value = TRIF
        Dash.Range('H60').Value = TRIM
        Dash.Range('H66').Value = TRWF
        Dash.Range('H64').Value = TRWM
        Dash.Range('H70').Value = TRPF
        Dash.Range('H68').Value = TRPM



        target.Close(SaveChanges=True)
        xlApp.Quit()


        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        target2 = xlApp.Workbooks.Open(pathOne)
        Mix = target.Worksheets(locationTwo)
        Assess = target.Worksheets(locationOne)
        Campus = target.Worksheets(locationFour)
        Dash = target.Worksheets("Dashboard")

        chilton = 0
        engeWashington = 0
        groesbeckHigh = 0
        groesbeckMiddle = 0
        whitehurst = 0
        laVegaEl = 0
        laVegaHS = 0
        laVegaInt = 0
        laVegaJH = 0
        laVegaPri = 0
        mexiaHS = 0
        mexiaJH = 0
        castelemanCreek = 0
        hewitt = 0
        midwayHS = 0
        midwayMiddle = 0
        quinn = 0
        teague = 0
        wacoCharter = 0
        altaVista = 0
        brazosHS = 0
        brookAvenue = 0
        cesarChavez = 0
        carver = 0
        hines = 0
        kendrick = 0
        providentHeights = 0
        tennyson = 0
        university = 0
        wacoHS = 0

        schoolMonth = 1
        columnticker = 1
        rowticker = 1
        dashTicker = 103
        columnmarker = ""
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        campusTicker = columnticker
        valueTransfer = 0

        while(schoolMonth != 2):
            if(schoolMonth == 1):
                if(dashTicker >= 163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in August" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('D'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('D'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 3):
            if(schoolMonth == 2):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in September" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('F'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('F'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;


        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 4):
            if(schoolMonth == 3):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in October" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('H'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('H'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 5):    
            if(schoolMonth == 4):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in November" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('J'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('J'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;
            
        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 6):
            if(schoolMonth == 5):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in December" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('M'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('M'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 7):    
            if(schoolMonth == 6):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in January" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('O'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('O'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 8):    
            if(schoolMonth == 7):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in February" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('Q'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('Q'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 9):
            if(schoolMonth == 8):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in March" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('S'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('S'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 10):
            if(schoolMonth == 9):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in April" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('U'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('U'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 11):
            if(schoolMonth == 10):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in May" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('X'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('X'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 12):
            if(schoolMonth == 11):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in June" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('Z'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('Z'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 103
        while(schoolMonth != 13):
            if(schoolMonth == 12):
                if(dashTicker >=163):
                    schoolMonth += 1
                    break
                if("Number of Services Entered in July" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('AB'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('AB'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        chiltonAssessed = 0
        chiltonComplete = 0
        chiltonEnrolled = 0
        chiltonInProgramTransfer = 0
        chiltonInactive = 0
        chiltonOutOfProgramTransfer = 0
        chiltonProgress = 0
        chiltonRegistered = 0

        engeWashingtonAssessed = 0
        engeWashingtonComplete = 0
        engeWashingtonEnrolled = 0
        engeWashingtonInProgramTransfer = 0
        engeWashingtonInactive = 0
        engeWashingtonOutOfProgramTransfer = 0
        engeWashingtonProgress = 0
        engeWashingtonRegistered = 0

        groesbeckHighAssessed = 0
        groesbeckHighComplete = 0
        groesbeckHighEnrolled = 0
        groesbeckHighInProgramTransfer = 0
        groesbeckHighInactive = 0
        groesbeckHighOutOfProgramTransfer = 0
        groesbeckHighProgress = 0
        groesbeckHighRegistered = 0

        groesbeckMiddleAssessed = 0
        groesbeckMiddleComplete = 0
        groesbeckMiddleEnrolled = 0
        groesbeckMiddleInProgramTransfer = 0
        groesbeckMiddleInactive = 0
        groesbeckMiddleOutOfProgramTransfer = 0
        groesbeckMiddleProgress = 0
        groesbeckMiddleRegistered = 0

        whitehurstAssessed = 0
        whitehurstComplete = 0
        whitehurstEnrolled = 0
        whitehurstInProgramTransfer = 0
        whitehurstInactive = 0
        whitehurstOutOfProgramTransfer = 0
        whitehurstProgress = 0
        whitehurstRegistered = 0

        laVegaElAssessed = 0
        laVegaElComplete = 0
        laVegaElEnrolled = 0
        laVegaElInProgramTransfer = 0
        laVegaElInactive = 0
        laVegaElOutOfProgramTransfer = 0
        laVegaElProgress = 0
        laVegaElRegistered = 0

        laVegaHSAssessed = 0
        laVegaHSComplete = 0
        laVegaHSEnrolled = 0
        laVegaHSInProgramTransfer = 0
        laVegaHSInactive = 0
        laVegaHSOutOfProgramTransfer = 0
        laVegaHSProgress = 0
        laVegaHSRegistered = 0

        laVegaIntAssessed = 0
        laVegaIntComplete = 0
        laVegaIntEnrolled = 0
        laVegaIntInProgramTransfer = 0
        laVegaIntInactive = 0
        laVegaIntOutOfProgramTransfer = 0
        laVegaIntProgress = 0
        laVegaIntRegistered = 0

        laVegaJHAssessed = 0
        laVegaJHComplete = 0
        laVegaJHEnrolled = 0
        laVegaJHInProgramTransfer = 0
        laVegaJHInactive = 0
        laVegaJHOutOfProgramTransfer = 0
        laVegaJHProgress = 0
        laVegaJHRegistered = 0

        laVegaPriAssessed = 0
        laVegaPriComplete = 0
        laVegaPriEnrolled = 0
        laVegaPriInProgramTransfer = 0
        laVegaPriInactive = 0
        laVegaPriOutOfProgramTransfer = 0
        laVegaPriProgress = 0
        laVegaPriRegistered = 0

        mexiaHSAssessed = 0
        mexiaHSComplete = 0
        mexiaHSEnrolled = 0
        mexiaHSInProgramTransfer = 0
        mexiaHSInactive = 0
        mexiaHSOutOfProgramTransfer = 0
        mexiaHSProgress = 0
        mexiaHSRegistered = 0

        mexiaJHAssessed = 0
        mexiaJHComplete = 0
        mexiaJHEnrolled = 0
        mexiaJHInProgramTransfer = 0
        mexiaJHInactive = 0
        mexiaJHOutOfProgramTransfer = 0
        mexiaJHProgress = 0
        mexiaJHRegistered = 0

        castelemanCreekAssessed = 0
        castelemanCreekComplete = 0
        castelemanCreekEnrolled = 0
        castelemanCreekInProgramTransfer = 0
        castelemanCreekInactive = 0
        castelemanCreekOutOfProgramTransfer = 0
        castelemanCreekProgress = 0
        castelemanCreekRegistered = 0

        hewittAssessed = 0
        hewittComplete = 0
        hewittEnrolled = 0
        hewittInProgramTransfer = 0
        hewittInactive = 0
        hewittOutOfProgramTransfer = 0
        hewittProgress = 0
        hewittRegistered = 0

        midwayHSAssessed = 0
        midwayHSComplete = 0
        midwayHSEnrolled = 0
        midwayHSInProgramTransfer = 0
        midwayHSInactive = 0
        midwayHSOutOfProgramTransfer = 0
        midwayHSProgress = 0
        midwayHSRegistered = 0

        midwayMiddleAssessed = 0
        midwayMiddleComplete = 0
        midwayMiddleEnrolled = 0
        midwayMiddleInProgramTransfer = 0
        midwayMiddleInactive = 0
        midwayMiddleOutOfProgramTransfer = 0
        midwayMiddleProgress = 0
        midwayMiddleRegistered = 0

        quinnAssessed = 0
        quinnComplete = 0
        quinnEnrolled = 0
        quinnInProgramTransfer = 0
        quinnInactive = 0
        quinnOutOfProgramTransfer = 0
        quinnProgress = 0
        quinnRegistered = 0

        teagueAssessed = 0
        teagueComplete = 0
        teagueEnrolled = 0
        teagueInProgramTransfer = 0
        teagueInactive = 0
        teagueOutOfProgramTransfer = 0
        teagueProgress = 0
        teagueRegistered = 0

        wacoCharterAssessed = 0
        wacoCharterComplete = 0
        wacoCharterEnrolled = 0
        wacoCharterInProgramTransfer = 0
        wacoCharterInactive = 0
        wacoCharterOutOfProgramTransfer = 0
        wacoCharterProgress = 0
        wacoCharterRegistered = 0

        altaVistaAssessed = 0
        altaVistaComplete = 0
        altaVistaEnrolled = 0
        altaVistaInProgramTransfer = 0
        altaVistaInactive = 0
        altaVistaOutOfProgramTransfer = 0
        altaVistaProgress = 0
        altaVistaRegistered = 0

        brazosHSAssessed = 0
        brazosHSComplete = 0
        brazosHSEnrolled = 0
        brazosHSInProgramTransfer = 0
        brazosHSInactive = 0
        brazosHSOutOfProgramTransfer = 0
        brazosHSProgress = 0
        brazosHSRegistered = 0

        brookAvenueAssessed = 0
        brookAvenueComplete = 0
        brookAvenueEnrolled = 0
        brookAvenueInProgramTransfer = 0
        brookAvenueInactive = 0
        brookAvenueOutOfProgramTransfer = 0
        brookAvenueProgress = 0
        brookAvenueRegistered = 0

        cesarChavezAssessed = 0
        cesarChavezComplete = 0
        cesarChavezEnrolled = 0
        cesarChavezInProgramTransfer = 0
        cesarChavezInactive = 0
        cesarChavezOutOfProgramTransfer = 0
        cesarChavezProgress = 0
        cesarChavezRegistered = 0

        carverAssessed = 0
        carverComplete = 0
        carverEnrolled = 0
        carverInProgramTransfer = 0
        carverInactive = 0
        carverOutOfProgramTransfer = 0
        carverProgress = 0
        carverRegistered = 0

        hinesAssessed = 0
        hinesComplete = 0
        hinesEnrolled = 0
        hinesInProgramTransfer = 0
        hinesInactive = 0
        hinesOutOfProgramTransfer = 0
        hinesProgress = 0
        hinesRegistered = 0

        kendrickAssessed = 0
        kendrickComplete = 0
        kendrickEnrolled = 0
        kendrickInProgramTransfer = 0
        kendrickInactive = 0
        kendrickOutOfProgramTransfer = 0
        kendrickProgress = 0
        kendrickRegistered = 0

        providentHeightsAssessed = 0
        providentHeightsComplete = 0
        providentHeightsEnrolled = 0
        providentHeightsInProgramTransfer = 0
        providentHeightsInactive = 0
        providentHeightsOutOfProgramTransfer = 0
        providentHeightsProgress = 0
        providentHeightsRegistered = 0

        tennysonAssessed = 0
        tennysonComplete = 0
        tennysonEnrolled = 0
        tennysonInProgramTransfer = 0
        tennysonInactive = 0
        tennysonOutOfProgramTransfer = 0
        tennysonProgress = 0
        tennysonRegistered = 0

        universityAssessed = 0
        universityComplete = 0
        universityEnrolled = 0
        universityInProgramTransfer = 0
        universityInactive = 0
        universityOutOfProgramTransfer = 0
        universityProgress = 0
        universityRegistered = 0

        wacoHSAssessed = 0
        wacoHSComplete = 0
        wacoHSEnrolled = 0
        wacoHSInProgramTransfer = 0
        wacoHSInactive = 0
        wacoHSOutOfProgramTransfer = 0
        wacoHSProgress = 0
        wacoHSRegistered = 0

        Casefile = 2
        DashTicker = 168
        while((str(Assess.Range('C'+str(Casefile)).Value) != 'None')):
            if("Progress" in str(Assess.Range('C'+str(Casefile)).Value)):
                if("CHILTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    chiltonProgress += 1
                if("ENGE-WASHINGTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    engeWashingtonProgress += 1
                if("GROESBECK H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckHighProgress += 1
                if("GROESBECK MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckMiddleProgress += 1
                if("H O WHITEHURST" in str(Assess.Range('A'+str(Casefile)).Value)):
                    whitehurstProgress += 1
                if("LA VEGA EL " in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaElProgress += 1
                if("LA VEGA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaHSProgress += 1
                if("LA VEGA INT" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaIntProgress += 1
                if("LA VEGA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaJHProgress += 1
                if("LA VEGA PRI" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaPriProgress += 1
                if("MEXIA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaHSProgress += 1
                if("MEXIA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaJHProgress += 1
                if("CASTLEMAN CREEK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    castelemanCreekProgress += 1
                if("HEWITT EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hewittProgress += 1
                if("MIDWAY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayHSProgress += 1
                if("MIDWAY MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayMiddleProgress += 1
                if("QUINN" in str(Assess.Range('A'+str(Casefile)).Value)):
                    quinnProgress += 1
                if("TEAGUE H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    teagueProgress += 1
                if("WACO CHARTER" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoCharterProgress += 1
                if("ALTA VISTA EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    altaVistaProgress += 1
                if("BRAZOS H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brazosHSProgress += 1
                if("BROOK AVENUE EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brookAvenueProgress += 1
                if("CESAR CHAVEZ MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    cesarChavezProgress += 1
                if("G W CARVER MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    carverProgress += 1
                if("J H HINES EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hinesProgress += 1
                if("KENDRICK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    kendrickProgress += 1
                if("PROVIDENT HEIGHTS EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    providentHeightsProgress += 1
                if("TENNYSON MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    tennysonProgress += 1
                if("UNIVERSITY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    universityProgress += 1
                if("WACO H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoHSProgress += 1
            if("Assessed" in str(Assess.Range('C'+str(Casefile)).Value)):
                if("CHILTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    chiltonAssessed += 1
                if("ENGE-WASHINGTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    engeWashingtonAssessed += 1
                if("GROESBECK H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckHighAssessed += 1
                if("GROESBECK MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckMiddleAssessed += 1
                if("H O WHITEHURST" in str(Assess.Range('A'+str(Casefile)).Value)):
                    whitehurstAssessed += 1
                if("LA VEGA EL " in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaElAssessed += 1
                if("LA VEGA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaHSAssessed += 1
                if("LA VEGA INT" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaIntAssessed += 1
                if("LA VEGA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaJHAssessed += 1
                if("LA VEGA PRI" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaPriAssessed += 1
                if("MEXIA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaHSAssessed += 1
                if("MEXIA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaJHAssessed += 1
                if("CASTLEMAN CREEK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    castelemanCreekAssessed += 1
                if("HEWITT EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hewittAssessed += 1
                if("MIDWAY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayHSAssessed += 1
                if("MIDWAY MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayMiddleAssessed += 1
                if("QUINN" in str(Assess.Range('A'+str(Casefile)).Value)):
                    quinnAssessed += 1
                if("TEAGUE H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    teagueAssessed += 1
                if("WACO CHARTER" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoCharterAssessed += 1
                if("ALTA VISTA EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    altaVistaAssessed += 1
                if("BRAZOS H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brazosHSAssessed += 1
                if("BROOK AVENUE EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brookAvenueAssessed += 1
                if("CESAR CHAVEZ MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    cesarChavezAssessed += 1
                if("G W CARVER MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    carverAssessed += 1
                if("J H HINES EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hinesAssessed += 1
                if("KENDRICK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    kendrickAssessed += 1
                if("PROVIDENT HEIGHTS EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    providentHeightsAssessed += 1
                if("TENNYSON MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    tennysonAssessed += 1
                if("UNIVERSITY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    universityAssessed += 1
                if("WACO H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoHSAssessed += 1
            if("Complete" in str(Assess.Range('C'+str(Casefile)).Value)):
                if("CHILTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    chiltonComplete += 1
                if("ENGE-WASHINGTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    engeWashingtonComplete += 1
                if("GROESBECK H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckHighComplete += 1
                if("GROESBECK MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckMiddleComplete += 1
                if("H O WHITEHURST" in str(Assess.Range('A'+str(Casefile)).Value)):
                    whitehurstComplete += 1
                if("LA VEGA EL " in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaElComplete += 1
                if("LA VEGA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaHSComplete += 1
                if("LA VEGA INT" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaIntComplete += 1
                if("LA VEGA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaJHComplete += 1
                if("LA VEGA PRI" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaPriComplete += 1
                if("MEXIA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaHSComplete += 1
                if("MEXIA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaJHComplete += 1
                if("CASTLEMAN CREEK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    castelemanCreekComplete += 1
                if("HEWITT EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hewittComplete += 1
                if("MIDWAY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayHSComplete += 1
                if("MIDWAY MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayMiddleComplete += 1
                if("QUINN" in str(Assess.Range('A'+str(Casefile)).Value)):
                    quinnComplete += 1
                if("TEAGUE H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    teagueComplete += 1
                if("WACO CHARTER" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoCharterComplete += 1
                if("ALTA VISTA EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    altaVistaComplete += 1
                if("BRAZOS H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brazosHSComplete += 1
                if("BROOK AVENUE EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brookAvenueComplete += 1
                if("CESAR CHAVEZ MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    cesarChavezComplete += 1
                if("G W CARVER MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    carverComplete += 1
                if("J H HINES EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hinesComplete += 1
                if("KENDRICK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    kendrickComplete += 1
                if("PROVIDENT HEIGHTS EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    providentHeightsComplete += 1
                if("TENNYSON MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    tennysonComplete += 1
                if("UNIVERSITY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    universityComplete += 1
                if("WACO H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoHSComplete += 1
            if("Enrolled" in str(Assess.Range('C'+str(Casefile)).Value)):
                if("CHILTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    chiltonEnrolled += 1
                if("ENGE-WASHINGTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    engeWashingtonEnrolled += 1
                if("GROESBECK H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckHighEnrolled += 1
                if("GROESBECK MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckMiddleEnrolled += 1
                if("H O WHITEHURST" in str(Assess.Range('A'+str(Casefile)).Value)):
                    whitehurstEnrolled += 1
                if("LA VEGA EL " in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaElEnrolled += 1
                if("LA VEGA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaHSEnrolled += 1
                if("LA VEGA INT" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaIntEnrolled += 1
                if("LA VEGA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaJHEnrolled += 1
                if("LA VEGA PRI" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaPriEnrolled += 1
                if("MEXIA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaHSEnrolled += 1
                if("MEXIA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaJHEnrolled += 1
                if("CASTLEMAN CREEK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    castelemanCreekEnrolled += 1
                if("HEWITT EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hewittEnrolled += 1
                if("MIDWAY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayHSEnrolled += 1
                if("MIDWAY MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayMiddleEnrolled += 1
                if("QUINN" in str(Assess.Range('A'+str(Casefile)).Value)):
                    quinnEnrolled += 1
                if("TEAGUE H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    teagueEnrolled += 1
                if("WACO CHARTER" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoCharterEnrolled += 1
                if("ALTA VISTA EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    altaVistaEnrolled += 1
                if("BRAZOS H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brazosHSEnrolled += 1
                if("BROOK AVENUE EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brookAvenueEnrolled += 1
                if("CESAR CHAVEZ MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    cesarChavezEnrolled += 1
                if("G W CARVER MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    carverEnrolled += 1
                if("J H HINES EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hinesEnrolled += 1
                if("KENDRICK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    kendrickEnrolled += 1
                if("PROVIDENT HEIGHTS EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    providentHeightsEnrolled += 1
                if("TENNYSON MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    tennysonEnrolled += 1
                if("UNIVERSITY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    universityEnrolled += 1
                if("WACO H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoHSEnrolled += 1
            if("In Program Transfer" in str(Assess.Range('C'+str(Casefile)).Value)):
                if("CHILTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    chiltonInProgramTransfer += 1
                if("ENGE-WASHINGTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    engeWashingtonInProgramTransfer += 1
                if("GROESBECK H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckHighInProgramTransfer += 1
                if("GROESBECK MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckMiddleInProgramTransfer += 1
                if("H O WHITEHURST" in str(Assess.Range('A'+str(Casefile)).Value)):
                    whitehurstInProgramTransfer += 1
                if("LA VEGA EL " in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaElInProgramTransfer += 1
                if("LA VEGA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaHSInProgramTransfer += 1
                if("LA VEGA INT" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaIntInProgramTransfer += 1
                if("LA VEGA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaJHInProgramTransfer += 1
                if("LA VEGA PRI" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaPriInProgramTransfer += 1
                if("MEXIA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaHSInProgramTransfer += 1
                if("MEXIA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaJHInProgramTransfer += 1
                if("CASTLEMAN CREEK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    castelemanCreekInProgramTransfer += 1
                if("HEWITT EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hewittInProgramTransfer += 1
                if("MIDWAY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayHSInProgramTransfer += 1
                if("MIDWAY MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayMiddleInProgramTransfer += 1
                if("QUINN" in str(Assess.Range('A'+str(Casefile)).Value)):
                    quinnInProgramTransfer += 1
                if("TEAGUE H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    teagueInProgramTransfer += 1
                if("WACO CHARTER" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoCharterInProgramTransfer += 1
                if("ALTA VISTA EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    altaVistaInProgramTransfer += 1
                if("BRAZOS H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brazosHSInProgramTransfer += 1
                if("BROOK AVENUE EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brookAvenueInProgramTransfer += 1
                if("CESAR CHAVEZ MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    cesarChavezInProgramTransfer += 1
                if("G W CARVER MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    carverInProgramTransfer += 1
                if("J H HINES EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hinesInProgramTransfer += 1
                if("KENDRICK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    kendrickInProgramTransfer += 1
                if("PROVIDENT HEIGHTS EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    providentHeightsInProgramTransfer += 1
                if("TENNYSON MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    tennysonInProgramTransfer += 1
                if("UNIVERSITY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    universityInProgramTransfer += 1
                if("WACO H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoHSInProgramTransfer += 1
            if("Inactive" in str(Assess.Range('C'+str(Casefile)).Value)):
                if("CHILTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    chiltonInactive += 1
                if("ENGE-WASHINGTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    engeWashingtonInactive += 1
                if("GROESBECK H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckHighInactive += 1
                if("GROESBECK MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckMiddleInactive += 1
                if("H O WHITEHURST" in str(Assess.Range('A'+str(Casefile)).Value)):
                    whitehurstInactive += 1
                if("LA VEGA EL " in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaElInactive += 1
                if("LA VEGA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaHSInactive += 1
                if("LA VEGA INT" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaIntInactive += 1
                if("LA VEGA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaJHInactive += 1
                if("LA VEGA PRI" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaPriInactive += 1
                if("MEXIA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaHSInactive += 1
                if("MEXIA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaJHInactive += 1
                if("CASTLEMAN CREEK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    castelemanCreekInactive += 1
                if("HEWITT EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hewittInactive += 1
                if("MIDWAY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayHSInactive += 1
                if("MIDWAY MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayMiddleInactive += 1
                if("QUINN" in str(Assess.Range('A'+str(Casefile)).Value)):
                    quinnInactive += 1
                if("TEAGUE H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    teagueInactive += 1
                if("WACO CHARTER" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoCharterInactive += 1
                if("ALTA VISTA EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    altaVistaInactive += 1
                if("BRAZOS H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brazosHSInactive += 1
                if("BROOK AVENUE EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brookAvenueInactive += 1
                if("CESAR CHAVEZ MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    cesarChavezInactive += 1
                if("G W CARVER MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    carverInactive += 1
                if("J H HINES EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hinesInactive += 1
                if("KENDRICK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    kendrickInactive += 1
                if("PROVIDENT HEIGHTS EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    providentHeightsInactive += 1
                if("TENNYSON MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    tennysonInactive += 1
                if("UNIVERSITY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    universityInactive += 1
                if("WACO H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoHSInactive += 1
            if("Out of Program Transfer" in str(Assess.Range('C'+str(Casefile)).Value)):
                if("CHILTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    chiltonOutOfProgramTransfer += 1
                if("ENGE-WASHINGTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    engeWashingtonOutOfProgramTransfer += 1
                if("GROESBECK H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckHighOutOfProgramTransfer += 1
                if("GROESBECK MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckMiddleOutOfProgramTransfer += 1
                if("H O WHITEHURST" in str(Assess.Range('A'+str(Casefile)).Value)):
                    whitehurstOutOfProgramTransfer += 1
                if("LA VEGA EL " in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaElOutOfProgramTransfer += 1
                if("LA VEGA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaHSOutOfProgramTransfer += 1
                if("LA VEGA INT" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaIntOutOfProgramTransfer += 1
                if("LA VEGA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaJHOutOfProgramTransfer += 1
                if("LA VEGA PRI" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaPriOutOfProgramTransfer += 1
                if("MEXIA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaHSOutOfProgramTransfer += 1
                if("MEXIA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaJHOutOfProgramTransfer += 1
                if("CASTLEMAN CREEK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    castelemanCreekOutOfProgramTransfer += 1
                if("HEWITT EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hewittOutOfProgramTransfer += 1
                if("MIDWAY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayHSOutOfProgramTransfer += 1
                if("MIDWAY MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayMiddleOutOfProgramTransfer += 1
                if("QUINN" in str(Assess.Range('A'+str(Casefile)).Value)):
                    quinnOutOfProgramTransfer += 1
                if("TEAGUE H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    teagueOutOfProgramTransfer += 1
                if("WACO CHARTER" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoCharterOutOfProgramTransfer += 1
                if("ALTA VISTA EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    altaVistaOutOfProgramTransfer += 1
                if("BRAZOS H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brazosHSOutOfProgramTransfer += 1
                if("BROOK AVENUE EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brookAvenueOutOfProgramTransfer += 1
                if("CESAR CHAVEZ MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    cesarChavezOutOfProgramTransfer += 1
                if("G W CARVER MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    carverOutOfProgramTransfer += 1
                if("J H HINES EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hinesOutOfProgramTransfer += 1
                if("KENDRICK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    kendrickOutOfProgramTransfer += 1
                if("PROVIDENT HEIGHTS EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    providentHeightsOutOfProgramTransfer += 1
                if("TENNYSON MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    tennysonOutOfProgramTransfer += 1
                if("UNIVERSITY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    universityOutOfProgramTransfer += 1
                if("WACO H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoHSOutOfProgramTransfer += 1
            if("Registered" in str(Assess.Range('C'+str(Casefile)).Value)):
                if("CHILTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    chiltonRegistered += 1
                if("ENGE-WASHINGTON" in str(Assess.Range('A'+str(Casefile)).Value)):
                    engeWashingtonRegistered += 1
                if("GROESBECK H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckHighRegistered += 1
                if("GROESBECK MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    groesbeckMiddleRegistered += 1
                if("H O WHITEHURST" in str(Assess.Range('A'+str(Casefile)).Value)):
                    whitehurstRegistered += 1
                if("LA VEGA EL " in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaElRegistered += 1
                if("LA VEGA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaHSRegistered += 1
                if("LA VEGA INT" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaIntRegistered += 1
                if("LA VEGA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaJHRegistered += 1
                if("LA VEGA PRI" in str(Assess.Range('A'+str(Casefile)).Value)):
                    laVegaPriRegistered += 1
                if("MEXIA H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaHSRegistered += 1
                if("MEXIA J H" in str(Assess.Range('A'+str(Casefile)).Value)):
                    mexiaJHRegistered += 1
                if("CASTLEMAN CREEK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    castelemanCreekRegistered += 1
                if("HEWITT EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hewittRegistered += 1
                if("MIDWAY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayHSRegistered += 1
                if("MIDWAY MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    midwayMiddleRegistered += 1
                if("QUINN" in str(Assess.Range('A'+str(Casefile)).Value)):
                    quinnRegistered += 1
                if("TEAGUE H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    teagueRegistered += 1
                if("WACO CHARTER" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoCharterRegistered += 1
                if("ALTA VISTA EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    altaVistaRegistered += 1
                if("BRAZOS H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brazosHSRegistered += 1
                if("BROOK AVENUE EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    brookAvenueRegistered += 1
                if("CESAR CHAVEZ MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    cesarChavezRegistered += 1
                if("G W CARVER MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    carverRegistered += 1
                if("J H HINES EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    hinesRegistered += 1
                if("KENDRICK EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    kendrickRegistered += 1
                if("PROVIDENT HEIGHTS EL" in str(Assess.Range('A'+str(Casefile)).Value)):
                    providentHeightsRegistered += 1
                if("TENNYSON MIDDLE" in str(Assess.Range('A'+str(Casefile)).Value)):
                    tennysonRegistered += 1
                if("UNIVERSITY H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    universityRegistered += 1
                if("WACO H S" in str(Assess.Range('A'+str(Casefile)).Value)):
                    wacoHSRegistered += 1
            Casefile +=1
        DashTicker = 168
        while(str(Dash.Range('B'+str(DashTicker)).Value) != 'None'):
            if("CHILTON" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = chiltonAssessed;
                Dash.Range('F'+str(DashTicker)).Value = chiltonComplete;
                Dash.Range('H'+str(DashTicker)).Value = chiltonEnrolled;
                Dash.Range('J'+str(DashTicker)).Value = chiltonInProgramTransfer;
                Dash.Range('M'+str(DashTicker)).Value = chiltonInactive;
                Dash.Range('O'+str(DashTicker)).Value = chiltonOutOfProgramTransfer;
                Dash.Range('Q'+str(DashTicker)).Value = chiltonProgress;
                Dash.Range('S'+str(DashTicker)).Value = chiltonRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = chiltonComplete + chiltonProgress + chiltonOutOfProgramTransfer + chiltonInProgramTransfer; 
            if("ENGE-WASHINGTON INT" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = engeWashingtonAssessed;
                Dash.Range('F'+str(DashTicker)).Value = engeWashingtonComplete; 
                Dash.Range('H'+str(DashTicker)).Value = engeWashingtonEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = engeWashingtonInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = engeWashingtonInactive; 
                Dash.Range('O'+str(DashTicker)).Value = engeWashingtonOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = engeWashingtonProgress; 
                Dash.Range('S'+str(DashTicker)).Value = engeWashingtonRegistered; 
                Dash.Range('Z'+str(DashTicker)).Value = engeWashingtonComplete + engeWashingtonProgress + engeWashingtonOutOfProgramTransfer + engeWashingtonInProgramTransfer; 
            if("GROESBECK H S" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = groesbeckHighAssessed;
                Dash.Range('F'+str(DashTicker)).Value = groesbeckHighComplete; 
                Dash.Range('H'+str(DashTicker)).Value = groesbeckHighEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = groesbeckHighInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = groesbeckHighInactive; 
                Dash.Range('O'+str(DashTicker)).Value = groesbeckHighOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = groesbeckHighProgress; 
                Dash.Range('S'+str(DashTicker)).Value = groesbeckHighRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = groesbeckHighComplete + groesbeckHighProgress + groesbeckHighOutOfProgramTransfer + groesbeckHighInProgramTransfer; 
            if("GROESBECK MIDDLE" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = groesbeckMiddleAssessed;
                Dash.Range('F'+str(DashTicker)).Value = groesbeckMiddleComplete; 
                Dash.Range('H'+str(DashTicker)).Value = groesbeckMiddleEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = groesbeckMiddleInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = groesbeckMiddleInactive; 
                Dash.Range('O'+str(DashTicker)).Value = groesbeckMiddleOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = groesbeckMiddleProgress; 
                Dash.Range('S'+str(DashTicker)).Value = groesbeckMiddleRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = groesbeckMiddleComplete + groesbeckMiddleProgress + groesbeckMiddleOutOfProgramTransfer + groesbeckMiddleInProgramTransfer; 
            if("H O WHITEHURST EL" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = whitehurstAssessed;
                Dash.Range('F'+str(DashTicker)).Value = whitehurstComplete; 
                Dash.Range('H'+str(DashTicker)).Value = whitehurstEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = whitehurstInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = whitehurstInactive; 
                Dash.Range('O'+str(DashTicker)).Value = whitehurstOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = whitehurstProgress; 
                Dash.Range('S'+str(DashTicker)).Value = whitehurstRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = whitehurstComplete + whitehurstProgress + whitehurstOutOfProgramTransfer + whitehurstInProgramTransfer; 
            if("LA VEGA EL" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = laVegaElAssessed;
                Dash.Range('F'+str(DashTicker)).Value = laVegaElComplete; 
                Dash.Range('H'+str(DashTicker)).Value = laVegaElEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = laVegaElInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = laVegaElInactive; 
                Dash.Range('O'+str(DashTicker)).Value = laVegaElOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = laVegaElProgress; 
                Dash.Range('S'+str(DashTicker)).Value = laVegaElRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = laVegaElComplete + laVegaElProgress + laVegaElOutOfProgramTransfer + laVegaElInProgramTransfer; 
            if("LA VEGA H S" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = laVegaHSAssessed;
                Dash.Range('F'+str(DashTicker)).Value = laVegaHSComplete; 
                Dash.Range('H'+str(DashTicker)).Value = laVegaHSEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = laVegaHSInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = laVegaHSInactive; 
                Dash.Range('O'+str(DashTicker)).Value = laVegaHSOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = laVegaHSProgress; 
                Dash.Range('S'+str(DashTicker)).Value = laVegaHSRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = laVegaHSComplete + laVegaHSProgress + laVegaHSOutOfProgramTransfer + laVegaHSInProgramTransfer; 
            if("LA VEGA INT" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = laVegaIntAssessed;
                Dash.Range('F'+str(DashTicker)).Value = laVegaIntComplete; 
                Dash.Range('H'+str(DashTicker)).Value = laVegaIntEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = laVegaIntInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = laVegaIntInactive; 
                Dash.Range('O'+str(DashTicker)).Value = laVegaIntOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = laVegaIntProgress; 
                Dash.Range('S'+str(DashTicker)).Value = laVegaIntRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = laVegaIntComplete + laVegaIntProgress + laVegaIntOutOfProgramTransfer + laVegaIntInProgramTransfer; 
            if("LA VEGA J H" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = laVegaJHAssessed;
                Dash.Range('F'+str(DashTicker)).Value = laVegaJHComplete; 
                Dash.Range('H'+str(DashTicker)).Value = laVegaJHEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = laVegaJHInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = laVegaJHInactive; 
                Dash.Range('O'+str(DashTicker)).Value = laVegaJHOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = laVegaJHProgress; 
                Dash.Range('S'+str(DashTicker)).Value = laVegaJHRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = laVegaJHComplete + laVegaJHProgress + laVegaJHOutOfProgramTransfer + laVegaJHInProgramTransfer; 
            if("LA VEGA PR" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = laVegaPriAssessed;
                Dash.Range('F'+str(DashTicker)).Value = laVegaPriComplete; 
                Dash.Range('H'+str(DashTicker)).Value = laVegaPriEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = laVegaPriInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = laVegaPriInactive; 
                Dash.Range('O'+str(DashTicker)).Value = laVegaPriOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = laVegaPriProgress; 
                Dash.Range('S'+str(DashTicker)).Value = laVegaPriRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = laVegaPriComplete + laVegaPriProgress + laVegaPriOutOfProgramTransfer + laVegaPriInProgramTransfer; 
            if("MEXIA H S" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = mexiaHSAssessed;
                Dash.Range('F'+str(DashTicker)).Value = mexiaHSComplete; 
                Dash.Range('H'+str(DashTicker)).Value = mexiaHSEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = mexiaHSInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = mexiaHSInactive; 
                Dash.Range('O'+str(DashTicker)).Value = mexiaHSOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = mexiaHSProgress; 
                Dash.Range('S'+str(DashTicker)).Value = mexiaHSRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = mexiaHSComplete + mexiaHSProgress + mexiaHSOutOfProgramTransfer + mexiaHSInProgramTransfer; 
            if("MEXIA J H" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = mexiaJHAssessed;
                Dash.Range('F'+str(DashTicker)).Value = mexiaJHComplete; 
                Dash.Range('H'+str(DashTicker)).Value = mexiaJHEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = mexiaJHInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = mexiaJHInactive; 
                Dash.Range('O'+str(DashTicker)).Value = mexiaJHOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = mexiaJHProgress; 
                Dash.Range('S'+str(DashTicker)).Value = mexiaJHRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = mexiaJHComplete + mexiaJHProgress + mexiaJHOutOfProgramTransfer + mexiaJHInProgramTransfer; 
            if("CASTLEMAN" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = castelemanCreekAssessed;
                Dash.Range('F'+str(DashTicker)).Value = castelemanCreekComplete; 
                Dash.Range('H'+str(DashTicker)).Value = castelemanCreekEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = castelemanCreekInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = castelemanCreekInactive; 
                Dash.Range('O'+str(DashTicker)).Value = castelemanCreekOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = castelemanCreekProgress; 
                Dash.Range('S'+str(DashTicker)).Value = castelemanCreekRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = castelemanCreekComplete + castelemanCreekProgress + castelemanCreekOutOfProgramTransfer + castelemanCreekInProgramTransfer; 
            if("HEWITT" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = hewittAssessed;
                Dash.Range('F'+str(DashTicker)).Value = hewittComplete; 
                Dash.Range('H'+str(DashTicker)).Value = hewittEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = hewittInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = hewittInactive; 
                Dash.Range('O'+str(DashTicker)).Value = hewittOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = hewittProgress; 
                Dash.Range('S'+str(DashTicker)).Value = hewittRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = hewittComplete + hewittProgress + hewittOutOfProgramTransfer + hewittInProgramTransfer; 
            if("MIDWAY H S" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = midwayHSAssessed;
                Dash.Range('F'+str(DashTicker)).Value = midwayHSComplete; 
                Dash.Range('H'+str(DashTicker)).Value = midwayHSEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = midwayHSInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = midwayHSInactive; 
                Dash.Range('O'+str(DashTicker)).Value = midwayHSOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = midwayHSProgress; 
                Dash.Range('S'+str(DashTicker)).Value = midwayHSRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = midwayHSComplete + midwayHSProgress + midwayHSOutOfProgramTransfer + midwayHSInProgramTransfer; 
            if("MIDWAY MIDDLE" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = midwayMiddleAssessed;
                Dash.Range('F'+str(DashTicker)).Value = midwayMiddleComplete; 
                Dash.Range('H'+str(DashTicker)).Value = midwayMiddleEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = midwayMiddleInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = midwayMiddleInactive; 
                Dash.Range('O'+str(DashTicker)).Value = midwayMiddleOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = midwayMiddleProgress; 
                Dash.Range('S'+str(DashTicker)).Value = midwayMiddleRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = midwayMiddleComplete + midwayMiddleProgress + midwayMiddleOutOfProgramTransfer + midwayMiddleInProgramTransfer; 
            if("QUINN" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = quinnAssessed;
                Dash.Range('F'+str(DashTicker)).Value = quinnComplete; 
                Dash.Range('H'+str(DashTicker)).Value = quinnEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = quinnInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = quinnInactive; 
                Dash.Range('O'+str(DashTicker)).Value = quinnOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = quinnProgress; 
                Dash.Range('S'+str(DashTicker)).Value = quinnRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = quinnComplete + quinnProgress + quinnOutOfProgramTransfer + quinnInProgramTransfer; 
            if("TEAGUE" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = teagueAssessed;
                Dash.Range('F'+str(DashTicker)).Value = teagueComplete; 
                Dash.Range('H'+str(DashTicker)).Value = teagueEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = teagueInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = teagueInactive; 
                Dash.Range('O'+str(DashTicker)).Value = teagueOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = teagueProgress; 
                Dash.Range('S'+str(DashTicker)).Value = teagueRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = teagueComplete + teagueProgress + teagueOutOfProgramTransfer + teagueInProgramTransfer; 
            if("WACO CHARTER" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = wacoCharterAssessed;
                Dash.Range('F'+str(DashTicker)).Value = wacoCharterComplete; 
                Dash.Range('H'+str(DashTicker)).Value = wacoCharterEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = wacoCharterInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = wacoCharterInactive; 
                Dash.Range('O'+str(DashTicker)).Value = wacoCharterOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = wacoCharterProgress; 
                Dash.Range('S'+str(DashTicker)).Value = wacoCharterRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = wacoCharterComplete + wacoCharterProgress + wacoCharterOutOfProgramTransfer + wacoCharterInProgramTransfer; 
            if("ALTA VISTA" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = altaVistaAssessed;
                Dash.Range('F'+str(DashTicker)).Value = altaVistaComplete; 
                Dash.Range('H'+str(DashTicker)).Value = altaVistaEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = altaVistaInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = altaVistaInactive; 
                Dash.Range('O'+str(DashTicker)).Value = altaVistaOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = altaVistaProgress; 
                Dash.Range('S'+str(DashTicker)).Value = altaVistaRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = altaVistaComplete + altaVistaProgress + altaVistaOutOfProgramTransfer + altaVistaInProgramTransfer; 
            if("BRAZOS" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = brazosHSAssessed;
                Dash.Range('F'+str(DashTicker)).Value = brazosHSComplete; 
                Dash.Range('H'+str(DashTicker)).Value = brazosHSEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = brazosHSInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = brazosHSInactive; 
                Dash.Range('O'+str(DashTicker)).Value = brazosHSOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = brazosHSProgress; 
                Dash.Range('S'+str(DashTicker)).Value = brazosHSRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = brazosHSComplete + brazosHSProgress + brazosHSOutOfProgramTransfer + brazosHSInProgramTransfer; 
            if("BROOK AVENUE" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = brookAvenueAssessed;
                Dash.Range('F'+str(DashTicker)).Value = brookAvenueComplete; 
                Dash.Range('H'+str(DashTicker)).Value = brookAvenueEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = brookAvenueInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = brookAvenueInactive; 
                Dash.Range('O'+str(DashTicker)).Value = brookAvenueOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = brookAvenueProgress; 
                Dash.Range('S'+str(DashTicker)).Value = brookAvenueRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = brookAvenueComplete + brookAvenueProgress + brookAvenueOutOfProgramTransfer + brookAvenueInProgramTransfer; 
            if("CEASER CHAVEZ" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = cesarChavezAssessed;
                Dash.Range('F'+str(DashTicker)).Value = cesarChavezComplete; 
                Dash.Range('H'+str(DashTicker)).Value = cesarChavezEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = cesarChavezInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = cesarChavezInactive; 
                Dash.Range('O'+str(DashTicker)).Value = cesarChavezOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = cesarChavezProgress; 
                Dash.Range('S'+str(DashTicker)).Value = cesarChavezRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = cesarChavezComplete + cesarChavezProgress + cesarChavezOutOfProgramTransfer + cesarChavezInProgramTransfer; 
            if("CARVER" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = carverAssessed;
                Dash.Range('F'+str(DashTicker)).Value = carverComplete; 
                Dash.Range('H'+str(DashTicker)).Value = carverEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = carverInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = carverInactive; 
                Dash.Range('O'+str(DashTicker)).Value = carverOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = carverProgress; 
                Dash.Range('S'+str(DashTicker)).Value = carverRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = carverComplete + carverProgress + carverOutOfProgramTransfer + carverInProgramTransfer; 
            if("J H HINES" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = hinesAssessed;
                Dash.Range('F'+str(DashTicker)).Value = hinesComplete; 
                Dash.Range('H'+str(DashTicker)).Value = hinesEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = hinesInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = hinesInactive; 
                Dash.Range('O'+str(DashTicker)).Value = hinesOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = hinesProgress; 
                Dash.Range('S'+str(DashTicker)).Value = hinesRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = hinesComplete + hinesProgress + hinesOutOfProgramTransfer + hinesInProgramTransfer; 
            if("KENDRICK" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = kendrickAssessed;
                Dash.Range('F'+str(DashTicker)).Value = kendrickComplete; 
                Dash.Range('H'+str(DashTicker)).Value = kendrickEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = kendrickInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = kendrickInactive; 
                Dash.Range('O'+str(DashTicker)).Value = kendrickOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = kendrickProgress; 
                Dash.Range('S'+str(DashTicker)).Value = kendrickRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = kendrickComplete + kendrickProgress + kendrickOutOfProgramTransfer + kendrickInProgramTransfer; 
            if("PROVIDENT" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = providentHeightsAssessed;
                Dash.Range('F'+str(DashTicker)).Value = providentHeightsComplete; 
                Dash.Range('H'+str(DashTicker)).Value = providentHeightsEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = providentHeightsInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = providentHeightsInactive; 
                Dash.Range('O'+str(DashTicker)).Value = providentHeightsOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = providentHeightsProgress; 
                Dash.Range('S'+str(DashTicker)).Value = providentHeightsRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = providentHeightsComplete + providentHeightsProgress + providentHeightsOutOfProgramTransfer + providentHeightsInProgramTransfer; 
            if("TENNYSON" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = tennysonAssessed;
                Dash.Range('F'+str(DashTicker)).Value = tennysonComplete; 
                Dash.Range('H'+str(DashTicker)).Value = tennysonEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = tennysonInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = tennysonInactive; 
                Dash.Range('O'+str(DashTicker)).Value = tennysonOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = tennysonProgress; 
                Dash.Range('S'+str(DashTicker)).Value = tennysonRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = tennysonComplete + tennysonProgress + tennysonOutOfProgramTransfer + tennysonInProgramTransfer; 
            if("UNIVERSITY" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = universityAssessed;
                Dash.Range('F'+str(DashTicker)).Value = universityComplete; 
                Dash.Range('H'+str(DashTicker)).Value = universityEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = universityInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = universityInactive; 
                Dash.Range('O'+str(DashTicker)).Value = universityOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = universityProgress; 
                Dash.Range('S'+str(DashTicker)).Value = universityRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = universityComplete + universityProgress + universityOutOfProgramTransfer + universityInProgramTransfer; 
            if("WACO H S" in str(Dash.Range('B'+str(DashTicker)).Value)):
                Dash.Range('D'+str(DashTicker)).Value = wacoHSAssessed;
                Dash.Range('F'+str(DashTicker)).Value = wacoHSComplete; 
                Dash.Range('H'+str(DashTicker)).Value = wacoHSEnrolled; 
                Dash.Range('J'+str(DashTicker)).Value = wacoHSInProgramTransfer; 
                Dash.Range('M'+str(DashTicker)).Value = wacoHSInactive; 
                Dash.Range('O'+str(DashTicker)).Value = wacoHSOutOfProgramTransfer; 
                Dash.Range('Q'+str(DashTicker)).Value = wacoHSProgress; 
                Dash.Range('S'+str(DashTicker)).Value = wacoHSRegistered;
                Dash.Range('Z'+str(DashTicker)).Value = wacoHSComplete + wacoHSProgress + wacoHSOutOfProgramTransfer + wacoHSInProgramTransfer;
            DashTicker += 2



        target2.Close(SaveChanges=True)
        target.Close(SaveChanges=True)
        xlApp.Quit()


        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        target2 = xlApp.Workbooks.Open(pathOne)
        Mix = target.Worksheets(locationTwo)
        Assess = target.Worksheets(locationOne)
        Campus = target.Worksheets(locationFour)
        Dash = target.Worksheets("Dashboard")



        Dash.Range('D233').Value = str(AMSecondary)
        Dash.Range('H233').Value = str(AMPreparation)
        Dash.Range('M233').Value = str(AMExploration)
        Dash.Range('Q233').Value = str(AMEmployment)
        Dash.Range('U233').Value = str(AMFAFSA)

        Dash.Range('D235').Value = str(AFSecondary)
        Dash.Range('H235').Value = str(AFPreparation)
        Dash.Range('M235').Value = str(AFExploration)
        Dash.Range('Q235').Value = str(AFEmployment)
        Dash.Range('U235').Value = str(AFFAFSA)

        Dash.Range('D237').Value = str(BMSecondary)
        Dash.Range('H237').Value = str(BMPreparation)
        Dash.Range('M237').Value = str(BMExploration)
        Dash.Range('Q237').Value = str(BMEmployment)
        Dash.Range('U237').Value = str(BMFAFSA)

        Dash.Range('D239').Value = str(BFSecondary)
        Dash.Range('H239').Value = str(BFPreparation)
        Dash.Range('M239').Value = str(BFExploration)
        Dash.Range('Q239').Value = str(BFEmployment)
        Dash.Range('U239').Value = str(BFFAFSA)

        Dash.Range('D241').Value = str(HMSecondary)
        Dash.Range('H241').Value = str(HMPreparation)
        Dash.Range('M241').Value = str(HMExploration)
        Dash.Range('Q241').Value = str(HMEmployment)
        Dash.Range('U241').Value = str(HMFAFSA)

        Dash.Range('D243').Value = str(HFSecondary)
        Dash.Range('H243').Value = str(HFPreparation)
        Dash.Range('M243').Value = str(HFExploration)
        Dash.Range('Q243').Value = str(HFEmployment)
        Dash.Range('U243').Value = str(HFFAFSA)

        Dash.Range('D245').Value = str(IMSecondary)
        Dash.Range('H245').Value = str(IMPreparation)
        Dash.Range('M245').Value = str(IMExploration)
        Dash.Range('Q245').Value = str(IMEmployment)
        Dash.Range('U245').Value = str(IMFAFSA)

        Dash.Range('D247').Value = str(IFSecondary)
        Dash.Range('H247').Value = str(IFPreparation)
        Dash.Range('M247').Value = str(IFExploration)
        Dash.Range('Q247').Value = str(IFEmployment)
        Dash.Range('U247').Value = str(IFFAFSA)

        Dash.Range('D249').Value = str(WMSecondary)
        Dash.Range('H249').Value = str(WMPreparation)
        Dash.Range('M249').Value = str(WMExploration)
        Dash.Range('Q249').Value = str(WMEmployment)
        Dash.Range('U249').Value = str(WMFAFSA)

        Dash.Range('D251').Value = str(WFSecondary)
        Dash.Range('H251').Value = str(WFPreparation)
        Dash.Range('M251').Value = str(WFExploration)
        Dash.Range('Q251').Value = str(WFEmployment)
        Dash.Range('U251').Value = str(WFFAFSA)

        Dash.Range('D253').Value = str(PMSecondary)
        Dash.Range('H253').Value = str(PMPreparation)
        Dash.Range('M253').Value = str(PMExploration)
        Dash.Range('Q253').Value = str(PMEmployment)
        Dash.Range('U253').Value = str(PMFAFSA)

        Dash.Range('D255').Value = str(PFSecondary)
        Dash.Range('H255').Value = str(PFPreparation)
        Dash.Range('M255').Value = str(PFExploration)
        Dash.Range('Q255').Value = str(PFEmployment)
        Dash.Range('U255').Value = str(PFFAFSA)



        Dash.Range('F233').Value = str(AMSecondaryFull)
        Dash.Range('J233').Value = str(AMPreparationFull)
        Dash.Range('O233').Value = str(AMExplorationFull)
        Dash.Range('S233').Value = str(AMEmploymentFull)
        Dash.Range('X233').Value = str(AMFAFSAFull)

        Dash.Range('F235').Value = str(AFSecondaryFull)
        Dash.Range('J235').Value = str(AFPreparationFull)
        Dash.Range('O235').Value = str(AFExplorationFull)
        Dash.Range('S235').Value = str(AFEmploymentFull)
        Dash.Range('X235').Value = str(AFFAFSAFull)

        Dash.Range('F237').Value = str(BMSecondaryFull)
        Dash.Range('J237').Value = str(BMPreparationFull)
        Dash.Range('O237').Value = str(BMExplorationFull)
        Dash.Range('S237').Value = str(BMEmploymentFull)
        Dash.Range('X237').Value = str(BMFAFSAFull)

        Dash.Range('F239').Value = str(BFSecondaryFull)
        Dash.Range('J239').Value = str(BFPreparationFull)
        Dash.Range('O239').Value = str(BFExplorationFull)
        Dash.Range('S239').Value = str(BFEmploymentFull)
        Dash.Range('X239').Value = str(BFFAFSAFull)

        Dash.Range('F241').Value = str(HMSecondaryFull)
        Dash.Range('J241').Value = str(HMPreparationFull)
        Dash.Range('O241').Value = str(HMExplorationFull)
        Dash.Range('S241').Value = str(HMEmploymentFull)
        Dash.Range('X241').Value = str(HMFAFSAFull)

        Dash.Range('F243').Value = str(HFSecondaryFull)
        Dash.Range('J243').Value = str(HFPreparationFull)
        Dash.Range('O243').Value = str(HFExplorationFull)
        Dash.Range('S243').Value = str(HFEmploymentFull)
        Dash.Range('X243').Value = str(HFFAFSAFull)

        Dash.Range('F245').Value = str(IMSecondaryFull)
        Dash.Range('J245').Value = str(IMPreparationFull)
        Dash.Range('O245').Value = str(IMExplorationFull)
        Dash.Range('S245').Value = str(IMEmploymentFull)
        Dash.Range('X245').Value = str(IMFAFSAFull)

        Dash.Range('F247').Value = str(IFSecondaryFull)
        Dash.Range('J247').Value = str(IFPreparationFull)
        Dash.Range('O247').Value = str(IFExplorationFull)
        Dash.Range('S247').Value = str(IFEmploymentFull)
        Dash.Range('X247').Value = str(IFFAFSAFull)

        Dash.Range('F249').Value = str(WMSecondaryFull)
        Dash.Range('J249').Value = str(WMPreparationFull)
        Dash.Range('O249').Value = str(WMExplorationFull)
        Dash.Range('S249').Value = str(WMEmploymentFull)
        Dash.Range('X249').Value = str(WMFAFSAFull)

        Dash.Range('F251').Value = str(WFSecondaryFull)
        Dash.Range('J251').Value = str(WFPreparationFull)
        Dash.Range('O251').Value = str(WFExplorationFull)
        Dash.Range('S251').Value = str(WFEmploymentFull)
        Dash.Range('X251').Value = str(WFFAFSAFull)

        Dash.Range('F253').Value = str(PMSecondaryFull)
        Dash.Range('J253').Value = str(PMPreparationFull)
        Dash.Range('O253').Value = str(PMExplorationFull)
        Dash.Range('S253').Value = str(PMEmploymentFull)
        Dash.Range('X253').Value = str(PMFAFSAFull)

        Dash.Range('F255').Value = str(PFSecondaryFull)
        Dash.Range('J255').Value = str(PFPreparationFull)
        Dash.Range('O255').Value = str(PFExplorationFull)
        Dash.Range('S255').Value = str(PFEmploymentFull)
        Dash.Range('X255').Value = str(PFFAFSAFull)

        #Service Data

        DashTicker = 262
        Dash.Range('D'+str(DashTicker)).Value = str(asianMaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(asianFemaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(blackMaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(blackFemaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(hispanicMaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(hispanicFemaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(nativeMaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(nativeFemaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(whiteMaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(whiteFemaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(pacificMaleMonthly); DashTicker += 2
        Dash.Range('D'+str(DashTicker)).Value = str(pacificFemaleMonthly); DashTicker += 2

        DashTicker = 262
        Dash.Range('H'+str(DashTicker)).Value = str(asianMaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(asianFemaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(blackMaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(blackFemaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(hispanicMaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(hispanicFemaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(nativeMaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(nativeFemaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(whiteMaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(whiteFemaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(pacificMaleService); DashTicker += 2
        Dash.Range('H'+str(DashTicker)).Value = str(pacificFemaleService); DashTicker += 2

        DashTicker = 262
        Dash.Range('AB'+str(DashTicker)).Value = str(asianMaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(asianFemaleService); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(blackMaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(blackFemaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(hispanicMaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(hispanicFemaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(nativeMaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(nativeFemaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(whiteMaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(whiteFemaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(pacificMaleServiced); DashTicker += 2
        Dash.Range('AB'+str(DashTicker)).Value = str(pacificFemaleServiced); DashTicker += 2

        DashTicker = 262
        Dash.Range('U'+str(DashTicker)).Value = str(asianMaleMonthly + asianMaleService + asianMaleIndirect + asianMaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(asianFemaleMonthly + asianFemaleService + asianFemaleIndirect + asianFemaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(blackMaleMonthly + blackMaleService + blackMaleIndirect + blackMaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(blackFemaleMonthly + blackFemaleService + blackFemaleIndirect + blackFemaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(hispanicMaleMonthly + hispanicMaleService + hispanicMaleIndirect + hispanicMaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(hispanicFemaleMonthly + hispanicFemaleService + hispanicFemaleIndirect + hispanicFemaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(nativeMaleMonthly + nativeMaleService + nativeMaleIndirect + nativeMaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(nativeFemaleMonthly + nativeFemaleService + nativeFemaleIndirect + nativeFemaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(whiteMaleMonthly + whiteMaleService + whiteMaleIndirect + whiteMaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(whiteFemaleMonthly + whiteFemaleService + whiteFemaleIndirect + whiteFemaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(pacificMaleMonthly + pacificMaleService + pacificMaleIndirect + pacificMaleOther); DashTicker += 2
        Dash.Range('U'+str(DashTicker)).Value = str(pacificFemaleMonthly + pacificFemaleService + pacificFemaleIndirect + pacificFemaleOther); DashTicker += 2



        DashTicker = 262
        Dash.Range('M'+str(DashTicker)).Value = str(asianMaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(asianFemaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(blackMaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(blackFemaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(hispanicMaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(hispanicFemaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(nativeMaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(nativeFemaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(whiteMaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(whiteFemaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(pacificMaleBehavioral); DashTicker += 2
        Dash.Range('M'+str(DashTicker)).Value = str(pacificFemaleBehavioral); DashTicker += 2
        DashTicker = 262
        Dash.Range('O'+str(DashTicker)).Value = str(asianMaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(asianFemaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(blackMaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(blackFemaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(hispanicMaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(hispanicFemaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(nativeMaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(nativeFemaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(whiteMaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(whiteFemaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(pacificMaleAcademics); DashTicker += 2
        Dash.Range('O'+str(DashTicker)).Value = str(pacificFemaleAcademics); DashTicker += 2
        DashTicker = 262
        Dash.Range('Q'+str(DashTicker)).Value = str(asianMaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(asianFemaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(blackMaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(blackFemaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(hispanicMaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(hispanicFemaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(nativeMaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(nativeFemaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(whiteMaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(whiteFemaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(pacificMaleSocial); DashTicker += 2
        Dash.Range('Q'+str(DashTicker)).Value = str(pacificFemaleSocial); DashTicker += 2
        DashTicker = 262
        Dash.Range('S'+str(DashTicker)).Value = str(asianMaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(asianFemaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(blackMaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(blackFemaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(hispanicMaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(hispanicFemaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(nativeMaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(nativeFemaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(whiteMaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(whiteFemaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(pacificMaleAttendance); DashTicker += 2
        Dash.Range('S'+str(DashTicker)).Value = str(pacificFemaleAttendance); DashTicker += 2


        target.Close(SaveChanges=True)
        target2.Close(SaveChanges=True)

        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        target2 = xlApp.Workbooks.Open(pathOne)
        Mix = target.Worksheets(locationTwo)
        Assess = target.Worksheets(locationOne)
        Campus = target.Worksheets(locationFour)
        Progress = target.Worksheets(locationFive)
        Dash = target.Worksheets("Dashboard")

        #Improved

        WMAcademicI = 0
        WMAttendanceI = 0
        WMBehaviorI = 0
        WFAcademicI = 0
        WFAttendanceI = 0
        WFBehaviorI = 0

        AMAcademicI = 0
        AMAttendanceI = 0
        AMBehaviorI = 0
        AFAcademicI = 0
        AFAttendanceI = 0
        AFBehaviorI = 0

        BMAcademicI = 0
        BMAttendanceI = 0
        BMBehaviorI = 0
        BFAcademicI = 0
        BFAttendanceI = 0
        BFBehaviorI = 0

        HMAcademicI = 0
        HMAttendanceI = 0
        HMBehaviorI = 0
        HFAcademicI = 0
        HFAttendanceI = 0
        HFBehaviorI = 0

        IMAcademicI = 0
        IMAttendanceI = 0
        IMBehaviorI = 0
        IFAcademicI = 0
        IFAttendanceI = 0
        IFBehaviorI = 0

        PMAcademicI = 0
        PMAttendanceI = 0
        PMBehaviorI = 0
        PFAcademicI = 0
        PFAttendanceI = 0
        PFBehaviorI = 0

        #Complete Assessed with Progress

        CWMAcademicI = 0
        CWMAttendanceI = 0
        CWMBehaviorI = 0
        CWFAcademicI = 0
        CWFAttendanceI = 0
        CWFBehaviorI = 0

        CAMAcademicI = 0
        CAMAttendanceI = 0
        CAMBehaviorI = 0
        CAFAcademicI = 0
        CAFAttendanceI = 0
        CAFBehaviorI = 0

        CBMAcademicI = 0
        CBMAttendanceI = 0
        CBMBehaviorI = 0
        CBFAcademicI = 0
        CBFAttendanceI = 0
        CBFBehaviorI = 0

        CHMAcademicI = 0
        CHMAttendanceI = 0
        CHMBehaviorI = 0
        CHFAcademicI = 0
        CHFAttendanceI = 0
        CHFBehaviorI = 0

        CIMAcademicI = 0
        CIMAttendanceI = 0
        CIMBehaviorI = 0
        CIFAcademicI = 0
        CIFAttendanceI = 0
        CIFBehaviorI = 0

        CPMAcademicI = 0
        CPMAttendanceI = 0
        CPMBehaviorI = 0
        CPFAcademicI = 0
        CPFAttendanceI = 0
        CPFBehaviorI = 0

        #declined

        WMAcademicD = 0
        WMAttendanceD = 0
        WMBehaviorD = 0
        WFAcademicD = 0
        WFAttendanceD = 0
        WFBehaviorD = 0

        AMAcademicD = 0
        AMAttendanceD = 0
        AMBehaviorD = 0
        AFAcademicD = 0
        AFAttendanceD = 0
        AFBehaviorD = 0

        BMAcademicD = 0
        BMAttendanceD = 0
        BMBehaviorD = 0
        BFAcademicD = 0
        BFAttendanceD = 0
        BFBehaviorD = 0

        HMAcademicD = 0
        HMAttendanceD = 0
        HMBehaviorD = 0
        HFAcademicD = 0
        HFAttendanceD = 0
        HFBehaviorD = 0

        IMAcademicD = 0
        IMAttendanceD = 0
        IMBehaviorD = 0
        IFAcademicD = 0
        IFAttendanceD = 0
        IFBehaviorD = 0

        PMAcademicD = 0
        PMAttendanceD = 0
        PMBehaviorD = 0
        PFAcademicD = 0
        PFAttendanceD = 0
        PFBehaviorD = 0

        #Complete Assess with Decline Progress

        CWMAcademicD = 0
        CWMAttendanceD = 0
        CWMBehaviorD = 0
        CWFAcademicD = 0
        CWFAttendanceD = 0
        CWFBehaviorD = 0

        CAMAcademicD = 0
        CAMAttendanceD = 0
        CAMBehaviorD = 0
        CAFAcademicD = 0
        CAFAttendanceD = 0
        CAFBehaviorD = 0

        CBMAcademicD = 0
        CBMAttendanceD = 0
        CBMBehaviorD = 0
        CBFAcademicD = 0
        CBFAttendanceD = 0
        CBFBehaviorD = 0

        CHMAcademicD = 0
        CHMAttendanceD = 0
        CHMBehaviorD = 0
        CHFAcademicD = 0
        CHFAttendanceD = 0
        CHFBehaviorD = 0

        CIMAcademicD = 0
        CIMAttendanceD = 0
        CIMBehaviorD = 0
        CIFAcademicD = 0
        CIFAttendanceD = 0
        CIFBehaviorD = 0

        CPMAcademicD = 0
        CPMAttendanceD = 0
        CPMBehaviorD = 0
        CPFAcademicD = 0
        CPFAttendanceD = 0
        CPFBehaviorD = 0


        AcademicImprove = 0
        AcademicDecline = 0

        AttendanceImprove = 0
        AttendanceDecline = 0

        BehaviorImprove = 0
        BehaviorDecline = 0

        rowShifter = 2
        whileLoopCondition = False

        while(whileLoopCondition != True):
            if(str(Progress.Range('AD'+str(rowShifter)).Value) != "None"):
                if(str(Progress.Range('U'+str(rowShifter)).Value) in "M" and str(Progress.Range('X'+str(rowShifter)).Value) in "W "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        WMAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CWMAcademicI += 1
                    else:
                        AcademicDecline += 1
                        WMAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CWMAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) in "F" and str(Progress.Range('X'+str(rowShifter)).Value) in "W "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        WFAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CWFAcademicI += 1
                    else:
                        AcademicDecline += 1
                        WFAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CWFAcademicD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "A "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        AMAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CAMAcademicI += 1
                    else:
                        AcademicDecline += 1
                        AMAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CAMAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "A "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        AFAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CAFAcademicI += 1
                    else:
                        AcademicDecline += 1
                        AFAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CAFAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "B "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        BMAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CBMAcademicI += 1
                    else:
                        AcademicDecline += 1
                        BMAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CBMAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "B "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        BFAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CBFAcademicI += 1
                    else:
                        AcademicDecline += 1
                        BFAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CBFAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "H "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        HMAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CHMAcademicI += 1
                    else:
                        AcademicDecline += 1
                        HMAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CHMAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "H "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        HFAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CHFAcademicI += 1
                    else:
                        AcademicDecline += 1
                        HFAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CHFAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "I "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        IMAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CIMAcademicI += 1
                    else:
                        AcademicDecline += 1
                        IMAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CIMAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "I "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        IFAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CIFAcademicI += 1
                    else:
                        AcademicDecline += 1
                        IFAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CIFAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "P "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        PMAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CPMAcademicI += 1
                    else:
                        AcademicDecline += 1
                        PMAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CPMAcademicD += 1
                            
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "P "):
                    if((Progress.Range('AA'+str(rowShifter)).Value) == 1 and str(Progress.Range('AD'+str(rowShifter)).Value) != ""):
                        AcademicImprove += 1
                        PFAcademicI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CPFAcademicI += 1
                    else:
                        AcademicDecline += 1
                        PFAcademicD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CPFAcademicD += 1


            if(str(Progress.Range('AD'+str(rowShifter)).Value) != "None"):
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "W "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        WMAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CWMAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        WMAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CWMAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "W "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        WFAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CWFAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        WFAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CWFAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "A "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        AMAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CAMAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        AMAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CAMAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "A "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        AFAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CAFAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        AFAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CAFAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "B "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        BMAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CBMAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        BMAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CBMAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "B "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        BFAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CBFAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        BFAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CBFAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "H "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        HMAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CHMAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        HMAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CHMAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "H "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        HFAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CHFAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        HFAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CHFAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "I "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        IMAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CIMAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        IMAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CIMAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "I "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        IFAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CIFAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        IFAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CIFAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "P "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        PMAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CPMAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        PMAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CPMAttendanceD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "P "):
                    if((Progress.Range('AB'+str(rowShifter)).Value) == 1 and str(Progress.Range('AE'+str(rowShifter)).Value) != ""):
                        AttendanceImprove += 1
                        PFAttendanceI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CPFAttendanceI += 1
                    else:
                        AttendanceDecline += 1
                        PFAttendanceD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CPFAttendanceD += 1


            if(str(Progress.Range('AD'+str(rowShifter)).Value) != "None"):
                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "W "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        WMBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CWMBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        WMBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CWMBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "W "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        WFBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CWFBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        WFBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CWFBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "A "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        AMBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CAMBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        AMBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CAMBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "A "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        AFBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CAFBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        AFBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CAFBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "B "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        BMBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CBMBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        BMBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CBMBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "B "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        BFBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CBFBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        BFBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CBFBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "H "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        HMBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CHMBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        HMBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CHMBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "H "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        HFBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CHFBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        HFBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CHFBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "I "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        IMBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CIMBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        IMBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CIMBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "I "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        IFBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CIFBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        IFBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CIFBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "M" and str(Progress.Range('X'+str(rowShifter)).Value) == "P "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        PMBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CPMBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        PMBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CPMBehaviorD += 1

                if(str(Progress.Range('U'+str(rowShifter)).Value) == "F" and str(Progress.Range('X'+str(rowShifter)).Value) == "P "):
                    if((Progress.Range('AC'+str(rowShifter)).Value) == 1 and str(Progress.Range('AF'+str(rowShifter)).Value) != ""):
                        BehaviorImprove += 1
                        PFBehaviorI += 1
                    if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                        CPFBehaviorI += 1
                    else:
                        BehaviorDecline += 1
                        PFBehaviorD += 1
                        if(str(Progress.Range('S'+str(rowShifter)).Value) in "Complete"):
                            CPFBehaviorD += 1

            rowShifter += 1
            if(Progress.Range('A'+str(rowShifter)).Value == None):
                whileLoopCondition = True

        Dash.Range('D291').Value = AMAcademicI
        Dash.Range('F291').Value = AMAttendanceI
        Dash.Range('H291').Value = AMBehaviorI
        Dash.Range('J291').Value = AMAcademicI + AMAttendanceI + AMBehaviorI

        Dash.Range('D293').Value = AFAcademicI
        Dash.Range('F293').Value = AFAttendanceI
        Dash.Range('H293').Value = AFBehaviorI
        Dash.Range('J293').Value = AFAcademicI + AFAttendanceI + AFBehaviorI

        Dash.Range('D295').Value = BMAcademicI
        Dash.Range('F295').Value = BMAttendanceI
        Dash.Range('H295').Value = BMBehaviorI
        Dash.Range('J295').Value = BMAcademicI + BMAttendanceI + BMBehaviorI

        Dash.Range('D297').Value = BFAcademicI
        Dash.Range('F297').Value = BFAttendanceI
        Dash.Range('H297').Value = BFBehaviorI
        Dash.Range('J297').Value = BFAcademicI + BFAttendanceI + BFBehaviorI

        Dash.Range('D299').Value = HMAcademicI
        Dash.Range('F299').Value = HMAttendanceI
        Dash.Range('H299').Value = HMBehaviorI
        Dash.Range('J299').Value = HMAcademicI + HMAttendanceI + HMBehaviorI

        Dash.Range('D301').Value = HFAcademicI
        Dash.Range('F301').Value = HFAttendanceI
        Dash.Range('H301').Value = HFBehaviorI
        Dash.Range('J301').Value = HFAcademicI + HFAttendanceI + HFBehaviorI

        Dash.Range('D303').Value = IMAcademicI
        Dash.Range('F303').Value = IMAttendanceI
        Dash.Range('H303').Value = IMBehaviorI
        Dash.Range('J303').Value = IMAcademicI + IMAttendanceI + IMBehaviorI

        Dash.Range('D305').Value = IFAcademicI
        Dash.Range('F305').Value = IFAttendanceI
        Dash.Range('H305').Value = IFBehaviorI
        Dash.Range('J305').Value = IFAcademicI + IFAttendanceI + IFBehaviorI

        Dash.Range('D307').Value = WMAcademicI
        Dash.Range('F307').Value = WMAttendanceI
        Dash.Range('H307').Value = WMBehaviorI
        Dash.Range('J307').Value = WMAcademicI + WMAttendanceI + WMBehaviorI

        Dash.Range('D309').Value = WFAcademicI
        Dash.Range('F309').Value = WFAttendanceI
        Dash.Range('H309').Value = WFBehaviorI
        Dash.Range('J309').Value = WFAcademicI + WFAttendanceI + WFBehaviorI

        Dash.Range('D311').Value = PMAcademicI
        Dash.Range('F311').Value = PMAttendanceI
        Dash.Range('H311').Value = PMBehaviorI
        Dash.Range('J311').Value = PMAcademicI + PMAttendanceI + PMBehaviorI

        Dash.Range('D313').Value = PFAcademicI
        Dash.Range('F313').Value = PFAttendanceI
        Dash.Range('H313').Value = PFBehaviorI
        Dash.Range('J313').Value = PFAcademicI + PFAttendanceI + PFBehaviorI



        Dash.Range('D319').Value = AMAcademicI + AFAcademicI
        Dash.Range('F319').Value = AMAttendanceI + AFAttendanceI
        Dash.Range('H319').Value = AMBehaviorI + AFBehaviorI
        Dash.Range('J319').Value = AMAcademicI + AMAttendanceI + AMBehaviorI + AFAcademicI + AFAttendanceI + AFBehaviorI

        Dash.Range('D321').Value = BMAcademicI + BFAcademicI
        Dash.Range('F321').Value = BMAttendanceI + BFAttendanceI
        Dash.Range('H321').Value = BMBehaviorI + BFBehaviorI
        Dash.Range('J321').Value = BMAcademicI + BMAttendanceI + BMBehaviorI + BFAcademicI + BFAttendanceI + BFBehaviorI

        Dash.Range('D323').Value = HMAcademicI + HFAcademicI
        Dash.Range('F323').Value = HMAttendanceI + HFAttendanceI
        Dash.Range('H323').Value = HMBehaviorI + HFBehaviorI
        Dash.Range('J323').Value = HMAcademicI + HMAttendanceI + HMBehaviorI + HFAcademicI + HFAttendanceI + HFBehaviorI

        Dash.Range('D325').Value = IMAcademicI + IFAcademicI
        Dash.Range('F325').Value = IMAttendanceI + IFAttendanceI
        Dash.Range('H325').Value = IMBehaviorI + IFBehaviorI
        Dash.Range('J325').Value = IMAcademicI + IMAttendanceI + IMBehaviorI + IFAcademicI + IFAttendanceI + IFBehaviorI

        Dash.Range('D327').Value = WMAcademicI + WFAcademicI
        Dash.Range('F327').Value = WMAttendanceI + WFAttendanceI
        Dash.Range('H327').Value = WMBehaviorI + WFBehaviorI
        Dash.Range('J327').Value = WMAcademicI + WMAttendanceI + WMBehaviorI + WFAcademicI + WFAttendanceI + WFBehaviorI

        Dash.Range('D329').Value = PMAcademicI + PFAcademicI
        Dash.Range('F329').Value = PMAttendanceI + PFAttendanceI
        Dash.Range('H329').Value = PMBehaviorI + PFBehaviorI
        Dash.Range('J329').Value = PMAcademicI + PMAttendanceI + PMBehaviorI + PFAcademicI + PFAttendanceI + PFBehaviorI



        Dash.Range('D335').Value = AMAcademicI + BMAcademicI + HMAcademicI + IMAcademicI + WMAcademicI + PMAcademicI
        Dash.Range('F335').Value = AMAttendanceI + BMAttendanceI + HMAttendanceI + IMAttendanceI + WMAttendanceI + PMAttendanceI
        Dash.Range('H335').Value = AMBehaviorI + BMBehaviorI + HMBehaviorI + IMBehaviorI + WMBehaviorI + PMBehaviorI
        Dash.Range('J335').Value = AMAcademicI + AMAttendanceI + AMBehaviorI + BMAcademicI + BMAttendanceI + BMBehaviorI + HMAcademicI + HMAttendanceI + HMBehaviorI + IMAcademicI + IMAttendanceI + IMBehaviorI + WMAcademicI + WMAttendanceI + WMBehaviorI + PMAcademicI + PMAttendanceI + PMBehaviorI

        Dash.Range('D337').Value = AFAcademicI + BFAcademicI + HFAcademicI + IFAcademicI + WFAcademicI + PFAcademicI
        Dash.Range('F337').Value = AFAttendanceI + BFAttendanceI + HFAttendanceI + IFAttendanceI + WFAttendanceI + PFAttendanceI
        Dash.Range('H337').Value = AFBehaviorI + BFBehaviorI + HFBehaviorI + IFBehaviorI + WFBehaviorI + PFBehaviorI
        Dash.Range('J337').Value = AFAcademicI + AFAttendanceI + AFBehaviorI + BFAcademicI + BFAttendanceI + BFBehaviorI + HFAcademicI + HFAttendanceI + HFBehaviorI + IFAcademicI + IFAttendanceI + IFBehaviorI + WFAcademicI + WFAttendanceI + WFBehaviorI + PFAcademicI + PFAttendanceI + PFBehaviorI



        Dash.Range('M291').Value = AMAcademicD 
        Dash.Range('O291').Value = AMAttendanceD 
        Dash.Range('Q291').Value = AMBehaviorD 
        Dash.Range('S291').Value = AMAcademicD + AMAttendanceD + AMBehaviorD 

        Dash.Range('M293').Value = AFAcademicD 
        Dash.Range('O293').Value = AFAttendanceD 
        Dash.Range('Q293').Value = AFBehaviorD 
        Dash.Range('S293').Value = AFAcademicD + AFAttendanceD + AFBehaviorD 

        Dash.Range('M295').Value = BMAcademicD 
        Dash.Range('O295').Value = BMAttendanceD 
        Dash.Range('Q295').Value = BMBehaviorD 
        Dash.Range('S295').Value = BMAcademicD + BMAttendanceD + BMBehaviorD 

        Dash.Range('M297').Value = BFAcademicD 
        Dash.Range('O297').Value = BFAttendanceD 
        Dash.Range('Q297').Value = BFBehaviorD 
        Dash.Range('S297').Value = BFAcademicD + BFAttendanceD + BFBehaviorD 

        Dash.Range('M299').Value = HMAcademicD 
        Dash.Range('O299').Value = HMAttendanceD 
        Dash.Range('Q299').Value = HMBehaviorD 
        Dash.Range('S299').Value = HMAcademicD + HMAttendanceD + HMBehaviorD 

        Dash.Range('M301').Value = HFAcademicD 
        Dash.Range('O301').Value = HFAttendanceD 
        Dash.Range('Q301').Value = HFBehaviorD 
        Dash.Range('S301').Value = HFAcademicD + HFAttendanceD + HFBehaviorD 

        Dash.Range('M303').Value = IMAcademicD 
        Dash.Range('O303').Value = IMAttendanceD 
        Dash.Range('Q303').Value = IMBehaviorD 
        Dash.Range('S303').Value = IMAcademicD + IMAttendanceD + IMBehaviorD 

        Dash.Range('M305').Value = IFAcademicD 
        Dash.Range('O305').Value = IFAttendanceD 
        Dash.Range('Q305').Value = IFBehaviorD 
        Dash.Range('S305').Value = IFAcademicD + IFAttendanceD + IFBehaviorD 

        Dash.Range('M307').Value = WMAcademicD 
        Dash.Range('O307').Value = WMAttendanceD 
        Dash.Range('Q307').Value = WMBehaviorD 
        Dash.Range('S307').Value = WMAcademicD + WMAttendanceD + WMBehaviorD 

        Dash.Range('M309').Value = WFAcademicD 
        Dash.Range('O309').Value = WFAttendanceD 
        Dash.Range('Q309').Value = WFBehaviorD 
        Dash.Range('S309').Value = WFAcademicD + WFAttendanceD + WFBehaviorD 

        Dash.Range('M311').Value = PMAcademicD 
        Dash.Range('O311').Value = PMAttendanceD 
        Dash.Range('Q311').Value = PMBehaviorD 
        Dash.Range('S311').Value = PMAcademicD + PMAttendanceD + PMBehaviorD 

        Dash.Range('M313').Value = PFAcademicD 
        Dash.Range('O313').Value = PFAttendanceD 
        Dash.Range('Q313').Value = PFBehaviorD 
        Dash.Range('S313').Value = PFAcademicD + PFAttendanceD + PFBehaviorD 



        Dash.Range('M319').Value = AMAcademicD + AFAcademicD 
        Dash.Range('O319').Value = AMAttendanceD + AFAttendanceD 
        Dash.Range('Q319').Value = AMBehaviorD + AFBehaviorD 
        Dash.Range('S319').Value = AMAcademicD + AMAttendanceD + AMBehaviorD + AFAcademicD + AFAttendanceD + AFBehaviorD 

        Dash.Range('M321').Value = BMAcademicD + BFAcademicD 
        Dash.Range('O321').Value = BMAttendanceD + BFAttendanceD 
        Dash.Range('Q321').Value = BMBehaviorD + BFBehaviorD 
        Dash.Range('S321').Value = BMAcademicD + BMAttendanceD + BMBehaviorD + BFAcademicD + BFAttendanceD + BFBehaviorD 

        Dash.Range('M323').Value = HMAcademicD + HFAcademicD 
        Dash.Range('O323').Value = HMAttendanceD + HFAttendanceD 
        Dash.Range('Q323').Value = HMBehaviorD + HFBehaviorD 
        Dash.Range('S323').Value = HMAcademicD + HMAttendanceD + HMBehaviorD + HFAcademicD + HFAttendanceD + HFBehaviorI

        Dash.Range('M325').Value = IMAcademicD + IFAcademicD 
        Dash.Range('O325').Value = IMAttendanceD + IFAttendanceD 
        Dash.Range('Q325').Value = IMBehaviorD + IFBehaviorD 
        Dash.Range('S325').Value = IMAcademicD + IMAttendanceD + IMBehaviorD + IFAcademicD + IFAttendanceD + IFBehaviorI

        Dash.Range('M327').Value = WMAcademicD + WFAcademicD 
        Dash.Range('O327').Value = WMAttendanceD + WFAttendanceD 
        Dash.Range('Q327').Value = WMBehaviorD + WFBehaviorD 
        Dash.Range('S327').Value = WMAcademicD + WMAttendanceD + WMBehaviorD + WFAcademicD + WFAttendanceD + WFBehaviorD 

        Dash.Range('M329').Value = PMAcademicD + PFAcademicD 
        Dash.Range('O329').Value = PMAttendanceD + PFAttendanceD 
        Dash.Range('Q329').Value = PMBehaviorD + PFBehaviorD 
        Dash.Range('S329').Value = PMAcademicD + PMAttendanceD + PMBehaviorD + PFAcademicD + PFAttendanceD + PFBehaviorD 



        Dash.Range('M335').Value = AMAcademicD + BMAcademicD + HMAcademicD + IMAcademicD + WMAcademicD + PMAcademicD 
        Dash.Range('O335').Value = AMAttendanceD + BMAttendanceD + HMAttendanceD + IMAttendanceD + WMAttendanceD + PMAttendanceD 
        Dash.Range('Q335').Value = AMBehaviorD + BMBehaviorD + HMBehaviorD + IMBehaviorD + WMBehaviorD + PMBehaviorD 
        Dash.Range('S335').Value = AMAcademicD + AMAttendanceD + AMBehaviorD + BMAcademicD + BMAttendanceD + BMBehaviorD + HMAcademicD + HMAttendanceD + HMBehaviorD + IMAcademicD + IMAttendanceD + IMBehaviorD + WMAcademicD + WMAttendanceD + WMBehaviorD + PMAcademicD + PMAttendanceD + PMBehaviorD 

        Dash.Range('M337').Value = AFAcademicD + BFAcademicD + HFAcademicD + IFAcademicD + WFAcademicD + PFAcademicD 
        Dash.Range('O337').Value = AFAttendanceD + BFAttendanceD + HFAttendanceD + IFAttendanceD + WFAttendanceD + PFAttendanceD  
        Dash.Range('Q337').Value = AFBehaviorD + BFBehaviorD + HFBehaviorD + IFBehaviorD + WFBehaviorD + PFBehaviorD 
        Dash.Range('S337').Value = AFAcademicD + AFAttendanceD + AFBehaviorD + BFAcademicD + BFAttendanceD + BFBehaviorD + HFAcademicD + HFAttendanceD + HFBehaviorD + IFAcademicD + IFAttendanceD + IFBehaviorD + WFAcademicD + WFAttendanceD + WFBehaviorD + PFAcademicD + PFAttendanceD + PFBehaviorD 



        Dash.Range('X291').Value = CAMAcademicI
        Dash.Range('Z291').Value = CAMAttendanceI
        Dash.Range('AB291').Value = CAMBehaviorI
        Dash.Range('AD291').Value = CAMAcademicI + CAMAttendanceI + CAMBehaviorI

        Dash.Range('X293').Value = CAFAcademicI
        Dash.Range('Z293').Value = CAFAttendanceI
        Dash.Range('AB293').Value = CAFBehaviorI
        Dash.Range('AD293').Value = CAFAcademicI + CAFAttendanceI + CAFBehaviorI

        Dash.Range('X295').Value = CBMAcademicI
        Dash.Range('Z295').Value = CBMAttendanceI
        Dash.Range('AB295').Value = CBMBehaviorI
        Dash.Range('AD295').Value = CBMAcademicI + CBMAttendanceI + CBMBehaviorI

        Dash.Range('X297').Value = CBFAcademicI
        Dash.Range('Z297').Value = CBFAttendanceI
        Dash.Range('AB297').Value = CBFBehaviorI
        Dash.Range('AD297').Value = CBFAcademicI + CBFAttendanceI + CBFBehaviorI

        Dash.Range('X299').Value = CHMAcademicI
        Dash.Range('Z299').Value = CHMAttendanceI
        Dash.Range('AB299').Value = CHMBehaviorI
        Dash.Range('AD299').Value = CHMAcademicI + CHMAttendanceI + CHMBehaviorI

        Dash.Range('X301').Value = CHFAcademicI
        Dash.Range('Z301').Value = CHFAttendanceI
        Dash.Range('AB301').Value = CHFBehaviorI
        Dash.Range('AD301').Value = CHFAcademicI + CHFAttendanceI + CHFBehaviorI

        Dash.Range('X303').Value = CIMAcademicI
        Dash.Range('Z303').Value = CIMAttendanceI
        Dash.Range('AB303').Value = CIMBehaviorI
        Dash.Range('AD303').Value = CIMAcademicI + CIMAttendanceI + CIMBehaviorI

        Dash.Range('X305').Value = CIFAcademicI
        Dash.Range('Z305').Value = CIFAttendanceI
        Dash.Range('AB305').Value = CIFBehaviorI
        Dash.Range('AD305').Value = CIFAcademicI + CIFAttendanceI + CIFBehaviorI

        Dash.Range('X307').Value = CWMAcademicI
        Dash.Range('Z307').Value = CWMAttendanceI
        Dash.Range('AB307').Value = CWMBehaviorI
        Dash.Range('AD307').Value = CWMAcademicI + CWMAttendanceI + CWMBehaviorI

        Dash.Range('X309').Value = CWFAcademicI
        Dash.Range('Z309').Value = CWFAttendanceI
        Dash.Range('AB309').Value = CWFBehaviorI
        Dash.Range('AD309').Value = CWFAcademicI + CWFAttendanceI + CWFBehaviorI

        Dash.Range('X311').Value = CPMAcademicI
        Dash.Range('Z311').Value = CPMAttendanceI
        Dash.Range('AB311').Value = CPMBehaviorI
        Dash.Range('AD311').Value = CPMAcademicI + CPMAttendanceI + CPMBehaviorI

        Dash.Range('X313').Value = CPFAcademicI
        Dash.Range('Z313').Value = CPFAttendanceI
        Dash.Range('AB313').Value = CPFBehaviorI
        Dash.Range('AD313').Value = CPFAcademicI + CPFAttendanceI + CPFBehaviorI



        Dash.Range('X319').Value = CAMAcademicI + CAFAcademicI
        Dash.Range('Z319').Value = CAMAttendanceI + CAFAttendanceI
        Dash.Range('AB319').Value = CAMBehaviorI + CAFBehaviorI
        Dash.Range('AD319').Value = CAMAcademicI + CAMAttendanceI + CAMBehaviorI + CAFAcademicI + CAFAttendanceI + CAFBehaviorI

        Dash.Range('X321').Value = CBMAcademicI + CBFAcademicI
        Dash.Range('Z321').Value = CBMAttendanceI + CBFAttendanceI
        Dash.Range('AB321').Value = CBMBehaviorI + CBFBehaviorI
        Dash.Range('AD321').Value = CBMAcademicI + CBMAttendanceI + CBMBehaviorI + CBFAcademicI + CBFAttendanceI + CBFBehaviorI

        Dash.Range('X323').Value = CHMAcademicI + CHFAcademicI
        Dash.Range('Z323').Value = CHMAttendanceI + CHFAttendanceI
        Dash.Range('AB323').Value = CHMBehaviorI + CHFBehaviorI
        Dash.Range('AD323').Value = CHMAcademicI + CHMAttendanceI + CHMBehaviorI + CHFAcademicI + CHFAttendanceI + CHFBehaviorI

        Dash.Range('X325').Value = CIMAcademicI + CIFAcademicI
        Dash.Range('Z325').Value = CIMAttendanceI + CIFAttendanceI
        Dash.Range('AB325').Value = CIMBehaviorI + CIFBehaviorI
        Dash.Range('AD325').Value = CIMAcademicI + CIMAttendanceI + CIMBehaviorI + CIFAcademicI + CIFAttendanceI + CIFBehaviorI

        Dash.Range('X327').Value = CWMAcademicI + CWFAcademicI
        Dash.Range('Z327').Value = CWMAttendanceI + CWFAttendanceI
        Dash.Range('AB327').Value = CWMBehaviorI + CWFBehaviorI
        Dash.Range('AD327').Value = CWMAcademicI + CWMAttendanceI + CWMBehaviorI + CWFAcademicI + CWFAttendanceI + CWFBehaviorI

        Dash.Range('X329').Value = CPMAcademicI + CPFAcademicI
        Dash.Range('Z329').Value = CPMAttendanceI + CPFAttendanceI
        Dash.Range('AB329').Value = CPMBehaviorI + CPFBehaviorI
        Dash.Range('AD329').Value = CPMAcademicI + CPMAttendanceI + CPMBehaviorI + CPFAcademicI + CPFAttendanceI + CPFBehaviorI



        Dash.Range('X335').Value = CAMAcademicI + CBMAcademicI + CHMAcademicI + CIMAcademicI + CWMAcademicI + CPMAcademicI
        Dash.Range('Z335').Value = CAMAttendanceI + CBMAttendanceI + CHMAttendanceI + CIMAttendanceI + CWMAttendanceI + CPMAttendanceI
        Dash.Range('AB335').Value = CAMBehaviorI + CBMBehaviorI + CHMBehaviorI + CIMBehaviorI + CWMBehaviorI + CPMBehaviorI
        Dash.Range('AD335').Value = CAMAcademicI + CAMAttendanceI + CAMBehaviorI + CBMAcademicI + CBMAttendanceI + CBMBehaviorI + CHMAcademicI + CHMAttendanceI + CHMBehaviorI + CIMAcademicI + CIMAttendanceI + CIMBehaviorI + CWMAcademicI + CWMAttendanceI + CWMBehaviorI + CPMAcademicI + CPMAttendanceI + CPMBehaviorI

        Dash.Range('X337').Value = CAFAcademicI + CBFAcademicI + CHFAcademicI + CIFAcademicI + CWFAcademicI + CPFAcademicI
        Dash.Range('Z337').Value = CAFAttendanceI + CBFAttendanceI + CHFAttendanceI + CIFAttendanceI + CWFAttendanceI + CPFAttendanceI
        Dash.Range('AB337').Value = CAFBehaviorI + CBFBehaviorI + CHFBehaviorI + CIFBehaviorI + CWFBehaviorI + CPFBehaviorI
        Dash.Range('AD337').Value = CAFAcademicI + CAFAttendanceI + CAFBehaviorI + CBFAcademicI + CBFAttendanceI + CBFBehaviorI + CHFAcademicI + CHFAttendanceI + CHFBehaviorI + CIFAcademicI + CIFAttendanceI + CIFBehaviorI + CWFAcademicI + CWFAttendanceI + CWFBehaviorI + CPFAcademicI + CPFAttendanceI + CPFBehaviorI

        Dash.Range('AI291').Value = CAMAcademicD
        Dash.Range('AK291').Value = CAMAttendanceD
        Dash.Range('AM291').Value = CAMBehaviorD
        Dash.Range('AO291').Value = CAMAcademicD+ CAMAttendanceD+ CAMBehaviorD

        Dash.Range('AI293').Value = CAFAcademicD
        Dash.Range('AK293').Value = CAFAttendanceD
        Dash.Range('AM293').Value = CAFBehaviorD
        Dash.Range('AO293').Value = CAFAcademicD+ CAFAttendanceD+ CAFBehaviorD

        Dash.Range('AI295').Value = CBMAcademicD
        Dash.Range('AK295').Value = CBMAttendanceD
        Dash.Range('AM295').Value = CBMBehaviorD
        Dash.Range('AO295').Value = CBMAcademicD+ CBMAttendanceD+ CBMBehaviorD

        Dash.Range('AI297').Value = CBFAcademicD
        Dash.Range('AK297').Value = CBFAttendanceD
        Dash.Range('AM297').Value = CBFBehaviorD
        Dash.Range('AO297').Value = CBFAcademicD+ CBFAttendanceD+ CBFBehaviorD

        Dash.Range('AI299').Value = CHMAcademicD
        Dash.Range('AK299').Value = CHMAttendanceD
        Dash.Range('AM299').Value = CHMBehaviorD
        Dash.Range('AO299').Value = CHMAcademicD+ CHMAttendanceD+ CHMBehaviorD

        Dash.Range('AI301').Value = CHFAcademicD
        Dash.Range('AK301').Value = CHFAttendanceD
        Dash.Range('AM301').Value = CHFBehaviorD
        Dash.Range('AO301').Value = CHFAcademicD+ CHFAttendanceD+ CHFBehaviorD

        Dash.Range('AI303').Value = CIMAcademicD
        Dash.Range('AK303').Value = CIMAttendanceD
        Dash.Range('AM303').Value = CIMBehaviorD
        Dash.Range('AO303').Value = CIMAcademicD+ CIMAttendanceD+ CIMBehaviorD

        Dash.Range('AI305').Value = CIFAcademicD
        Dash.Range('AK305').Value = CIFAttendanceD
        Dash.Range('AM305').Value = CIFBehaviorD
        Dash.Range('AO305').Value = CIFAcademicD+ CIFAttendanceD+ CIFBehaviorD

        Dash.Range('AI307').Value = CWMAcademicD
        Dash.Range('AK307').Value = CWMAttendanceD
        Dash.Range('AM307').Value = CWMBehaviorD
        Dash.Range('AO307').Value = CWMAcademicD+ CWMAttendanceD+ CWMBehaviorD

        Dash.Range('AI309').Value = CWFAcademicD
        Dash.Range('AK309').Value = CWFAttendanceD
        Dash.Range('AM309').Value = CWFBehaviorD
        Dash.Range('AO309').Value = CWFAcademicD+ CWFAttendanceD+ CWFBehaviorD

        Dash.Range('AI311').Value = CPMAcademicD
        Dash.Range('AK311').Value = CPMAttendanceD
        Dash.Range('AM311').Value = CPMBehaviorD
        Dash.Range('AO311').Value = CPMAcademicD+ CPMAttendanceD+ CPMBehaviorD

        Dash.Range('AI313').Value = CPFAcademicD
        Dash.Range('AK313').Value = CPFAttendanceD
        Dash.Range('AM313').Value = CPFBehaviorD
        Dash.Range('AO313').Value = CPFAcademicD+ CPFAttendanceD+ CPFBehaviorD



        Dash.Range('AI319').Value = CAMAcademicD+ CAFAcademicD
        Dash.Range('AK319').Value = CAMAttendanceD+ CAFAttendanceD
        Dash.Range('AM319').Value = CAMBehaviorD+ CAFBehaviorD
        Dash.Range('AO319').Value = CAMAcademicD+ CAMAttendanceD+ CAMBehaviorD+ CAFAcademicD+ CAFAttendanceD+ CAFBehaviorD

        Dash.Range('AI321').Value = CBMAcademicD+ CBFAcademicD
        Dash.Range('AK321').Value = CBMAttendanceD+ CBFAttendanceD
        Dash.Range('AM321').Value = CBMBehaviorD+ CBFBehaviorD
        Dash.Range('AO321').Value = CBMAcademicD+ CBMAttendanceD+ CBMBehaviorD+ CBFAcademicD+ CBFAttendanceD+ CBFBehaviorD

        Dash.Range('AI323').Value = CHMAcademicD+ CHFAcademicD
        Dash.Range('AK323').Value = CHMAttendanceD+ CHFAttendanceD
        Dash.Range('AM323').Value = CHMBehaviorD+ CHFBehaviorD
        Dash.Range('AO323').Value = CHMAcademicD+ CHMAttendanceD+ CHMBehaviorD+ CHFAcademicD+ CHFAttendanceD+ CHFBehaviorD

        Dash.Range('AI325').Value = CIMAcademicD+ CIFAcademicD
        Dash.Range('AK325').Value = CIMAttendanceD+ CIFAttendanceD
        Dash.Range('AM325').Value = CIMBehaviorD+ CIFBehaviorD
        Dash.Range('AO325').Value = CIMAcademicD+ CIMAttendanceD+ CIMBehaviorD+ CIFAcademicD+ CIFAttendanceD+ CIFBehaviorD

        Dash.Range('AI327').Value = CWMAcademicD+ CWFAcademicD
        Dash.Range('AK327').Value = CWMAttendanceD+ CWFAttendanceD
        Dash.Range('AM327').Value = CWMBehaviorD+ CWFBehaviorD
        Dash.Range('AO327').Value = CWMAcademicD+ CWMAttendanceD+ CWMBehaviorD+ CWFAcademicD+ CWFAttendanceD+ CWFBehaviorD

        Dash.Range('AI329').Value = CPMAcademicD+ CPFAcademicD
        Dash.Range('AK329').Value = CPMAttendanceD+ CPFAttendanceD
        Dash.Range('AM329').Value = CPMBehaviorD+ CPFBehaviorD
        Dash.Range('AO329').Value = CPMAcademicD+ CPMAttendanceD+ CPMBehaviorD+ CPFAcademicD+ CPFAttendanceD+ CPFBehaviorD



        Dash.Range('AI335').Value = CAMAcademicD+ CBMAcademicD+ CHMAcademicD+ CIMAcademicD+ CWMAcademicD+ CPMAcademicD
        Dash.Range('AK335').Value = CAMAttendanceD+ CBMAttendanceD+ CHMAttendanceD+ CIMAttendanceD+ CWMAttendanceD+ CPMAttendanceD
        Dash.Range('AM335').Value = CAMBehaviorD+ CBMBehaviorD+ CHMBehaviorD+ CIMBehaviorD+ CWMBehaviorD+ CPMBehaviorD
        Dash.Range('AO335').Value = CAMAcademicD+ CAMAttendanceD+ CAMBehaviorD+ CBMAcademicD+ CBMAttendanceD+ CBMBehaviorD+ CHMAcademicD+ CHMAttendanceD+ CHMBehaviorD+ CIMAcademicD+ CIMAttendanceD+ CIMBehaviorD+ CWMAcademicD+ CWMAttendanceD+ CWMBehaviorD+ CPMAcademicD+ CPMAttendanceD+ CPMBehaviorD

        Dash.Range('AI337').Value = CAFAcademicD+ CBFAcademicD+ CHFAcademicD+ CIFAcademicD+ CWFAcademicD+ CPFAcademicD
        Dash.Range('AK337').Value = CAFAttendanceD+ CBFAttendanceD+ CHFAttendanceD+ CIFAttendanceD+ CWFAttendanceD+ CPFAttendanceD
        Dash.Range('AM337').Value = CAFBehaviorD+ CBFBehaviorD+ CHFBehaviorD+ CIFBehaviorD+ CWFBehaviorD+ CPFBehaviorD
        Dash.Range('AO337').Value = CAFAcademicD+ CAFAttendanceD+ CAFBehaviorD+ CBFAcademicD+ CBFAttendanceD+ CBFBehaviorD+ CHFAcademicD+ CHFAttendanceD+ CHFBehaviorD+ CIFAcademicD+ CIFAttendanceD+ CIFBehaviorD+ CWFAcademicD+ CWFAttendanceD+ CWFBehaviorD+ CPFAcademicD+ CPFAttendanceD+ CPFBehaviorD


        target.Close(SaveChanges=True)
        target2.Close(SaveChanges=True)

        xlApp = win32.Dispatch('Excel.Application')
        target = xlApp.Workbooks.Open(pathTwo)
        target2 = xlApp.Workbooks.Open(pathOne)
        Mix = target.Worksheets(locationTwo)
        Assess = target.Worksheets(locationOne)
        Campus = target.Worksheets(locationFour)
        Progress = target.Worksheets(locationFive)
        Dash = target.Worksheets("Dashboard")
        
        chiltonAM = 0
        chiltonAF = 0
        chiltonBM = 0
        chiltonBF = 0
        chiltonHM = 0
        chiltonHF = 0
        chiltonNM = 0
        chiltonNF = 0
        chiltonWM = 0
        chiltonWF = 0
        chiltonPM = 0
        chiltonPF = 0

        engeAM = 0
        engeAF = 0
        engeBM = 0
        engeBF = 0
        engeHM = 0
        engeHF = 0
        engeNM = 0
        engeNF = 0
        engeWM = 0
        engeWF = 0
        engePM = 0
        engePF = 0

        grosHSAM = 0
        grosHSAF = 0
        grosHSBM = 0
        grosHSBF = 0
        grosHSHM = 0
        grosHSHF = 0
        grosHSNM = 0
        grosHSNF = 0
        grosHSWM = 0
        grosHSWF = 0
        grosHSPM = 0
        grosHSPF = 0

        grosMSAM = 0
        grosMSAF = 0
        grosMSBM = 0
        grosMSBF = 0
        grosMSHM = 0
        grosMSHF = 0
        grosMSNM = 0
        grosMSNF = 0
        grosMSWM = 0
        grosMSWF = 0
        grosMSPM = 0
        grosMSPF = 0

        hurstAM = 0
        hurstAF = 0
        hurstBM = 0
        hurstBF = 0
        hurstHM = 0
        hurstHF = 0
        hurstNM = 0
        hurstNF = 0
        hurstWM = 0
        hurstWF = 0
        hurstPM = 0
        hurstPF = 0

        vegaElAM = 0
        vegaElAF = 0
        vegaElBM = 0
        vegaElBF = 0
        vegaElHM = 0
        vegaElHF = 0
        vegaElNM = 0
        vegaElNF = 0
        vegaElWM = 0
        vegaElWF = 0
        vegaElPM = 0
        vegaElPF = 0

        vegaHSAM = 0
        vegaHSAF = 0
        vegaHSBM = 0
        vegaHSBF = 0
        vegaHSHM = 0
        vegaHSHF = 0
        vegaHSNM = 0
        vegaHSNF = 0
        vegaHSWM = 0
        vegaHSWF = 0
        vegaHSPM = 0
        vegaHSPF = 0

        vegaINTAM = 0
        vegaINTAF = 0
        vegaINTBM = 0
        vegaINTBF = 0
        vegaINTHM = 0
        vegaINTHF = 0
        vegaINTNM = 0
        vegaINTNF = 0
        vegaINTWM = 0
        vegaINTWF = 0
        vegaINTPM = 0
        vegaINTPF = 0

        vegaJHAM = 0
        vegaJHAF = 0
        vegaJHBM = 0
        vegaJHBF = 0
        vegaJHHM = 0
        vegaJHHF = 0
        vegaJHNM = 0
        vegaJHNF = 0
        vegaJHWM = 0
        vegaJHWF = 0
        vegaJHPM = 0
        vegaJHPF = 0

        vegaPRIAM = 0
        vegaPRIAF = 0
        vegaPRIBM = 0
        vegaPRIBF = 0
        vegaPRIHM = 0
        vegaPRIHF = 0
        vegaPRINM = 0
        vegaPRINF = 0
        vegaPRIWM = 0
        vegaPRIWF = 0
        vegaPRIPM = 0
        vegaPRIPF = 0

        mexiaHSAM = 0
        mexiaHSAF = 0
        mexiaHSBM = 0
        mexiaHSBF = 0
        mexiaHSHM = 0
        mexiaHSHF = 0
        mexiaHSNM = 0
        mexiaHSNF = 0
        mexiaHSWM = 0
        mexiaHSWF = 0
        mexiaHSPM = 0
        mexiaHSPF = 0

        mexiaJHAM = 0
        mexiaJHAF = 0
        mexiaJHBM = 0
        mexiaJHBF = 0
        mexiaJHHM = 0
        mexiaJHHF = 0
        mexiaJHNM = 0
        mexiaJHNF = 0
        mexiaJHWM = 0
        mexiaJHWF = 0
        mexiaJHPM = 0
        mexiaJHPF = 0

        creekELAM = 0
        creekELAF = 0
        creekELBM = 0
        creekELBF = 0
        creekELHM = 0
        creekELHF = 0
        creekELNM = 0
        creekELNF = 0
        creekELWM = 0
        creekELWF = 0
        creekELPM = 0
        creekELPF = 0

        hewittELAM = 0
        hewittELAF = 0
        hewittELBM = 0
        hewittELBF = 0
        hewittELHM = 0
        hewittELHF = 0
        hewittELNM = 0
        hewittELNF = 0
        hewittELWM = 0
        hewittELWF = 0
        hewittELPM = 0
        hewittELPF = 0

        midwayHSAM = 0
        midwayHSAF = 0
        midwayHSBM = 0
        midwayHSBF = 0
        midwayHSHM = 0
        midwayHSHF = 0
        midwayHSNM = 0
        midwayHSNF = 0
        midwayHSWM = 0
        midwayHSWF = 0
        midwayHSPM = 0
        midwayHSPF = 0

        midwayMSAM = 0
        midwayMSAF = 0
        midwayMSBM = 0
        midwayMSBF = 0
        midwayMSHM = 0
        midwayMSHF = 0
        midwayMSNM = 0
        midwayMSNF = 0
        midwayMSWM = 0
        midwayMSWF = 0
        midwayMSPM = 0
        midwayMSPF = 0

        quinnMSAM = 0
        quinnMSAF = 0
        quinnMSBM = 0
        quinnMSBF = 0
        quinnMSHM = 0
        quinnMSHF = 0
        quinnMSNM = 0
        quinnMSNF = 0
        quinnMSWM = 0
        quinnMSWF = 0
        quinnMSPM = 0
        quinnMSPF = 0

        teagueHSAM = 0
        teagueHSAF = 0
        teagueHSBM = 0
        teagueHSBF = 0
        teagueHSHM = 0
        teagueHSHF = 0
        teagueHSNM = 0
        teagueHSNF = 0
        teagueHSWM = 0
        teagueHSWF = 0
        teagueHSPM = 0
        teagueHSPF = 0

        wacoCharterAM = 0
        wacoCharterAF = 0
        wacoCharterBM = 0
        wacoCharterBF = 0
        wacoCharterHM = 0
        wacoCharterHF = 0
        wacoCharterNM = 0
        wacoCharterNF = 0
        wacoCharterWM = 0
        wacoCharterWF = 0
        wacoCharterPM = 0
        wacoCharterPF = 0

        vistaElAM = 0
        vistaElAF = 0
        vistaElBM = 0
        vistaElBF = 0
        vistaElHM = 0
        vistaElHF = 0
        vistaElNM = 0
        vistaElNF = 0
        vistaElWM = 0
        vistaElWF = 0
        vistaElPM = 0
        vistaElPF = 0

        brazosHSAM = 0
        brazosHSAF = 0
        brazosHSBM = 0
        brazosHSBF = 0
        brazosHSHM = 0
        brazosHSHF = 0
        brazosHSNM = 0
        brazosHSNF = 0
        brazosHSWM = 0
        brazosHSWF = 0
        brazosHSPM = 0
        brazosHSPF = 0

        brookELAM = 0
        brookELAF = 0
        brookELBM = 0
        brookELBF = 0
        brookELHM = 0
        brookELHF = 0
        brookELNM = 0
        brookELNF = 0
        brookELWM = 0
        brookELWF = 0
        brookELPM = 0
        brookELPF = 0

        chavezAM = 0
        chavezAF = 0
        chavezBM = 0
        chavezBF = 0
        chavezHM = 0
        chavezHF = 0
        chavezNM = 0
        chavezNF = 0
        chavezWM = 0
        chavezWF = 0
        chavezPM = 0
        chavezPF = 0

        carverMSAM = 0
        carverMSAF = 0
        carverMSBM = 0
        carverMSBF = 0
        carverMSHM = 0
        carverMSHF = 0
        carverMSNM = 0
        carverMSNF = 0
        carverMSWM = 0
        carverMSWF = 0
        carverMSPM = 0
        carverMSPF = 0

        hinesElAM = 0
        hinesElAF = 0
        hinesElBM = 0
        hinesElBF = 0
        hinesElHM = 0
        hinesElHF = 0
        hinesElNM = 0
        hinesElNF = 0
        hinesElWM = 0
        hinesElWF = 0
        hinesElPM = 0
        hinesElPF = 0

        kendrickElAM = 0
        kendrickElAF = 0
        kendrickElBM = 0
        kendrickElBF = 0
        kendrickElHM = 0
        kendrickElHF = 0
        kendrickElNM = 0
        kendrickElNF = 0
        kendrickElWM = 0
        kendrickElWF = 0
        kendrickElPM = 0
        kendrickElPF = 0

        heightsELAM = 0
        heightsELAF = 0
        heightsELBM = 0
        heightsELBF = 0
        heightsELHM = 0
        heightsELHF = 0
        heightsELNM = 0
        heightsELNF = 0
        heightsELWM = 0
        heightsELWF = 0
        heightsELPM = 0
        heightsELPF = 0

        tennysonHSAM = 0
        tennysonHSAF = 0
        tennysonHSBM = 0
        tennysonHSBF = 0
        tennysonHSHM = 0
        tennysonHSHF = 0
        tennysonHSNM = 0
        tennysonHSNF = 0
        tennysonHSWM = 0
        tennysonHSWF = 0
        tennysonHSPM = 0
        tennysonHSPF = 0

        universityHSAM = 0
        universityHSAF = 0
        universityHSBM = 0
        universityHSBF = 0
        universityHSHM = 0
        universityHSHF = 0
        universityHSNM = 0
        universityHSNF = 0
        universityHSWM = 0
        universityHSWF = 0
        universityHSPM = 0
        universityHSPF = 0

        wacoHSAM = 0
        wacoHSAF = 0
        wacoHSBM = 0
        wacoHSBF = 0
        wacoHSHM = 0
        wacoHSHF = 0
        wacoHSNM = 0
        wacoHSNF = 0
        wacoHSWM = 0
        wacoHSWF = 0
        wacoHSPM = 0
        wacoHSPF = 0

        rowShifter = 2
        whileLoopCondition = False

        while(whileLoopCondition != True):

            if("CHILTON" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chiltonPF += 1

            elif("ENGE-WASHINGTON" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engeWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engePM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    engePF += 1

            elif("GROESBECK H S" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosHSPF += 1

            elif("GROESBECK MIDDLE" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    grosMSPF += 1

            elif("H O WHITEHURST EL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hurstPF += 1

            elif("LA VEGA EL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaElPF += 1

            elif("LA VEGA H S" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaHSPF += 1

            elif("LA VEGA INT H P MILES CAMPUS" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaINTPF += 1

            elif("LA VEGA J H GEORGE DIXON CAMPUS" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaJHPF += 1

            elif("LA VEGA PRI PHIL BANCALE CAMPUS" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRINM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRINF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vegaPRIPF += 1

            elif("MEXIA H S" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaHSPF += 1

            elif("MEXIA J H" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    mexiaJHPF += 1

            elif("CASTLEMAN CREEK EL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    creekELPF += 1

            elif("HEWITT EL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hewittELPF += 1

            elif("MIDWAY H S" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayHSPF += 1

            elif("MIDWAY MIDDLE" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    midwayMSPF += 1

            elif("QUINN CAMPUS PUBLIC MIDDLE" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    quinnMSPF += 1

            elif("TEAGUE H S" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    teagueHSPF += 1

            elif("WACO CHARTER SCHOOL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoCharterPF += 1

            elif("ALTA VISTA EL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    vistaElPF += 1

            elif("BRAZOS H S" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brazosHSPF += 1

            elif("BROOK AVENUE EL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    brookELPF += 1

            elif("CESAR CHAVEZ MIDDLE" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    chavezPF += 1

            elif("J H HINES EL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    hinesElPF += 1

            elif("G W CARVER MIDDLE" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    carverMSPF += 1
                    
            elif("KENDRICK EL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    kendrickElPF += 1
                
            elif("PROVIDENT HEIGHTS EL" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    heightsELPF += 1

            elif("TENNYSON MIDDLE" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    tennysonHSPF += 1

            elif("UNIVERSITY H S" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    universityHSPF += 1
                
            elif("WACO H S" in str(Progress.Range('G'+str(rowShifter)).Value)):
                if("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSAM += 1
                elif("A" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSAF += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSBM += 1
                elif("B" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSBF += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSHM += 1
                elif("H" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSHF += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSNM += 1
                elif("I" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSNF += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSWM += 1
                elif("W" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSWF += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "M" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSPM += 1
                elif("P" in str(Progress.Range('X'+str(rowShifter)).Value) and "F" in str(Progress.Range('U'+str(rowShifter)).Value)):
                    wacoHSPF += 1

            rowShifter += 1

            if(Progress.Range('A'+str(rowShifter)).Value == None):
                whileLoopCondition = True

        moveSpace = 354
        while(str(Dash.Range('B'+str(moveSpace)).Value) != 'None'):
            if("CHILTON" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = chiltonAM; 
                Dash.Range('F'+str(moveSpace)).Value = chiltonAF; 
                Dash.Range('H'+str(moveSpace)).Value = chiltonBM; 
                Dash.Range('J'+str(moveSpace)).Value = chiltonBF; 
                Dash.Range('M'+str(moveSpace)).Value = chiltonHM; 
                Dash.Range('O'+str(moveSpace)).Value = chiltonHF; 
                Dash.Range('Q'+str(moveSpace)).Value = chiltonNM; 
                Dash.Range('S'+str(moveSpace)).Value = chiltonNF; 
                Dash.Range('U'+str(moveSpace)).Value = chiltonWM; 
                Dash.Range('X'+str(moveSpace)).Value = chiltonWF; 
                Dash.Range('Z'+str(moveSpace)).Value = chiltonPM; 
                Dash.Range('AB'+str(moveSpace)).Value = chiltonPF;
            if("ENGE-WASHINGTON" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = engeAM; 
                Dash.Range('F'+str(moveSpace)).Value = engeAF; 
                Dash.Range('H'+str(moveSpace)).Value = engeBM; 
                Dash.Range('J'+str(moveSpace)).Value = engeBF; 
                Dash.Range('M'+str(moveSpace)).Value = engeHM; 
                Dash.Range('O'+str(moveSpace)).Value = engeHF; 
                Dash.Range('Q'+str(moveSpace)).Value = engeNM; 
                Dash.Range('S'+str(moveSpace)).Value = engeNF; 
                Dash.Range('U'+str(moveSpace)).Value = engeWM; 
                Dash.Range('X'+str(moveSpace)).Value = engeWF; 
                Dash.Range('Z'+str(moveSpace)).Value = engePM; 
                Dash.Range('AB'+str(moveSpace)).Value = engePF;
            if("GROESBECK H S" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = grosHSAM; 
                Dash.Range('F'+str(moveSpace)).Value = grosHSAF; 
                Dash.Range('H'+str(moveSpace)).Value = grosHSBM; 
                Dash.Range('J'+str(moveSpace)).Value = grosHSBF; 
                Dash.Range('M'+str(moveSpace)).Value = grosHSHM; 
                Dash.Range('O'+str(moveSpace)).Value = grosHSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = grosHSNM; 
                Dash.Range('S'+str(moveSpace)).Value = grosHSNF; 
                Dash.Range('U'+str(moveSpace)).Value = grosHSWM; 
                Dash.Range('X'+str(moveSpace)).Value = grosHSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = grosHSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = grosHSPF;
            if("GROESBECK MIDDLE" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = grosMSAM; 
                Dash.Range('F'+str(moveSpace)).Value = grosMSAF; 
                Dash.Range('H'+str(moveSpace)).Value = grosMSBM; 
                Dash.Range('J'+str(moveSpace)).Value = grosMSBF; 
                Dash.Range('M'+str(moveSpace)).Value = grosMSHM; 
                Dash.Range('O'+str(moveSpace)).Value = grosMSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = grosMSNM; 
                Dash.Range('S'+str(moveSpace)).Value = grosMSNF; 
                Dash.Range('U'+str(moveSpace)).Value = grosMSWM; 
                Dash.Range('X'+str(moveSpace)).Value = grosMSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = grosMSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = grosMSPF;
            if("H O WHITEHURST" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = hurstAM; 
                Dash.Range('F'+str(moveSpace)).Value = hurstAF; 
                Dash.Range('H'+str(moveSpace)).Value = hurstBM; 
                Dash.Range('J'+str(moveSpace)).Value = hurstBF; 
                Dash.Range('M'+str(moveSpace)).Value = hurstHM; 
                Dash.Range('O'+str(moveSpace)).Value = hurstHF; 
                Dash.Range('Q'+str(moveSpace)).Value = hurstNM; 
                Dash.Range('S'+str(moveSpace)).Value = hurstNF; 
                Dash.Range('U'+str(moveSpace)).Value = hurstWM; 
                Dash.Range('X'+str(moveSpace)).Value = hurstWF; 
                Dash.Range('Z'+str(moveSpace)).Value = hurstPM; 
                Dash.Range('AB'+str(moveSpace)).Value = hurstPF;
            if("LA VEGA EL" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = vegaElAM; 
                Dash.Range('F'+str(moveSpace)).Value = vegaElAF; 
                Dash.Range('H'+str(moveSpace)).Value = vegaElBM; 
                Dash.Range('J'+str(moveSpace)).Value = vegaElBF; 
                Dash.Range('M'+str(moveSpace)).Value = vegaElHM; 
                Dash.Range('O'+str(moveSpace)).Value = vegaElHF; 
                Dash.Range('Q'+str(moveSpace)).Value = vegaElNM; 
                Dash.Range('S'+str(moveSpace)).Value = vegaElNF; 
                Dash.Range('U'+str(moveSpace)).Value = vegaElWM; 
                Dash.Range('X'+str(moveSpace)).Value = vegaElWF; 
                Dash.Range('Z'+str(moveSpace)).Value = vegaElPM; 
                Dash.Range('AB'+str(moveSpace)).Value = vegaElPF;
            if("LA VEGA H S" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = vegaHSAM; 
                Dash.Range('F'+str(moveSpace)).Value = vegaHSAF; 
                Dash.Range('H'+str(moveSpace)).Value = vegaHSBM; 
                Dash.Range('J'+str(moveSpace)).Value = vegaHSBF; 
                Dash.Range('M'+str(moveSpace)).Value = vegaHSHM; 
                Dash.Range('O'+str(moveSpace)).Value = vegaHSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = vegaHSNM; 
                Dash.Range('S'+str(moveSpace)).Value = vegaHSNF; 
                Dash.Range('U'+str(moveSpace)).Value = vegaHSWM; 
                Dash.Range('X'+str(moveSpace)).Value = vegaHSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = vegaHSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = vegaHSPF;
            if("LA VEGA INT" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = vegaINTAM; 
                Dash.Range('F'+str(moveSpace)).Value = vegaINTAF; 
                Dash.Range('H'+str(moveSpace)).Value = vegaINTBM; 
                Dash.Range('J'+str(moveSpace)).Value = vegaINTBF; 
                Dash.Range('M'+str(moveSpace)).Value = vegaINTHM; 
                Dash.Range('O'+str(moveSpace)).Value = vegaINTHF; 
                Dash.Range('Q'+str(moveSpace)).Value = vegaINTNM; 
                Dash.Range('S'+str(moveSpace)).Value = vegaINTNF; 
                Dash.Range('U'+str(moveSpace)).Value = vegaINTWM; 
                Dash.Range('X'+str(moveSpace)).Value = vegaINTWF; 
                Dash.Range('Z'+str(moveSpace)).Value = vegaINTPM; 
                Dash.Range('AB'+str(moveSpace)).Value = vegaINTPF;
            if("LA VEGA J H" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = vegaJHAM; 
                Dash.Range('F'+str(moveSpace)).Value = vegaJHAF; 
                Dash.Range('H'+str(moveSpace)).Value = vegaJHBM; 
                Dash.Range('J'+str(moveSpace)).Value = vegaJHBF; 
                Dash.Range('M'+str(moveSpace)).Value = vegaJHHM; 
                Dash.Range('O'+str(moveSpace)).Value = vegaJHHF; 
                Dash.Range('Q'+str(moveSpace)).Value = vegaJHNM; 
                Dash.Range('S'+str(moveSpace)).Value = vegaJHNF; 
                Dash.Range('U'+str(moveSpace)).Value = vegaJHWM; 
                Dash.Range('X'+str(moveSpace)).Value = vegaJHWF; 
                Dash.Range('Z'+str(moveSpace)).Value = vegaJHPM; 
                Dash.Range('AB'+str(moveSpace)).Value = vegaJHPF;
            if("LA VEGA PRIMARY" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = vegaPRIAM; 
                Dash.Range('F'+str(moveSpace)).Value = vegaPRIAF; 
                Dash.Range('H'+str(moveSpace)).Value = vegaPRIBM; 
                Dash.Range('J'+str(moveSpace)).Value = vegaPRIBF; 
                Dash.Range('M'+str(moveSpace)).Value = vegaPRIHM; 
                Dash.Range('O'+str(moveSpace)).Value = vegaPRIHF; 
                Dash.Range('Q'+str(moveSpace)).Value = vegaPRINM; 
                Dash.Range('S'+str(moveSpace)).Value = vegaPRINF; 
                Dash.Range('U'+str(moveSpace)).Value = vegaPRIWM; 
                Dash.Range('X'+str(moveSpace)).Value = vegaPRIWF; 
                Dash.Range('Z'+str(moveSpace)).Value = vegaPRIPM; 
                Dash.Range('AB'+str(moveSpace)).Value = vegaPRIPF;
            if("MEXIA H S" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = mexiaHSAM; 
                Dash.Range('F'+str(moveSpace)).Value = mexiaHSAF; 
                Dash.Range('H'+str(moveSpace)).Value = mexiaHSBM; 
                Dash.Range('J'+str(moveSpace)).Value = mexiaHSBF; 
                Dash.Range('M'+str(moveSpace)).Value = mexiaHSHM; 
                Dash.Range('O'+str(moveSpace)).Value = mexiaHSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = mexiaHSNM; 
                Dash.Range('S'+str(moveSpace)).Value = mexiaHSNF; 
                Dash.Range('U'+str(moveSpace)).Value = mexiaHSWM; 
                Dash.Range('X'+str(moveSpace)).Value = mexiaHSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = mexiaHSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = mexiaHSPF;
            if("MEXIA J H" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = mexiaJHAM; 
                Dash.Range('F'+str(moveSpace)).Value = mexiaJHAF; 
                Dash.Range('H'+str(moveSpace)).Value = mexiaJHBM; 
                Dash.Range('J'+str(moveSpace)).Value = mexiaJHBF; 
                Dash.Range('M'+str(moveSpace)).Value = mexiaJHHM; 
                Dash.Range('O'+str(moveSpace)).Value = mexiaJHHF; 
                Dash.Range('Q'+str(moveSpace)).Value = mexiaJHNM; 
                Dash.Range('S'+str(moveSpace)).Value = mexiaJHNF; 
                Dash.Range('U'+str(moveSpace)).Value = mexiaJHWM; 
                Dash.Range('X'+str(moveSpace)).Value = mexiaJHWF; 
                Dash.Range('Z'+str(moveSpace)).Value = mexiaJHPM; 
                Dash.Range('AB'+str(moveSpace)).Value = mexiaJHPF;
            if("Creek EL" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = creekELAM; 
                Dash.Range('F'+str(moveSpace)).Value = creekELAF; 
                Dash.Range('H'+str(moveSpace)).Value = creekELBM; 
                Dash.Range('J'+str(moveSpace)).Value = creekELBF; 
                Dash.Range('M'+str(moveSpace)).Value = creekELHM; 
                Dash.Range('O'+str(moveSpace)).Value = creekELHF; 
                Dash.Range('Q'+str(moveSpace)).Value = creekELNM; 
                Dash.Range('S'+str(moveSpace)).Value = creekELNF; 
                Dash.Range('U'+str(moveSpace)).Value = creekELWM; 
                Dash.Range('X'+str(moveSpace)).Value = creekELWF; 
                Dash.Range('Z'+str(moveSpace)).Value = creekELPM; 
                Dash.Range('AB'+str(moveSpace)).Value = creekELPF;
            if("HEWITT EL" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = hewittELAM; 
                Dash.Range('F'+str(moveSpace)).Value = hewittELAF; 
                Dash.Range('H'+str(moveSpace)).Value = hewittELBM; 
                Dash.Range('J'+str(moveSpace)).Value = hewittELBF; 
                Dash.Range('M'+str(moveSpace)).Value = hewittELHM; 
                Dash.Range('O'+str(moveSpace)).Value = hewittELHF; 
                Dash.Range('Q'+str(moveSpace)).Value = hewittELNM; 
                Dash.Range('S'+str(moveSpace)).Value = hewittELNF; 
                Dash.Range('U'+str(moveSpace)).Value = hewittELWM; 
                Dash.Range('X'+str(moveSpace)).Value = hewittELWF; 
                Dash.Range('Z'+str(moveSpace)).Value = hewittELPM; 
                Dash.Range('AB'+str(moveSpace)).Value = hewittELPF;
            if("MIDWAY H S" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = midwayHSAM; 
                Dash.Range('F'+str(moveSpace)).Value = midwayHSAF; 
                Dash.Range('H'+str(moveSpace)).Value = midwayHSBM; 
                Dash.Range('J'+str(moveSpace)).Value = midwayHSBF; 
                Dash.Range('M'+str(moveSpace)).Value = midwayHSHM; 
                Dash.Range('O'+str(moveSpace)).Value = midwayHSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = midwayHSNM; 
                Dash.Range('S'+str(moveSpace)).Value = midwayHSNF; 
                Dash.Range('U'+str(moveSpace)).Value = midwayHSWM; 
                Dash.Range('X'+str(moveSpace)).Value = midwayHSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = midwayHSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = midwayHSPF;
            if("MIDWAY MIDDLE" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = midwayMSAM; 
                Dash.Range('F'+str(moveSpace)).Value = midwayMSAF; 
                Dash.Range('H'+str(moveSpace)).Value = midwayMSBM; 
                Dash.Range('J'+str(moveSpace)).Value = midwayMSBF; 
                Dash.Range('M'+str(moveSpace)).Value = midwayMSHM; 
                Dash.Range('O'+str(moveSpace)).Value = midwayMSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = midwayMSNM; 
                Dash.Range('S'+str(moveSpace)).Value = midwayMSNF; 
                Dash.Range('U'+str(moveSpace)).Value = midwayMSWM; 
                Dash.Range('X'+str(moveSpace)).Value = midwayMSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = midwayMSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = midwayMSPF;
            if("QUINN MIDDLE" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = quinnMSAM; 
                Dash.Range('F'+str(moveSpace)).Value = quinnMSAF; 
                Dash.Range('H'+str(moveSpace)).Value = quinnMSBM; 
                Dash.Range('J'+str(moveSpace)).Value = quinnMSBF; 
                Dash.Range('M'+str(moveSpace)).Value = quinnMSHM; 
                Dash.Range('O'+str(moveSpace)).Value = quinnMSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = quinnMSNM; 
                Dash.Range('S'+str(moveSpace)).Value = quinnMSNF; 
                Dash.Range('U'+str(moveSpace)).Value = quinnMSWM; 
                Dash.Range('X'+str(moveSpace)).Value = quinnMSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = quinnMSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = quinnMSPF;
            if("TEAGUE H S" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = teagueHSAM; 
                Dash.Range('F'+str(moveSpace)).Value = teagueHSAF; 
                Dash.Range('H'+str(moveSpace)).Value = teagueHSBM; 
                Dash.Range('J'+str(moveSpace)).Value = teagueHSBF; 
                Dash.Range('M'+str(moveSpace)).Value = teagueHSHM; 
                Dash.Range('O'+str(moveSpace)).Value = teagueHSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = teagueHSNM; 
                Dash.Range('S'+str(moveSpace)).Value = teagueHSNF; 
                Dash.Range('U'+str(moveSpace)).Value = teagueHSWM; 
                Dash.Range('X'+str(moveSpace)).Value = teagueHSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = teagueHSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = teagueHSPF;
            if("WACO CHARTER" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = wacoCharterAM; 
                Dash.Range('F'+str(moveSpace)).Value = wacoCharterAF; 
                Dash.Range('H'+str(moveSpace)).Value = wacoCharterBM; 
                Dash.Range('J'+str(moveSpace)).Value = wacoCharterBF; 
                Dash.Range('M'+str(moveSpace)).Value = wacoCharterHM; 
                Dash.Range('O'+str(moveSpace)).Value = wacoCharterHF; 
                Dash.Range('Q'+str(moveSpace)).Value = wacoCharterNM; 
                Dash.Range('S'+str(moveSpace)).Value = wacoCharterNF; 
                Dash.Range('U'+str(moveSpace)).Value = wacoCharterWM; 
                Dash.Range('X'+str(moveSpace)).Value = wacoCharterWF; 
                Dash.Range('Z'+str(moveSpace)).Value = wacoCharterPM; 
                Dash.Range('AB'+str(moveSpace)).Value = wacoCharterPF;
            if("ALTA VISTA" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = vistaElAM; 
                Dash.Range('F'+str(moveSpace)).Value = vistaElAF; 
                Dash.Range('H'+str(moveSpace)).Value = vistaElBM; 
                Dash.Range('J'+str(moveSpace)).Value = vistaElBF; 
                Dash.Range('M'+str(moveSpace)).Value = vistaElHM; 
                Dash.Range('O'+str(moveSpace)).Value = vistaElHF; 
                Dash.Range('Q'+str(moveSpace)).Value = vistaElNM; 
                Dash.Range('S'+str(moveSpace)).Value = vistaElNF; 
                Dash.Range('U'+str(moveSpace)).Value = vistaElWM; 
                Dash.Range('X'+str(moveSpace)).Value = vistaElWF; 
                Dash.Range('Z'+str(moveSpace)).Value = vistaElPM; 
                Dash.Range('AB'+str(moveSpace)).Value = vistaElPF;
            if("BRAZOS H S" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = brazosHSAM; 
                Dash.Range('F'+str(moveSpace)).Value = brazosHSAF; 
                Dash.Range('H'+str(moveSpace)).Value = brazosHSBM; 
                Dash.Range('J'+str(moveSpace)).Value = brazosHSBF; 
                Dash.Range('M'+str(moveSpace)).Value = brazosHSHM; 
                Dash.Range('O'+str(moveSpace)).Value = brazosHSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = brazosHSNM; 
                Dash.Range('S'+str(moveSpace)).Value = brazosHSNF; 
                Dash.Range('U'+str(moveSpace)).Value = brazosHSWM; 
                Dash.Range('X'+str(moveSpace)).Value = brazosHSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = brazosHSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = brazosHSPF;
            if("BROOK AVENUE" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = brookELAM; 
                Dash.Range('F'+str(moveSpace)).Value = brookELAF; 
                Dash.Range('H'+str(moveSpace)).Value = brookELBM; 
                Dash.Range('J'+str(moveSpace)).Value = brookELBF; 
                Dash.Range('M'+str(moveSpace)).Value = brookELHM; 
                Dash.Range('O'+str(moveSpace)).Value = brookELHF; 
                Dash.Range('Q'+str(moveSpace)).Value = brookELNM; 
                Dash.Range('S'+str(moveSpace)).Value = brookELNF; 
                Dash.Range('U'+str(moveSpace)).Value = brookELWM; 
                Dash.Range('X'+str(moveSpace)).Value = brookELWF; 
                Dash.Range('Z'+str(moveSpace)).Value = brookELPM; 
                Dash.Range('AB'+str(moveSpace)).Value = brookELPF;
            if("CHAVEZ" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = chavezAM; 
                Dash.Range('F'+str(moveSpace)).Value = chavezAF; 
                Dash.Range('H'+str(moveSpace)).Value = chavezBM; 
                Dash.Range('J'+str(moveSpace)).Value = chavezBF; 
                Dash.Range('M'+str(moveSpace)).Value = chavezHM; 
                Dash.Range('O'+str(moveSpace)).Value = chavezHF; 
                Dash.Range('Q'+str(moveSpace)).Value = chavezNM; 
                Dash.Range('S'+str(moveSpace)).Value = chavezNF; 
                Dash.Range('U'+str(moveSpace)).Value = chavezWM; 
                Dash.Range('X'+str(moveSpace)).Value = chavezWF; 
                Dash.Range('Z'+str(moveSpace)).Value = chavezPM; 
                Dash.Range('AB'+str(moveSpace)).Value = chavezPF;
            if("CARVER MIDDLE" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = carverMSAM; 
                Dash.Range('F'+str(moveSpace)).Value = carverMSAF; 
                Dash.Range('H'+str(moveSpace)).Value = carverMSBM; 
                Dash.Range('J'+str(moveSpace)).Value = carverMSBF; 
                Dash.Range('M'+str(moveSpace)).Value = carverMSHM; 
                Dash.Range('O'+str(moveSpace)).Value = carverMSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = carverMSNM; 
                Dash.Range('S'+str(moveSpace)).Value = carverMSNF; 
                Dash.Range('U'+str(moveSpace)).Value = carverMSWM; 
                Dash.Range('X'+str(moveSpace)).Value = carverMSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = carverMSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = carverMSPF;
            if("J H HINES" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = hinesElAM; 
                Dash.Range('F'+str(moveSpace)).Value = hinesElAF; 
                Dash.Range('H'+str(moveSpace)).Value = hinesElBM; 
                Dash.Range('J'+str(moveSpace)).Value = hinesElBF; 
                Dash.Range('M'+str(moveSpace)).Value = hinesElHM; 
                Dash.Range('O'+str(moveSpace)).Value = hinesElHF; 
                Dash.Range('Q'+str(moveSpace)).Value = hinesElNM; 
                Dash.Range('S'+str(moveSpace)).Value = hinesElNF; 
                Dash.Range('U'+str(moveSpace)).Value = hinesElWM; 
                Dash.Range('X'+str(moveSpace)).Value = hinesElWF; 
                Dash.Range('Z'+str(moveSpace)).Value = hinesElPM; 
                Dash.Range('AB'+str(moveSpace)).Value = hinesElPF;
            if("KENDRICK" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = kendrickElAM; 
                Dash.Range('F'+str(moveSpace)).Value = kendrickElAF; 
                Dash.Range('H'+str(moveSpace)).Value = kendrickElBM; 
                Dash.Range('J'+str(moveSpace)).Value = kendrickElBF; 
                Dash.Range('M'+str(moveSpace)).Value = kendrickElHM; 
                Dash.Range('O'+str(moveSpace)).Value = kendrickElHF; 
                Dash.Range('Q'+str(moveSpace)).Value = kendrickElNM; 
                Dash.Range('S'+str(moveSpace)).Value = kendrickElNF; 
                Dash.Range('U'+str(moveSpace)).Value = kendrickElWM; 
                Dash.Range('X'+str(moveSpace)).Value = kendrickElWF; 
                Dash.Range('Z'+str(moveSpace)).Value = kendrickElPM; 
                Dash.Range('AB'+str(moveSpace)).Value = kendrickElPF;
            if("PROVIDENT HEIGHTS" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = heightsELAM; 
                Dash.Range('F'+str(moveSpace)).Value = heightsELAF; 
                Dash.Range('H'+str(moveSpace)).Value = heightsELBM; 
                Dash.Range('J'+str(moveSpace)).Value = heightsELBF; 
                Dash.Range('M'+str(moveSpace)).Value = heightsELHM; 
                Dash.Range('O'+str(moveSpace)).Value = heightsELHF; 
                Dash.Range('Q'+str(moveSpace)).Value = heightsELNM; 
                Dash.Range('S'+str(moveSpace)).Value = heightsELNF; 
                Dash.Range('U'+str(moveSpace)).Value = heightsELWM; 
                Dash.Range('X'+str(moveSpace)).Value = heightsELWF; 
                Dash.Range('Z'+str(moveSpace)).Value = heightsELPM; 
                Dash.Range('AB'+str(moveSpace)).Value = heightsELPF;
            if("TENNYSON MIDDLE" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = tennysonHSAM; 
                Dash.Range('F'+str(moveSpace)).Value = tennysonHSAF; 
                Dash.Range('H'+str(moveSpace)).Value = tennysonHSBM; 
                Dash.Range('J'+str(moveSpace)).Value = tennysonHSBF; 
                Dash.Range('M'+str(moveSpace)).Value = tennysonHSHM; 
                Dash.Range('O'+str(moveSpace)).Value = tennysonHSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = tennysonHSNM; 
                Dash.Range('S'+str(moveSpace)).Value = tennysonHSNF; 
                Dash.Range('U'+str(moveSpace)).Value = tennysonHSWM; 
                Dash.Range('X'+str(moveSpace)).Value = tennysonHSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = tennysonHSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = tennysonHSPF;
            if("UNIVERSITY H S" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = universityHSAM; 
                Dash.Range('F'+str(moveSpace)).Value = universityHSAF; 
                Dash.Range('H'+str(moveSpace)).Value = universityHSBM; 
                Dash.Range('J'+str(moveSpace)).Value = universityHSBF; 
                Dash.Range('M'+str(moveSpace)).Value = universityHSHM; 
                Dash.Range('O'+str(moveSpace)).Value = universityHSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = universityHSNM; 
                Dash.Range('S'+str(moveSpace)).Value = universityHSNF; 
                Dash.Range('U'+str(moveSpace)).Value = universityHSWM; 
                Dash.Range('X'+str(moveSpace)).Value = universityHSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = universityHSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = universityHSPF;
            if("WACO H S" in str(Dash.Range('B'+str(moveSpace)).Value)):
                Dash.Range('D'+str(moveSpace)).Value = wacoHSAM; 
                Dash.Range('F'+str(moveSpace)).Value = wacoHSAF; 
                Dash.Range('H'+str(moveSpace)).Value = wacoHSBM; 
                Dash.Range('J'+str(moveSpace)).Value = wacoHSBF; 
                Dash.Range('M'+str(moveSpace)).Value = wacoHSHM; 
                Dash.Range('O'+str(moveSpace)).Value = wacoHSHF; 
                Dash.Range('Q'+str(moveSpace)).Value = wacoHSNM; 
                Dash.Range('S'+str(moveSpace)).Value = wacoHSNF; 
                Dash.Range('U'+str(moveSpace)).Value = wacoHSWM; 
                Dash.Range('X'+str(moveSpace)).Value = wacoHSWF; 
                Dash.Range('Z'+str(moveSpace)).Value = wacoHSPM; 
                Dash.Range('AB'+str(moveSpace)).Value = wacoHSPF;
            moveSpace += 2

        schoolMonth = 1
        columnticker = 1
        rowticker = 1
        dashTicker = 419
        columnmarker = ""
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        campusTicker = columnticker
        valueTransfer = 0

        while(schoolMonth != 2):
            if(schoolMonth == 1):
                if(dashTicker >= 478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in August" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('D'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('D'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 3):
            if(schoolMonth == 2):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in September" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('F'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('F'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;


        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 4):
            if(schoolMonth == 3):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in October" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('H'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('H'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 5):    
            if(schoolMonth == 4):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in November" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('J'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('J'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;
            
        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 6):
            if(schoolMonth == 5):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in December" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('M'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('M'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 7):    
            if(schoolMonth == 6):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in January" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('O'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('O'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 8):    
            if(schoolMonth == 7):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in February" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('Q'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('Q'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 9):
            if(schoolMonth == 8):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in March" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('S'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('S'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 10):
            if(schoolMonth == 9):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in April" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('U'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('U'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 11):
            if(schoolMonth == 10):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in May" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('X'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('X'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 12):
            if(schoolMonth == 11):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in June" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('Z'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('Z'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

        columnticker = 1
        rowticker = 1
        columnmarker = Campus.Cells(rowticker,columnticker).Value
        dashTicker = 419
        while(schoolMonth != 13):
            if(schoolMonth == 12):
                if(dashTicker >=478):
                    schoolMonth += 1
                    break
                if("Number of Students Enrolled in July" in str(columnmarker)):
                    campusTicker = columnticker
                    if(str(Dash.Range('B'+str(dashTicker)).Value) in str(Campus.Range('G'+str(rowticker)).Value)):
                        Dash.Range('AB'+str(dashTicker)).Value = Campus.Cells(rowticker,campusTicker).Value
                        dashTicker += 2
                        rowticker += 1
                    elif((str(Dash.Range('B'+str(dashTicker)).Value) not in str(Campus.Range('G'+str(rowticker)).Value)) and Campus.Range('G'+str(rowticker)).Value != None and Campus.Range('G'+str(rowticker)).Value not in "Campus"):
                        Dash.Range('AB'+str(dashTicker)).Value = 0
                        dashTicker += 2
                    else:
                        if(Campus.Range('G'+str(rowticker)).Value == None):
                            if(Campus.Range('G'+str(rowticker+1)).Value == None):
                                schoolMonth += 1
                                columnticker = 1
                                rowticker = 1
                                columnmarker = Campus.Cells(rowticker,columnticker).Value
                                break
                            rowticker += 1
                        else:
                            rowticker += 1
                elif(columnmarker == None):
                    schoolMonth += 1
                    columnticker = 1
                    rowticker = 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
                    break;
                else:
                    columnticker += 1
                    columnmarker = Campus.Cells(rowticker,columnticker).Value
            else:
                rowticker = 1
                columnticker = 1
                schoolMonth += 1
                break;

            
        target.Close(SaveChanges=True)
        target2.Close(SaveChanges=True)


        programDetails.Status = ""
        programDetails.Status = "Generated Demographic Success Chart..."
        programDetails.Status = ""
        programDetails.Status = "Survey Alignment Report Complete!"
        programDetails.endProgram = True

#GUI

startOff = 0
while(backgroundDetails.previousTab == 0):
    root = mtTkinter.Tk()
    root.lift()
    root.attributes('-topmost', True)
    root.grab_set()
    root.grab_release()
    root.focus_force()
    root.update()
    root.iconbitmap("C:/Users/T Choat/Desktop/Python Code/MST Builder/cis.ico")
    root.title("Survey Alignment Builder")
    root.geometry("525x350")
    root.config(background = "white")
    root.minsize(525,350)
    root.maxsize(525,350)

    pathSelectLabelA = Label(root, text="Select an Assessment Report:",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 9))

    pathSelectLabelM = Label(root, text="Select a Service Mix Report:",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 9))

    pathSelectLabelC = Label(root, text="Select a Campus Report:",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 9))

    pathSelectLabelP = Label(root, text="Select a Progress Report:",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 9))

    pathEntryA = Entry(width = 47)

    pathEntryM = Entry(width = 47)

    pathEntryC = Entry(width = 47)

    pathEntryP = Entry(width = 47)

    if(startOff == 0):
        pathBrowseButtonA = Button(root, text="Browse Folders", command= backgroundDetails.browse_button_One)
    else:
        backgroundDetails.recallPath()
        pathBrowseButtonA = Button(root, text="Browse Folders", command= backgroundDetails.browse_button_One)

    if(startOff == 0):
        pathBrowseButtonM = Button(root, text="Browse Folders", command= backgroundDetails.browse_button_Two)
    else:
        backgroundDetails.recallPath()
        pathBrowseButtonM = Button(root, text="Browse Folders", command= backgroundDetails.browse_button_Two)

    if(startOff == 0):
        pathBrowseButtonC = Button(root, text="Browse Folders", command= backgroundDetails.browse_button_Three)
    else:
        backgroundDetails.recallPath()
        pathBrowseButtonC = Button(root, text="Browse Folders", command= backgroundDetails.browse_button_Three)

    if(startOff == 0):
        pathBrowseButtonP = Button(root, text="Browse Folders", command= backgroundDetails.browse_button_Four)
    else:
        backgroundDetails.recallPath()
        pathBrowseButtonP = Button(root, text="Browse Folders", command= backgroundDetails.browse_button_Four)
        
    nextButton = Button(root, text="Next",height = 2, width = 5, command = backgroundDetails.closeNameWindow)

    root.protocol("WM_DELETE_WINDOW", backgroundDetails.close_window)

    root.bind("<Control-Return>",backgroundDetails.closeNameWindow)
    root.bind("<Control-Escape>",backgroundDetails.close_window)
    root.bind("<Control-a>",backgroundDetails.browse_button_One)
    root.bind("<Control-m>",backgroundDetails.browse_button_Two)
    root.bind("<Control-c>",backgroundDetails.browse_button_Three)
    root.bind("<Control-p>",backgroundDetails.browse_button_Four)

    pathSelectLabelA.place(x=20, y=30)

    pathSelectLabelM.place(x=20, y=90)

    pathSelectLabelC.place(x=20, y=150)

    pathSelectLabelP.place(x=20, y=210)
            
    pathEntryA.place(x=195, y=30)

    pathEntryM.place(x=195, y=90)

    pathEntryC.place(x=195, y=150)

    pathEntryP.place(x=195, y=210)
            
    pathBrowseButtonA.place(x=385, y=55)

    pathBrowseButtonM.place(x=385, y=115)

    pathBrowseButtonC.place(x=385, y=175)

    pathBrowseButtonP.place(x=385, y=235)
          
    nextButton.place(x=450, y = 290)

    root.mainloop()
    while(backgroundDetails.previousTab == 1):
        loading = mtTkinter.Tk()
        loading.lift()
        loading.attributes('-topmost', True)
        loading.grab_set()
        loading.grab_release()
        loading.focus_force()
        loading.update()
        loading.iconbitmap("C:/Users/T Choat/Desktop/Python Code/MST Builder/cis.ico")
        loading.title("Survey Alignment Builder")
        loading.geometry("525x350")
        loading.config(background = "white")
        loading.minsize(525,350)
        loading.maxsize(525,350)

        def run_program(eventRunProgram=None):
            def progressBar():
                previousButton.config(state = "disabled")
                runButton.config(state = "disabled")
                i = 0
                progress_Bar.start()

                for i in range(10000000):
                    if(programDetails.endProgram == False):
                        #progress_Bar.update()
                        statusLabel.config(text = programDetails.Status)
                        loading.after(10)
                        loading.unbind("<Control-Return>")
                        loading.unbind("<Control-BackSpace>")
                        loading.bind("<Control-Escape>",backgroundDetails.close_window)
                    if(programDetails.endProgram == True):
                        statusLabel.config(text = programDetails.Status)
                        #progress_Bar["value"] = 300
                        loading.after(10)
                        #progress_Bar.update()
                        progress_Bar.stop()
                        closeButton.config(state = "active")
                        loading.unbind("<Control-Return>")
                        loading.unbind("<Control-BackSpace>")
                        loading.bind("<Control-Escape>",backgroundDetails.close_window)
                        
            threading.Thread(target=progressBar, daemon = True).start()
            threading.Thread(target=programDetails.program, daemon = True).start()


        progress_Bar = ttk.Progressbar(loading, orient = "horizontal", length = 385, mode = "indeterminate")
        progress_Bar.pack()

        statusLabel = Label(loading, text="Press Run to Continue...",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 10))
        statusLabel.pack()

        def closeButtonEnd():
            progress_Bar.stop()
            loading.destroy
            exit()

        closeButton = Button(loading, text="Close",height = 2, width = 5, command = backgroundDetails.close_window)
        closeButton.pack()
        if(programDetails.endProgram == False):
            progress_Bar.stop()
            closeButton.config(state = "disable")

        previousButton = Button(loading, text="Back",height = 2, width = 5, command = backgroundDetails.previous_Button)
        previousButton.pack()
        
        if(programDetails.endProgram == False):
            progress_Bar.stop()
            previousButton.config(state = "active")
            loading.bind("<Control-Return>",run_program)
            loading.bind("<Control-BackSpace>", backgroundDetails.previous_Button)
            loading.bind("<Control-Escape>",backgroundDetails.close_window)

        runButton = Button(loading, text="Run",height = 2, width = 5, command = run_program)

        loading.protocol("WM_DELETE_WINDOW", backgroundDetails.close_window)

        progress_Bar.place(x=65, y = 75)

        statusLabel.place(x=65, y=100)

        previousButton.place(x=330, y = 290)

        runButton.place(x=390, y = 290)

        closeButton.place(x=450, y = 290)
        
        loading.mainloop()

        loading.update()

    startOff = 1
exit()
