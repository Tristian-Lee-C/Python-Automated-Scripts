#               Installing Pips Guide
#---------------------------------------------------------
#   
#   
#   
#
#   Last Update: 8/29/2022
#   Program Task: This program compiles excel sheets to
#                   generate the MST report.
#   
#
#                   ¯\_(ツ)_/¯
#
#                 Server Version
#---------------------------------------------------------
# Excel-Python Program for MSTs only (Openpyxl, Win32com, Ensurezip, and Pandas)
# Developed by CISHOT Data Entry Team

# Required Information ------
from openpyxl import load_workbook
import openpyxl as opxl
import win32com.client as win32
import pywintypes
import os
from pywintypes import com_error
from win32com.client import constants
import pandas as pd
import numpy as np
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from mttkinter import *
import time
import threading
import random
from datetime import date

class backgroundInfo():
    path = ''
    pathSR = ''
    nameEntrySubmit = ''
    dateEntrySubmit = ''
    previousTab = 0
    mstProgram = 0
    def browse_button_One(eventOne = None):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        pathEntry.delete(0, END)
        pathEntry.insert(END, filename)
        backgroundInfo.path = filename

    def mstGenProgram(eventOne = None):
        backgroundInfo.mstProgram = 1
        backgroundInfo.previousTab = 2
        choiceMenu.destroy()

    def cleaningProgram(eventOne = None):
        backgroundInfo.mstProgram = 2
        backgroundInfo.previousTab = 2
        choiceMenu.destroy()


    def browse_button_Two(eventTwo = None):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        rosterEntry.delete(0, END)
        rosterEntry.insert(END, filename)
        backgroundInfo.pathSR = filename

    def closeNameWindow(eventClose = None):
        backgroundInfo.nameEntrySubmit = nameEntry.get()
        backgroundInfo.dateEntrySubmit = dateEntry.get()
        backgroundInfo.path = pathEntry.get()
        backgroundInfo.pathSR = rosterEntry.get()
        if(backgroundInfo.nameEntrySubmit != "" and backgroundInfo.dateEntrySubmit != "" and backgroundInfo.path != "" and backgroundInfo.pathSR != ""):
            backgroundInfo.previousTab = 1
            root.destroy()

    def recallPath():
        pathEntry.delete(0, END)
        pathEntry.insert(END, backgroundInfo.path)
        rosterEntry.delete(0, END)
        rosterEntry.insert(END, backgroundInfo.pathSR)
        nameEntry.delete(0, END)
        nameEntry.insert(END, backgroundInfo.nameEntrySubmit)
        dateEntry.delete(0, END)
        dateEntry.insert(END, backgroundInfo.dateEntrySubmit)

    def close_window(eventLeave = None):
        exit();

    def previous_Button(eventBack = None):
        if(backgroundInfo.previousTab == 1):
            backgroundInfo.previousTab -= 1
            choiceMenu.destroy()

        if(backgroundInfo.previousTab == 2):
            backgroundInfo.previousTab -= 1
            loading.destroy()

def dispatch(app_name:str):
    try:
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    except AttributeError:
        # Corner case dependencies.
        import os
        import re
        import sys
        import shutil
        # Remove cache and try again.
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        from win32com import client
        app = client.gencache.EnsureDispatch(app_name)
    return app



#Phase 1 --- Clearing and Sorting

#while True:
    #locationOne = input('What year is the file located? (Example. "2022-2023")\n')
    #locationTwo = input('What month is the file located? (Example. "April 2022 - MSTs" or "Jan 2022 - MSTs")\n')
    #checkLocation = input('Files is located in ' + locationOne +' and ' + locationTwo + ' (Y/N)\n')
    #if (checkLocation == "yes" or checkLocation == "Yes" or checkLocation == "YES" or checkLocation == "Y" or checkLocation == "y"):
        #break




        #Start with the Openpyxl to begin
    
#while True:
    #file = input('====>Enter File Name<====\n')
    #filechange = file + '.xlsx'
    #mypath = 'S:/Programs/Community Youth Development/' + locationOne + '/6. Data/MSTs FY 22-23/' + locationTwo + '/'
    #myfile = filechange
    #path = browse_button() #os.path.join(mypath, myfile)
    #file2 = input('====>Enter Student Roster File Name<====\n')
    #filechange2 = file2 + '.xlsx'
    #choiceOne = input("Is your file name " + filechange + " and your Student Roster file name " + file2 + "? \n(Y/N) ---> ")
    #if(choiceOne == "yes" or choiceOne == "Yes" or choiceOne == "YES" or choiceOne == "Y" or choiceOne == "y"):
        #break

        # Deletion of Columns
class programCls():
    Status = ""
    endProgram = False
    def program(eventRun = None):
        if(backgroundInfo.mstProgram == 1):
            year_date = date.today()
            file = os.path.split(backgroundInfo.path)[1]
            file = file.replace(".xlsx","")
            file2 = os.path.split(backgroundInfo.pathSR)[1]
            file2 = file2.replace(".xlsx","")
                
            programCls.Status = "Sorting Documents..."
            xlApp = win32.Dispatch('Excel.Application')
            target = xlApp.Workbooks.Open(backgroundInfo.path)
            Services = target.Worksheets(1)
            try:
                Services.Name = file
            except:
                programCls.endProgram = True
                programCls.Status = "Program Complete!"
                target.Close(SaveChanges=True)
                xlApp.Quit()
                return
            

            rowticker = 1
            endTickerSupport = 60000
            columnticker = 1
            accesskey = Services.Cells(rowticker,columnticker).Value
            
            while(accesskey != None):
                accesskey = Services.Cells(rowticker,columnticker).Value
                if(accesskey == 'NAME_FIRST' or accesskey == 'NAME_LAST' or accesskey == 'SERVICE_DATE' or accesskey == 'GROUP_NAME' or accesskey == 'SERVICE_CODE' or accesskey == 'SERVICE_CODE_DESC' or accesskey == 'PRIMARY_PROVIDER' or accesskey == 'NOTES'):
                    columnticker += 1
                else:
                    localOne = Services.Cells(rowticker,columnticker)
                    localTwo = Services.Cells(endTickerSupport,columnticker)
                    deleteShift = Services.Range(Services.Cells(rowticker,columnticker), Services.Cells(endTickerSupport,columnticker))
                    deleteShift.Delete(Shift=-4159) #Current Project on Deleting


            target.Close(SaveChanges=True)
            xlApp.Quit()

            ws = load_workbook(backgroundInfo.path)
            wb = ws.active

            if (wb['C1'].value != 'FULL_NAME'):
                wb.insert_cols(3);
                wb['C1'] = 'FULL_NAME';
                wb['J1'] = 'In School (I) or Out of School (O)';


        #May be removed due to future Spreadsheet development--

        
            outStringOne = "After-School"
            outStringTwo = "After School"
            outStringThree = "Home Visit"
            outStringFour = "OUT OF SCHOOL"
            outStringFive = "AfterSchool"
            outStringService = "CYD Out of School"
            tick = 2
            fullString = str(wb["I"+str(tick)].value)
            fullStringGroup = str(wb["E"+str(tick)].value)
            fullStringService = str(wb["G"+str(tick)].value)
            while True:
                if(outStringOne.lower() in fullString.lower() or outStringTwo.lower() in fullString.lower() or outStringThree.lower() in fullString.lower() or outStringFour.lower() in fullString.lower() or outStringFive.lower() in fullString.lower()):
                    wb["j"+str(tick)].value = 'O'
                    tick = tick + 1
                elif(outStringService in fullStringService or outStringThree in fullStringService):
                    wb["j"+str(tick)].value = 'O'
                    tick = tick + 1
                else:
                    wb["j"+str(tick)].value = 'I'
                    tick = tick + 1
                if(wb["B"+str(tick)].value == None):
                    break
                fullString = str(wb["I"+str(tick)].value)
                fullStringGroup = str(wb["E"+str(tick)].value)
                fullStringService = str(wb["G"+str(tick)].value)

        #--------------------------------------------------

            ws.create_sheet("Pivot_Table",0)

            ws.save(backgroundInfo.path)
            ws.close()
            
            programCls.Status = ""
            programCls.Status = "Converting Service Codes..."
        #dest = 'S:/Programs/Community Youth Development/' + locationOne + '/6. Data/MSTs FY 22-23/' + locationTwo + '/'
            finalDest = backgroundInfo.path #dest + filechange

            xlApp2 = win32.Dispatch('Excel.Application')
            wbd1 = xlApp2.Workbooks.Open('C:\\Users\\T Choat\\Desktop\\Backup SVR\\OneDrive_2023-11-14 (1)\\MST Report Developer\\config\\CIStoCYD.xlsx')
            wbd2 = xlApp2.Workbooks.Open(finalDest)


                # Copying over the Service Codes
            wsd1 = wbd1.Worksheets(1)
            wsd1.Copy(Before=wbd2.Worksheets(1))



                # Pulling Student Roster
                

        #dest2 = 'S:\\Programs\\Community Youth Development\\MST Report Developer\\Student Roster\\'
            finalDestSR = backgroundInfo.pathSR #dest2 + filechange2

            wbd3 = xlApp2.Workbooks.Open(finalDestSR)

            wsd3 = wbd3.Worksheets(1)
            wsd3.Copy(Before=wbd2.Worksheets(1))


            wbd2.Close(SaveChanges=True)
            xlApp2.Quit()
            del xlApp2


                # Service Code Copy

            wsService = load_workbook(backgroundInfo.path)
            wbService = wsService.active
            log = wsService[file]
            service = wsService["LocalServiceCodes"]
            check = 2
            jump = 2

            while True:
                if(not(log["F"+str(check)].value == 1 or log["F"+str(check)].value == 2 or log["F"+str(check)].value == 3 or log["F"+str(check)].value == 4 or log["F"+str(check)].value == 5 or log["F"+str(check)].value == 6 or log["F"+str(check)].value == 7 or log["F"+str(check)].value == 8 or log["F"+str(check)].value == 9 or log["F"+str(check)].value == 10)):
                    if (log["F"+str(check)].value == service["B"+str(jump)].value):
                        log["F"+str(check)] = service["D"+str(jump)].value
                    if (log["F"+str(check)].value == "none"):
                        log.delete_rows(check,1);
                        check -=1
                    if (log["F"+str(check)].value == None):
                        jump = jump + 1
                        check = 1
                    if (service["B"+str(jump)].value == None):
                        break
                    check = check + 1
                else:
                    None
                    check = check + 1

            check = 2
        #while True:
        #    if (log["F"+str(check)].value == "none"):
        #        break
        #    if(int(log["F"+str(check)].value) > 11):
        #        log.delete_rows(check,1);
        #        check -=1
        #    check = check + 1

            # Creation of Name List

            count = 1
            partA = '=B'
            partB = '&", "&A'
            while True:
                count = count + 1
                if (log['A'+str(count)].value == None and log['A'+str(count)].value == None):
                    break
                log['C'+str(count)] = str(partA+str(count)+partB+str(count))

            wsService.save(backgroundInfo.path)
            wsService.close()


        #Phase 2 --- Pivot Table
            
            def clear_pts(wsw):
                for pt in wsw.PivotTables():
                    pt.TableRange2.Clear()

            def insert_pt_field_set1(pt):
                field_rows = {}
                field_rows['name'] = pt.PivotFields("FULL_NAME")
            
                field_values = {}
                field_values['Service'] = pt.PivotFields("SERVICE_CODE")

                field_columns = {}
                field_columns['ServiceC'] = pt.PivotFields("SERVICE_CODE")

                    # Insert row fields
                    # https://docs.microsoft.com/en-us/office/vba/api/excel.xlpivotfieldorientation
                    
                field_rows['name'].Orientation = 1
                field_rows['name'].Position = 1


                    # Insert values field
                    # https://docs.microsoft.com/en-us/office/vba/api/excel.xlconsolidationfunction

                field_values['Service'].Orientation = 4
                field_values['Service'].Function = -4112 # value reference
                field_values['Service'].NumberFormat = "#,##0"

                field_columns['ServiceC'].Orientation = 2
                field_columns['ServiceC'].Position = 1


                # Rquired to open excel in Win32Com
            xlApp = win32.Dispatch('Excel.Application')
            xlApp.Visible = False

                # Change the Path to Correct Folder if file is changed
        #dest = 'S:/Programs/Community Youth Development/' + locationOne + '/6. Data/MSTs FY 22-23/' + locationTwo + '/' #Where?
            finalDest = backgroundInfo.path#dest + filechange
            
                # Sheet Location and Name
            wbw = xlApp.Workbooks.Open(finalDest)
            wsw_data = wbw.Worksheets(file)
            wsw_pivot = wbw.Worksheets("Pivot_Table")


                # clear pivot tables on Report tab
            clear_pts(wsw_pivot)

                # create pt cache connection
            pt_cache = wbw.PivotCaches().Create(1, wsw_data.Range("A2").CurrentRegion)

                # insert pivot table designer/editor
            pt = pt_cache.CreatePivotTable(wsw_pivot.Range("B3"), "Service")

            insert_pt_field_set1(pt)

            wbw.Close(SaveChanges=True)
            xlApp.Quit()


        #Phase 3 --- MST Formatting

                # Opening Excel with Win32com
            xlApp2 = win32.Dispatch('Excel.Application')
            wbd1 = xlApp2.Workbooks.Open('C:\\Users\\T Choat\\Desktop\\Backup SVR\\OneDrive_2023-11-14 (1)\\MST Report Developer\\config\\Template MST.xlsx')
            wbd2 = xlApp2.Workbooks.Open(finalDest)


                # Copying over the MST
            wsd1 = wbd1.Worksheets(1)
            wsd1.Copy(Before=wbd2.Worksheets(1))

            wbd2.Close(SaveChanges=True)
            wbd1.Close(SaveChanges=True)

            xlApp2.Quit()

        #Phase 4 --- Birth and PEIRS Formatting

                #Line 607 and below hold this task

        #Phase 5 --- Moving Required Fills Over

                # Opening Excel with Openpyxl
            
            ws2 = load_workbook(backgroundInfo.path)
            wb = ws2.active
            PivotToName = ws2["Pivot_Table"]
            mstInfo = ws2["2022 - MST"]
            studSR = ws2[file2]

                # Splitting Names to gain last name
            num = 4
            lastName1 = '=LEFT(B'
            lastName2 = ',FIND(", ",B'
            lastName3 = ')-1)'

            while True:
                num = num + 1
                if (PivotToName['B'+str(num)].value == None or PivotToName['B'+str(num)].value == "Grand Total"):
                    break
                PivotToName['O'+str(num)] = str(lastName1) + str(num) + str(lastName2) + str(num) + str(lastName3)

                #Splitting Names to gain first name

            num2 = 4
            firstName1 = '=RIGHT(B'
            firstName2 = ',LEN(B'
            firstName3 = ')-FIND(", ",B'
            firstName4 = ')-1)'

            while True:
                num2 = num2 + 1
                if (PivotToName['B'+str(num2)].value == None or PivotToName['B'+str(num2)].value == "Grand Total"):
                    break
                PivotToName['P'+str(num2)] = str(firstName1) + str(num2) + str(firstName2) + str(num2) + str(firstName3) + str(num2) + str(firstName4)

                # Asking Questions?
            
            completion = backgroundInfo.nameEntrySubmit
            mstInfo['A10'].value = "Name of Person Completing this Form: " + (completion)
            completion2 = backgroundInfo.dateEntrySubmit
            mstInfo['A13'].value = "Service Month/Year: " + (completion2)

            ws2.save(backgroundInfo.path)

                    # Putting Names into MST  -- Current


            xlApp3 = win32.gencache.EnsureDispatch('Excel.Application')
            #dest2 = 'S:/Programs/Community Youth Development/' + locationOne + '/6. Data/MSTs FY 22-23/' + locationTwo + '/' #Where?
            finalDest2 = backgroundInfo.path #dest2 + filechange
            wbw2 = xlApp3.Workbooks.Open(finalDest2)

            num3 = 5
            altnum3 = 5
            numMST = 25
            altnumMST = 25

            wsw2_pivot = wbw2.Worksheets("Pivot_Table")
            wsw2_mst = wbw2.Worksheets("2022 - MST")
            wsw2_sr = wbw2.Worksheets(file2)
            current_mst_year = str(year_date.year) + " - MST"
            wsw2_mst.Name = current_mst_year
            wsw2_mst = wbw2.Worksheets(current_mst_year)
            while True:
                if (PivotToName['B'+str(num3)].value == None or PivotToName['B'+str(num3)].value == "Grand Total"):
                    break
                num3 = num3 + 1
                altnumMST = altnumMST + 1

            num3-=1
            altnumMST-=1

                    # Code for copying and pasting
                    
            programCls.Status = ""
            programCls.Status = "Inserting Names..."

            loc = "K"+str(altnum3)+":K"+str(num3)
            loc2 ="M"+str(numMST)+":M"+str(altnumMST)
            wsw2_pivot.Range("O"+str(altnum3)+":P"+str(num3)).Copy()
            wsw2_mst.Range("B"+str(numMST)+":C"+str(altnumMST)).PasteSpecial(Paste=-4163)

            programCls.Status = ""
            programCls.Status = "Placing service codes on MST..."
                    # Placing Service Codes
            serviceCheck = 1
            while (serviceCheck < 11):
                if (PivotToName['C4'].value == serviceCheck):
                    wsw2_pivot.Range("C"+str(altnum3)+":C"+str(num3)).Copy()
                    if (serviceCheck == 1):
                        wsw2_mst.Range("G"+str(numMST)+":G"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 2):
                        wsw2_mst.Range("H"+str(numMST)+":h"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 3):
                        wsw2_mst.Range("I"+str(numMST)+":I"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 4):
                        wsw2_mst.Range("J"+str(numMST)+":J"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 5):
                        wsw2_mst.Range("K"+str(numMST)+":K"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 6):
                        wsw2_mst.Range("L"+str(numMST)+":L"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 7):
                        wsw2_mst.Range("M"+str(numMST)+":M"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 8):
                        wsw2_mst.Range("N"+str(numMST)+":N"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 9):
                        wsw2_mst.Range("O"+str(numMST)+":O"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                serviceCheck+=1
            
            serviceCheck = 2
            while (serviceCheck < 11):
                if (PivotToName['D4'].value == serviceCheck):
                    wsw2_pivot.Range("D"+str(altnum3)+":D"+str(num3)).Copy()
                    if (serviceCheck == 2):
                        wsw2_mst.Range("H"+str(numMST)+":H"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 3):
                        wsw2_mst.Range("I"+str(numMST)+":I"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 4):
                        wsw2_mst.Range("J"+str(numMST)+":J"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 5):
                        wsw2_mst.Range("K"+str(numMST)+":K"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 6):
                        wsw2_mst.Range("L"+str(numMST)+":L"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 7):
                        wsw2_mst.Range("M"+str(numMST)+":M"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 8):
                        wsw2_mst.Range("N"+str(numMST)+":N"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 9):
                        wsw2_mst.Range("O"+str(numMST)+":O"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                serviceCheck+=1

            serviceCheck = 3
            while (serviceCheck < 11):
                if (PivotToName['E4'].value == serviceCheck):
                    wsw2_pivot.Range("E"+str(altnum3)+":E"+str(num3)).Copy()
                    if (serviceCheck == 3):
                        wsw2_mst.Range("I"+str(numMST)+":I"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 4):
                        wsw2_mst.Range("J"+str(numMST)+":J"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 5):
                        wsw2_mst.Range("K"+str(numMST)+":K"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 6):
                        wsw2_mst.Range("L"+str(numMST)+":L"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 7):
                        wsw2_mst.Range("M"+str(numMST)+":M"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 8):
                        wsw2_mst.Range("N"+str(numMST)+":N"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 9):
                        wsw2_mst.Range("O"+str(numMST)+":O"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                serviceCheck+=1

            serviceCheck = 4
            while (serviceCheck < 11):
                if (PivotToName['F4'].value == serviceCheck):
                    wsw2_pivot.Range("F"+str(altnum3)+":F"+str(num3)).Copy()
                    if (serviceCheck == 4):
                        wsw2_mst.Range("J"+str(numMST)+":J"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 5):
                        wsw2_mst.Range("K"+str(numMST)+":K"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 6):
                        wsw2_mst.Range("L"+str(numMST)+":L"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 7):
                        wsw2_mst.Range("M"+str(numMST)+":M"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 8):
                        wsw2_mst.Range("N"+str(numMST)+":N"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 9):
                        wsw2_mst.Range("O"+str(numMST)+":O"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                serviceCheck+=1

            serviceCheck = 5
            while (serviceCheck < 11):
                if (PivotToName['G4'].value == serviceCheck):
                    wsw2_pivot.Range("G"+str(altnum3)+":G"+str(num3)).Copy()
                    if (serviceCheck == 5):
                        wsw2_mst.Range("K"+str(numMST)+":k"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 6):
                        wsw2_mst.Range("L"+str(numMST)+":L"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 7):
                        wsw2_mst.Range("M"+str(numMST)+":M"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 8):
                        wsw2_mst.Range("N"+str(numMST)+":N"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 9):
                        wsw2_mst.Range("O"+str(numMST)+":O"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                serviceCheck+=1

            serviceCheck = 6
            while (serviceCheck < 11):
                if (PivotToName['H4'].value == serviceCheck):
                    wsw2_pivot.Range("H"+str(altnum3)+":H"+str(num3)).Copy()
                    if (serviceCheck == 6):
                        wsw2_mst.Range("L"+str(numMST)+":L"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 7):
                        wsw2_mst.Range("M"+str(numMST)+":M"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 8):
                        wsw2_mst.Range("N"+str(numMST)+":N"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 9):
                        wsw2_mst.Range("O"+str(numMST)+":O"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                serviceCheck+=1

            serviceCheck = 7
            while (serviceCheck < 11):
                if (PivotToName['I4'].value == serviceCheck):
                    wsw2_pivot.Range("I"+str(altnum3)+":I"+str(num3)).Copy()
                    if (serviceCheck == 7):
                        wsw2_mst.Range("M"+str(numMST)+":M"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 8):
                        wsw2_mst.Range("N"+str(numMST)+":N"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 9):
                        wsw2_mst.Range("O"+str(numMST)+":O"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                serviceCheck+=1

            serviceCheck = 8
            while (serviceCheck < 11):
                if (PivotToName['J4'].value == serviceCheck):
                    wsw2_pivot.Range("J"+str(altnum3)+":J"+str(num3)).Copy()
                    if (serviceCheck == 8):
                        wsw2_mst.Range("N"+str(numMST)+":N"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 9):
                        wsw2_mst.Range("O"+str(numMST)+":O"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                serviceCheck+=1

            serviceCheck = 9
            while (serviceCheck < 11):
                if (PivotToName['K4'].value == serviceCheck):
                    wsw2_pivot.Range("K"+str(altnum3)+":K"+str(num3)).Copy()
                    if (serviceCheck == 9):
                        wsw2_mst.Range("O"+str(numMST)+":O"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                        break
                serviceCheck+=1

            serviceCheck = 10
            while (serviceCheck < 11):
                if (PivotToName['L4'].value == serviceCheck):
                    wsw2_pivot.Range("L"+str(altnum3)+":L"+str(num3)).Copy()
                    if (serviceCheck == 10):
                        wsw2_mst.Range("P"+str(numMST)+":P"+str(altnumMST)).PasteSpecial(Paste=-4163)
                break

                    # Deleting Extra Space
            wsw2_mst.Range("A"+str(altnumMST+1)+":Q5377").Delete(Shift=-4162)

            wbw2.Close(SaveChanges=True)
            ws.close()
            xlApp3.Quit()
            del xlApp3

            programCls.Status = ""
            programCls.Status = "Inserting B-Days and Ids..."

            xlApp5 = win32.gencache.EnsureDispatch('Excel.Application')
            #dest3 = 'S:/Programs/Community Youth Development/' + locationOne + '/6. Data/MSTs FY 22-23/' + locationTwo + '/' #Where?
            finalDest3 = backgroundInfo.path #dest3 + filechange
            wbw5 = xlApp5.Workbooks.Open(finalDest2)

            wsw5_mst = wbw5.Worksheets(current_mst_year)
            wsw5_sr = wbw5.Worksheets(file2)
            numSR = 2
            numMSTSR = 25
            while True:
                if(wsw5_sr.Range('B'+str(numSR)).Value == None):
                    wsw5_mst.Range('A'+str(numMSTSR)).Interior.ColorIndex = 6
                    wsw5_mst.Range('A'+str(numMSTSR)).Interior.ColorIndex = 6
                    numMSTSR = numMSTSR + 1
                    numSR = 1
                if (numMSTSR == altnumMST+1):
                    break
                if(wsw5_mst.Range('B'+str(numMSTSR)).Value.lower() in wsw5_sr.Range('B'+str(numSR)).Value.lower()):
                    if(wsw5_mst.Range('C'+str(numMSTSR)).Value.lower() in wsw5_sr.Range('C'+str(numSR)).Value.lower()):
                        wsw5_sr.Range('A'+str(numSR)).Copy();
                        time.sleep(.5)
                        print(str(wsw5_sr.Range('A'+str(numSR)).Value))
                        wsw5_mst.Range('A'+str(numMSTSR)).PasteSpecial(Paste=-4163);
                        wsw5_sr.Range('E'+str(numSR)).Copy();
                        time.sleep(.5)
                        wsw5_mst.Range('E'+str(numMSTSR)).PasteSpecial(Paste=-4163);
                        numMSTSR = numMSTSR + 1;
                        numSR = 1;
                numSR = numSR+1

            wbw5.Close(SaveChanges=True)
            xlApp5.Quit()
            del xlApp5
            programCls.endProgram = True
            programCls.Status = ""
            programCls.Status = "Program Complete!"

        if(backgroundInfo.mstProgram == 2):
            programCls.Status = ""
            programCls.Status = "Creating Clean MST"
            file = os.path.split(backgroundInfo.path)[1]
            file = file.replace(".xlsx","")
            finalDest = backgroundInfo.path
            xlApp2 = win32.gencache.EnsureDispatch('Excel.Application')
            wbd3 = xlApp2.Workbooks.Open(finalDest)

            wsd3 = wbd3.Worksheets(1)
            wsd3.Copy(Before=wbd3.Worksheets(1))

            wbd = wbd3.Worksheets(1)
            wbd.Name = 'Clean - MST'

            '''wbd2 = wbd3.Worksheets(file)
            wbd2.Copy(Before=wbd3.Worksheets(2))

            wbd4 = wbd3.Worksheets(file+' (2)')
            wbd4.Name = 'Clean ' + file


            MST = wbd3.Worksheets('Clean - MST')
            Report = wbd3.Worksheets('Clean '+ file)
            cleanMST = 25
            cleanReport = 2
            cleanNameLast = MST.Range('B'+str(cleanMST)).Value
            cleanNameFirst = MST.Range('C'+str(cleanMST)).Value
            out = False
            while(True):
                while True:
                    if (Report.Range('B'+str(cleanReport)).Value == None and Report.Range('A'+str(cleanReport)).Value == None):
                        cleanReport = 2
                        break
                    if (Report.Range('B'+str(cleanReport)).Value == cleanNameLast and Report.Range('A'+str(cleanReport)).Value == cleanNameFirst):
                        Report.Range("A"+str(cleanReport)+":I"+str(cleanReport)).Delete(Shift=-4162)
                        cleanReport = cleanReport - 1
                    cleanReport = cleanReport + 1
                if (MST.Range('B'+str(cleanMST)).Value == None):
                    break
                if (MST.Range('A'+str(cleanMST)).Value == None):
                    cleanNameLast = MST.Range('B'+str(cleanMST)).Value
                    cleanNameFirst = MST.Range('C'+str(cleanMST)).Value
                    MST.Range("A"+str(cleanMST)+":Q"+str(cleanMST)).Delete(Shift=-4162)
                    cleanMST = cleanMST - 1
                cleanMST = cleanMST + 1
'''
            wbd3.Close(SaveChanges=True)
            xlApp2.Quit()
            del xlApp2
            programCls.Status = ""
            programCls.Status = "Creating Clean Service Log"
            wsService = load_workbook(backgroundInfo.path)
            wbService = wsService.active
            log = wsService[file]
            wsService.create_sheet("Clean Service Log")
            cleanService = wsService["Clean Service Log"]
            cleanMST = wsService['Clean - MST']

            CMManager = []
            CMStudent = []
            CMService = []
            CMCatch= []
            CMAmount = []
            tick = 2
            while True:
                if(str(log["H"+str(tick)].value) not in CMManager):
                    CMManager.append(str(log["H"+str(tick)].value))
                if(str(log["A"+str(tick)].value) not in CMStudent and str(log["B"+str(tick)].value) not in CMStudent):
                    CMStudent.append(str(log["A"+str(tick)].value) +", "+str(log["B"+str(tick)].value))
                if((log["H"+str(tick)].value) == None):
                    break
                tick += 1
            wsService.save(backgroundInfo.path)
            programCls.Status = ""
            programCls.Status = "Sorting Students"
            tick = 2
            slowtick=2
            checkTick = 25
            CMPlacement = 0
            tempCMService = CMService
            CMLength = len(CMManager)
            while True:
                if(CMPlacement >= CMLength):
                    break
                studentName = str(log["A"+str(tick)].value) +", "+str(log["B"+str(tick)].value)
                if(str(log["H"+str(tick)].value) in CMManager[CMPlacement] and (studentName in CMStudent and str(log["F"+str(tick)].value) not in CMService)):
                    comboLine = str(log["H"+str(tick)].value) + " - " + studentName + " - " + str(log["F"+str(tick)].value)
                    if(comboLine not in CMCatch):
                        CMCatch.append(comboLine)
                        checkTick = 25
                        breakoff = True
                        while(breakoff != False):
                            if(str(cleanMST["B"+str(checkTick)].value) in str(log["B"+str(tick)].value) and str(cleanMST["C"+str(checkTick)].value) in str(log["A"+str(tick)].value)):
                                if(str(cleanMST["A"+str(checkTick)].value) != 'None'):
                                    cleanService["A"+str(slowtick)] = str(log["A"+str(tick)].value)
                                    cleanService["B"+str(slowtick)] = str(log["B"+str(tick)].value)
                                    cleanService["C"+str(slowtick)] = str(cleanService["A"+str(slowtick)].value) +", "+str(cleanService["B"+str(slowtick)].value)
                                    cleanService["D"+str(slowtick)] = str(log["D"+str(tick)].value)
                                    cleanService["E"+str(slowtick)] = str(log["E"+str(tick)].value)
                                    cleanService["F"+str(slowtick)] = str(log["F"+str(tick)].value)
                                    cleanService["G"+str(slowtick)] = str(log["G"+str(tick)].value)
                                    cleanService["H"+str(slowtick)] = str(log["H"+str(tick)].value)
                                    cleanService["I"+str(slowtick)] = str(log["I"+str(tick)].value)
                                    cleanService["J"+str(slowtick)] = str(log["J"+str(tick)].value)
                                    breakoff = False
                                else:
                                    checkTick -= 1 
                                    breakoff = False
                                    slowtick -= 1
                            elif((cleanMST["B"+str(checkTick)].value) == None):
                                breakoff = False
                            else:
                                checkTick += 1 
                        slowtick += 1
                if((log["H"+str(tick)].value) == None):
                    CMPlacement += 1
                    tick = 1
                tick += 1
        
            wsService.save(backgroundInfo.path)
            wsService.close()

            xlApp2 = win32.gencache.EnsureDispatch('Excel.Application')
            wbDelete = xlApp2.Workbooks.Open(backgroundInfo.path)
            mstDelete = wbDelete.Worksheets('Clean - MST')
            cleanServiceLog = wbDelete.Worksheets('Clean Service Log')
            mainServiceLog = wbDelete.Worksheets(file)
            clearNull = 25
            while((mstDelete.Range("B"+str(clearNull)).Value) != None):
                if((mstDelete.Range("A"+str(clearNull)).Value) == None):
                    mstDelete.Range("A"+str(clearNull)+":R"+str(clearNull)).Delete(Shift=-4162)
                    clearNull -= 1
                clearNull += 1

            mainServiceLog.Range("A1:J1").Copy()
            cleanServiceLog.Range('A1').PasteSpecial(Paste=-4163);
            wbDelete.Close(SaveChanges=True)
            xlApp2.Quit()

            programCls.endProgram = True
            programCls.Status = ""
            programCls.Status = "Program Complete!"

startOff = 0
randomNum = random.random()
probability = int((randomNum*100)%2)

while(backgroundInfo.previousTab == 0):
    root = mtTkinter.Tk()
    root.lift()
    root.attributes('-topmost', True)
    root.grab_set()
    root.grab_release()
    root.focus_force()
    root.update()
    root.iconbitmap("C:/Users/T Choat/Desktop/Python Code/MST Builder/cis.ico")
    root.title("MST Builder")
    root.geometry("525x350")
    root.config(background = "white")
    root.minsize(525,350)
    root.maxsize(525,350)

    pathSelectLabel = Label(root, text="Select a MST File:",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 10))
            
    rosterSelectLabel = Label(root, text="Select a Roster File:",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 10))

    pathEntry = Entry(width = 53)

    rosterEntry = Entry(width = 53)

    nameEntryLabel = Label(root, text="Employee First and Last Name",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 10))

    dateEntryLabel = Label(root, text="MST Month and Year",
                                        bg = "white",
                                        fg = "black",
                                        font = ("Arial", 10))

    nameEntry = Entry(root, text = "Employee First and Last Name", width = 53)

    dateEntry = Entry(root, text = "MST Month and Year", width = 53)

    if(startOff == 0):
        pathBrowseButton = Button(root, text="Browse Folders", command= backgroundInfo.browse_button_One)
        rosterBrowseButton = Button(root, text="Browse Folders", command= backgroundInfo.browse_button_Two)
    else:
        backgroundInfo.recallPath()
        pathBrowseButton = Button(root, text="Browse Folders", command= backgroundInfo.browse_button_One)
        rosterBrowseButton = Button(root, text="Browse Folders", command= backgroundInfo.browse_button_Two)
            
    nextButton = Button(root, text="Next",height = 2, width = 5, command = backgroundInfo.closeNameWindow)

    root.protocol("WM_DELETE_WINDOW", backgroundInfo.close_window)

    root.bind("<Control-Return>",backgroundInfo.closeNameWindow)
    root.bind("<Control-Escape>",backgroundInfo.close_window)
    root.bind("<Control-m>",backgroundInfo.browse_button_One)
    root.bind("<Control-r>",backgroundInfo.browse_button_Two)

    pathSelectLabel.place(x=20, y=30)
            
    pathEntry.place(x=155, y=30)
            
    pathBrowseButton.place(x=385, y=60)

    rosterSelectLabel.place(x=20, y=110)

    rosterEntry.place(x=155, y=110)
            
    rosterBrowseButton.place(x=385, y=140)

    nameEntryLabel.place(x=20, y=200)

    nameEntry.place(x=20, y=220)

    dateEntryLabel.place(x=20, y=260)

    dateEntry.place(x=20, y=280)
            
    nextButton.place(x=450, y = 290)

    root.mainloop()
    while(backgroundInfo.previousTab == 1):
        choiceMenu = mtTkinter.Tk()
        choiceMenu.lift()
        choiceMenu.attributes('-topmost', True)
        choiceMenu.grab_set()
        choiceMenu.grab_release()
        choiceMenu.focus_force()
        choiceMenu.update()
        choiceMenu.iconbitmap("C:/Users/T Choat/Desktop/Python Code/MST Builder/cis.ico")
        choiceMenu.title("MST Builder")
        choiceMenu.geometry("525x350")
        choiceMenu.config(background = "white")
        choiceMenu.minsize(525,350)
        choiceMenu.maxsize(525,350)

        menuDialog = Label(choiceMenu, text="Choose the action you would like to automate!",
                                        bg = "white",
                                        fg = "black",
                                        justify = CENTER,
                                        font = ("Arial", 10))
        menuDialog.pack()
        menuDialog.place(x=125, y = 100)

        mstButton = Button(choiceMenu, text="MST",height = 2, width = 7, command = backgroundInfo.mstGenProgram)
        mstButton.pack()
        mstButton.place(x=175, y = 150)

        cleanButton = Button(choiceMenu, text="Service Log",height = 2, width = 8, command = backgroundInfo.cleaningProgram)
        cleanButton.pack()
        cleanButton.place(x=270, y = 150)

        previousButton = Button(choiceMenu, text="Back",height = 2, width = 5, command = backgroundInfo.previous_Button)
        previousButton.pack()
        previousButton.place(x=460, y = 290)

        choiceMenu.protocol("WM_DELETE_WINDOW", backgroundInfo.close_window)

        choiceMenu.mainloop()
        while(backgroundInfo.previousTab == 2):
            loading = mtTkinter.Tk()
            loading.lift()
            loading.attributes('-topmost', True)
            loading.grab_set()
            loading.grab_release()
            loading.focus_force()
            loading.update()
            loading.iconbitmap("C:/Users/T Choat/Desktop/Python Code/MST Builder/cis.ico")
            loading.title("MST Builder")
            loading.geometry("525x350")
            loading.config(background = "white")
            loading.minsize(525,350)
            loading.maxsize(525,350)

            def run_program(eventExecute = None):
                def progressBar():
                    previousButton.config(state = "disabled")
                    runButton.config(state = "disabled")
                    i = 0
                    progress_Bar.start()

                    
                    for i in range(600):
                        progress_Bar.update()
                        statusLabel.config(text = programCls.Status)
                        loading.after(900)
                        if(programCls.endProgram == True):
                            statusLabel.config(text = programCls.Status)
                            progress_Bar["value"] = 600
                            loading.after(10)
                            progress_Bar.update()
                            closeButton.config(state = "active")
                            loading.unbind("<Control-Return>")
                            loading.unbind("<Control-BackSpace>")
                            loading.bind("<Control-Escape>",backgroundInfo.close_window)
                            
                threading.Thread(target=progressBar, daemon = True).start()
                threading.Thread(target=programCls.program, daemon = True).start()


            progress_Bar = ttk.Progressbar(loading, orient = "horizontal", length=400, maximum = 350, mode = "determinate")
            progress_Bar.pack()

            statusLabel = Label(loading, text="Press Run to Continue...",
                                            bg = "white",
                                            fg = "black",
                                            font = ("Arial", 10))
            statusLabel.pack()

            def closeButtonEnd():
                progress_Bar.stop()
                loading.destroy
                exit()

            closeButton = Button(loading, text="Close",height = 2, width = 5, command = backgroundInfo.close_window)
            closeButton.pack()
            if(programCls.endProgram == False):
                progress_Bar.stop()
                closeButton.config(state = "disable")

            previousButton = Button(loading, text="Back",height = 2, width = 5, command = backgroundInfo.previous_Button)
            previousButton.pack()
            
            if(programCls.endProgram == False):
                progress_Bar.stop()
                previousButton.config(state = "active")
                loading.bind("<Control-Return>",run_program)
                loading.bind("<Control-BackSpace>", backgroundInfo.previous_Button)
                loading.bind("<Control-Escape>",backgroundInfo.close_window)

            runButton = Button(loading, text="Run",height = 2, width = 5, command = run_program)

            tipLabelOne = Label(loading, text="Tip: Remember to update the Student \nRosters and Service Code Excel Sheets",
                                            bg = "white",
                                            fg = "black",
                                            justify = LEFT,
                                            font = ("Arial", 9))
            tipLabelOne.pack()

            tipLabelTwo = Label(loading, text="Tip: If the pivot table contains a service code \ngreater than 10. The Service Code Sheet isn't updated!",
                                            bg = "white",
                                            fg = "black",
                                            justify = LEFT,
                                            font = ("Arial", 9))
            tipLabelTwo.pack()

            loading.protocol("WM_DELETE_WINDOW", backgroundInfo.close_window)

            progress_Bar.place(x=65, y = 75)

            statusLabel.place(x=65, y=100)

            previousButton.place(x=330, y = 290)

            runButton.place(x=390, y = 290)

            closeButton.place(x=450, y = 290)
            
            if(probability == 1):
                tipLabelTwo.pack_forget()
                tipLabelOne.place(x = 15, y = 295)
            else:
                tipLabelOne.pack_forget()
                tipLabelTwo.place(x = 15, y = 295)
            
            loading.mainloop()

            loading.update()

        startOff = 1
exit()
